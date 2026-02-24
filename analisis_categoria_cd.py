#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar informe semanal de artículos de categoría C y D
que deberían eliminarse del stock pero todavía están presentes.

Este script:
1. Lee los archivos de clasificación ABC+D de cada sección
2. Compara con el archivo de stock actual
3. Identifica artículos de categoría C y D que todavía están en stock
4. Genera un archivo Excel con el análisis por sección
"""

import sys
import os
import re
from pathlib import Path

# Añadir el directorio del script al path para poder importar src
_script_dir = Path(__file__).resolve().parent
if str(_script_dir) not in sys.path:
    sys.path.insert(0, str(_script_dir))

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import glob
import json
import smtplib
import ssl
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import formatdate

# Importar rutas centralizadas
from src.paths import INPUT_DIR, OUTPUT_DIR, CONFIG_DIR, ARCHIVO_STOCK_ACTUAL, PATRON_CLASIFICACION_ABC, ANALISIS_CATEGORIA_CD_DIR
from src.date_utils import get_periodo_y_año_dinamico, get_periodo_info_detallada

# Color de encabezado RGB[0,128,0] (verde)
HEADER_COLOR = "FF008000"  # Verde en formato hex para openpyxl

# Definición de secciones
SECCIONES = [
    "MAF",
    "TIERRA_ARIDOS", 
    "DECO_EXTERIOR",
    "DECO_INTERIOR",
    "FITOS",
    "INTERIOR",
    "MASCOTAS_MANUFACTURADO",
    "MASCOTAS_VIVO",
    "SEMILLAS",
    "UTILES_JARDIN",
    "VIVERO"
]

# Escenarios que corresponden a categoría C y D
# Basado en el análisis: Escenario 3 y 7 tienen riesgo Medio y acciones preventivas
CATEGORIA_C_D_ESCENARIOS = ["3", "7"]

# ============================================================================
# CONFIGURACIÓN DE EMAIL
# ============================================================================

# Destinatarios: Ivan y Sandra
DESTINATARIOS = [
    {'nombre': 'Ivan', 'email': 'ivan.delgado@viveverde.es'},
    {'nombre': 'Sandra', 'email': 'ivan.delgado@viveverde.es'}
]

# Configuración del servidor SMTP
SMTP_CONFIG = {
    'servidor': 'smtp.serviciodecorreo.es',
    'puerto': 465,
    'remitente_email': 'ivan.delgado@viveverde.es',
    'remitente_nombre': 'Sistema de Pedidos Viveverde'
}


def cargar_configuracion():
    """Carga la configuración del sistema."""
    config_path = CONFIG_DIR / "config.json"
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def obtener_archivo_clasificacion(seccion):
    """
    Obtiene el archivo de clasificación ABC+D para el período actual.
    
    Utiliza la lógica dinámica para determinar el período y año:
    - Período: actual (basado en la fecha del sistema)
    - Año: anterior (año anterior al actual)
    """
    # Obtener período y año dinámicamente (período actual, año anterior)
    datos_dinamicos = get_periodo_y_año_dinamico(tipo_calculo="actual")
    periodo = datos_dinamicos['periodo']
    año = datos_dinamicos['año']
    
    # Mostrar información de cálculo automático (solo una vez)
    if seccion == SECCIONES[0]:  # Solo mostrar en la primera sección
        info_periodo = get_periodo_info_detallada()
        print(f"\n*** MODO AUTOMÁTICO: Determinando período y año dinámicamente ***")
        print(f"  Fecha actual: {info_periodo['fecha_actual']} (semana {info_periodo['semana_actual']})")
        print(f"  Período actual del sistema: {info_periodo['periodo_actual']}")
        print(f"  Período a procesar: {periodo}")
        print(f"  Año a procesar: {año}")
        print("*" * 60)
    
    # Buscar archivos que coincidan con el patrón específico (período y año)
    patrones = [
        INPUT_DIR / f"CLASIFICACION_ABC+D_{seccion}_{periodo}_{año}.xlsx",
        INPUT_DIR / f"1CLASIFICACION_ABC+D_{seccion}_{periodo}_{año}.xlsx"
    ]
    
    archivos = []
    for patron in patrones:
        archivos.extend(glob.glob(str(patron)))
    
    if not archivos:
        print(f"  ⚠️ No se encontró archivo de clasificación para {seccion} ({periodo}_{año})")
        return None
    
    # Devolver el primer archivo encontrado (debería haber solo uno con el filtro específico)
    return archivos[0]


def normalizar_codigo_articulo(codigo):
    """
    Normaliza el código de artículo para comparación.
    Convierte a string y elimina decimales.
    """
    if pd.isna(codigo):
        return None
    
    # Convertir a string
    codigo_str = str(codigo).strip()
    
    # Eliminar decimales si existen (por ejemplo, "101010001.0" -> "101010001")
    if '.' in codigo_str:
        try:
            codigo_str = str(int(float(codigo_str)))
        except:
            pass
    
    return codigo_str


def cargar_stock_actual():
    """
    Carga el archivo de stock actual.
    Llena las celdas vacías de artículo con el valor de la celda superior.
    """
    stock_path = ARCHIVO_STOCK_ACTUAL
    if not stock_path.exists():
        raise FileNotFoundError(f"No se encontró el archivo de stock: {stock_path}")
    
    df = pd.read_excel(stock_path)
    
    # Rellenar celdas vacías de artículo con el valor de la celda superior
    # Esto es necesario porque en el Excel los artículos se agrupan y solo 
    # el primero de cada grupo tiene el código
    df['Artículo'] = df['Artículo'].ffill()
    
    # Normalizar código de artículo a string para comparación
    df['Artículo'] = df['Artículo'].apply(normalizar_codigo_articulo)
    # Eliminar filas con código nulo
    df = df[df['Artículo'].notna()]
    return df


def cargar_clasificacion(seccion):
    """
    Carga el archivo de clasificación para una sección específica.
    """
    archivo = obtener_archivo_clasificacion(seccion)
    if archivo is None:
        return None
    
    df = pd.read_excel(archivo)
    # Normalizar código de artículo
    df['Artículo'] = df['Artículo'].apply(normalizar_codigo_articulo)
    
    return df


def identificar_articulos_categoria_c_d(df_clasificacion):
    """
    Identifica los artículos de categoría C y D basados en los escenarios.
    """
    if df_clasificacion is None or df_clasificacion.empty:
        return pd.DataFrame()
    
    # Filtrar por escenarios de categoría C y D
    df_categoria_cd = df_clasificacion[
        df_clasificacion['Escenario'].isin(CATEGORIA_C_D_ESCENARIOS)
    ].copy()
    
    return df_categoria_cd


def comparar_con_stock(df_categoria_cd, df_stock):
    """
    Compara los artículos de categoría C y D con el stock actual.
    Devuelve los artículos que todavía están en stock.
    """
    if df_categoria_cd.empty:
        return pd.DataFrame()
    
    # Obtener códigos de artículos en stock
    articulos_en_stock = set(df_stock['Artículo'].unique())
    
    # Filtrar artículos que todavía están en stock
    df_en_stock = df_categoria_cd[
        df_categoria_cd['Artículo'].isin(articulos_en_stock)
    ].copy()
    
    return df_en_stock


def obtener_unidades_stock(articulo, talla, color, df_stock):
    """
    Obtiene las unidades en stock para un artículo específico.
    """
    if pd.isna(talla):
        talla = None
    if pd.isna(color):
        color = None
    
    # Buscar en stock
    mask = df_stock['Artículo'] == articulo
    
    if talla is not None:
        mask = mask & (df_stock['Talla'].astype(str).str.strip() == str(talla).strip())
    if color is not None:
        mask = mask & (df_stock['Color'].astype(str).str.strip() == str(color).strip())
    
    unidades = df_stock.loc[mask, 'Unidades'].sum()
    return unidades if pd.notna(unidades) else 0


def calcular_metricas(df_en_stock, df_stock, seccion):
    """
    Calcula métricas de resumen para la sección.
    """
    if df_en_stock.empty:
        return {
            'seccion': seccion,
            'total_articulos_cd': 0,
            'total_articulos_en_stock': 0,
            'unidades_totales_en_stock': 0,
            'articulos_sin_stock': 0,
            'porcentaje_sin_eliminar': 0
        }
    
    # Total de artículos de C y D identificados
    total_cd = len(df_en_stock)
    
    # Unidades totales en stock
    # Para cada artículo, sumar sus unidades en stock
    unidades_totales = 0
    for _, row in df_en_stock.iterrows():
        unidades = obtener_unidades_stock(
            row['Artículo'], 
            row.get('Talla'), 
            row.get('Color'), 
            df_stock
        )
        unidades_totales += unidades
    
    # Calcular métricas
    metricas = {
        'seccion': seccion,
        'total_articulos_cd': total_cd,
        'total_articulos_en_stock': total_cd,  # Todos los que encontramos están en stock
        'unidades_totales_en_stock': unidades_totales,
        'articulos_sin_stock': 0,  # Los que deberíamos haber eliminado pero ya no están
        'porcentaje_sin_eliminar': 100.0 if total_cd > 0 else 0
    }
    
    return metricas


def crear_excel(df_en_stock, metricas, seccion, workbook):
    """
    Crea una hoja en el workbook con los datos de la sección.
    """
    ws = workbook.create_sheet(title=seccion)
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FF008000", end_color="FF008000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título de la sección
    ws['A1'] = f"ARTÍCULOS DE CATEGORÍA C Y D PENDIENTES DE ELIMINAR - {seccion}"
    ws['A1'].font = Font(bold=True, size=14, color="FF008000")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Fecha de generación
    ws['A2'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    
    # Encabezados de la tabla
    headers = ['Artículo', 'Nombre artículo', 'Talla', 'Color', 'unidades']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos de la tabla
    if not df_en_stock.empty:
        # Obtener unidades en stock para cada artículo
        datos_tabla = []
        for _, row in df_en_stock.iterrows():
            articulo = row['Artículo']
            nombre = row.get('Nombre artículo', '')
            talla = row.get('Talla', '')
            color = row.get('Color', '')
            
            # Obtener unidades en stock
            unidades = obtener_unidades_stock(articulo, talla, color, 
                                             df_stock_global[seccion])
            
            datos_tabla.append({
                'Artículo': articulo,
                'Nombre artículo': nombre,
                'Talla': '' if pd.isna(talla) else str(talla),
                'Color': '' if pd.isna(color) else str(color),
                'unidades': unidades
            })
        
        # Escribir datos
        for row_idx, datos in enumerate(datos_tabla, start=5):
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=datos[header])
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
    else:
        ws['A6'] = "No hay artículos de categoría C y D pendientes de eliminar"
        ws['A6'].font = Font(italic=True, color="808080")
        ws.merge_cells('A6:E6')
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    
    # Sección de Métricas de Resumen
    fila_metricas = 6 + len(df_en_stock) + 2
    
    ws[f'A{fila_metricas}'] = "MÉTRICAS DE RESUMEN"
    ws[f'A{fila_metricas}'].font = Font(bold=True, size=12, color="FF008000")
    ws.merge_cells(f'A{fila_metricas}:E{fila_metricas}')
    
    # Encabezados de métricas
    fila_metricas += 1
    metricas_headers = ['Métrica', 'Valor']
    for col_idx, header in enumerate(metricas_headers, start=1):
        cell = ws.cell(row=fila_metricas, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos de métricas
    metricas_data = [
        ('Total artículos C+D identificados', metricas['total_articulos_cd']),
        ('Artículos todavía en stock', metricas['total_articulos_en_stock']),
        ('Unidades totales en stock', metricas['unidades_totales_en_stock']),
        ('Artículos ya eliminados del stock', metricas['articulos_sin_stock']),
        ('Porcentaje sin eliminar (%)', f"{metricas['porcentaje_sin_eliminar']:.1f}%")
    ]
    
    for row_idx, (metrica, valor) in enumerate(metricas_data, start=fila_metricas + 1):
        ws.cell(row=row_idx, column=1, value=metrica).border = thin_border
        ws.cell(row=row_idx, column=2, value=valor).border = thin_border
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="center")
    
    # Añadir nota explicativa
    fila_nota = fila_metricas + len(metricas_data) + 2
    ws[f'A{fila_nota}'] = "Nota: Estos artículos deberían haber sido eliminados del stock según el análisis de clasificación ABC+D,"
    ws[f'A{fila_nota}'].font = Font(italic=True, size=9, color="808080")
    ws.merge_cells(f'A{fila_nota}:E{fila_nota}')
    
    ws[f'A{fila_nota + 1}'] = "pero todavía están presentes en el inventario actual."
    ws[f'A{fila_nota + 1}'].font = Font(italic=True, size=9, color="808080")
    ws.merge_cells(f'A{fila_nota + 1}:E{fila_nota + 1}')


# Variable global para almacenar el stock por sección
df_stock_global = {}


def generar_informe():
    """
    Genera el informe completo de artículos de categoría C y D.
    """
    print("=" * 60)
    print("GENERANDO INFORME DE ARTÍCULOS C Y D PENDIENTES")
    print("=" * 60)
    
    # Cargar configuración
    config = cargar_configuracion()
    
    # Cargar stock actual
    print("\n📊 Cargando stock actual...")
    df_stock = cargar_stock_actual()
    print(f"  ✓ Stock cargado: {len(df_stock)} registros")
    
    # Cargar archivos de clasificación y procesar cada sección
    resultados = {}
    metricas_todas = []
    
    print("\n📁 Procesando secciones...")
    for seccion in SECCIONES:
        print(f"\n  ▶ Procesando {seccion}...")
        
        # Cargar clasificación
        df_clasificacion = cargar_clasificacion(seccion)
        
        if df_clasificacion is None:
            print(f"    ⚠️ Saltando {seccion} - no hay archivo de clasificación")
            continue
        
        # Identificar artículos de categoría C y D
        df_categoria_cd = identificar_articulos_categoria_c_d(df_clasificacion)
        print(f"    ✓ Artículos C+D en clasificación: {len(df_categoria_cd)}")
        
        # Comparar con stock actual
        df_en_stock = comparar_con_stock(df_categoria_cd, df_stock)
        print(f"    ✓ Artículos todavía en stock: {len(df_en_stock)}")
        
        # Guardar stock por sección para usarlo en crear_excel
        df_stock_global[seccion] = df_stock.copy()
        
        # Calcular métricas
        metricas = calcular_metricas(df_en_stock, df_stock, seccion)
        metricas_todas.append(metricas)
        
        resultados[seccion] = {
            'clasificacion': df_clasificacion,
            'categoria_cd': df_categoria_cd,
            'en_stock': df_en_stock,
            'metricas': metricas
        }
    
    # Crear archivo Excel
    print("\n📝 Generando archivo Excel...")
    workbook = Workbook()
    
    # Eliminar la hoja por defecto
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    
    # Crear hojas para cada sección
    for seccion in SECCIONES:
        if seccion in resultados:
            df_en_stock = resultados[seccion]['en_stock']
            metricas = resultados[seccion]['metricas']
            crear_excel(df_en_stock, metricas, seccion, workbook)
    
    # Generar nombre de archivo con fecha
    fecha_actual = datetime.now().strftime("%d%m%Y")
    nombre_archivo = f"Analisis_Categorias_C_y_D_{fecha_actual}.xlsx"
    ruta_salida = ANALISIS_CATEGORIA_CD_DIR / nombre_archivo
    
    # Guardar archivo
    workbook.save(ruta_salida)
    print(f"\n✅ Archivo generado: {ruta_salida}")
    
    # Mostrar resumen
    print("\n" + "=" * 60)
    print("RESUMEN DE RESULTADOS")
    print("=" * 60)
    
    for metricas in metricas_todas:
        print(f"\n📦 {metricas['seccion']}:")
        print(f"   Artículos C+D: {metricas['total_articulos_cd']}")
        print(f"   En stock: {metricas['total_articulos_en_stock']}")
        print(f"   Unidades: {metricas['unidades_totales_en_stock']}")
        print(f"   Sin eliminar: {metricas['porcentaje_sin_eliminar']:.1f}%")
    
    # Calcular totales
    total_articulos = sum(m['total_articulos_cd'] for m in metricas_todas)
    total_en_stock = sum(m['total_articulos_en_stock'] for m in metricas_todas)
    total_unidades = sum(m['unidades_totales_en_stock'] for m in metricas_todas)
    
    print("\n" + "=" * 60)
    print("TOTALES GENERALES")
    print("=" * 60)
    print(f"Total artículos C+D: {total_articulos}")
    print(f"Total artículos en stock: {total_en_stock}")
    print(f"Total unidades en stock: {total_unidades}")
    
    return ruta_salida


def buscar_archivo_semana_anterior(fecha_actual=None):
    """
    Busca el archivo de análisis de la semana anterior.
    Busca en el directorio de output y en user_input_files.
    """
    import re
    from datetime import datetime, timedelta
    
    if fecha_actual is None:
        fecha_actual = datetime.now()
    
    # Patrón para buscar archivos de análisis de categoría CD
    patron = r"Analisis_Categorias_C_y_D_(\d{8})\.xlsx"
    
    archivos_encontrados = []
    
    # Buscar en directorio de análisis
    for archivo in ANALISIS_CATEGORIA_CD_DIR.iterdir():
        match = re.match(patron, archivo.name)
        if match:
            fecha_str = match.group(1)
            try:
                fecha = datetime.strptime(fecha_str, "%d%m%Y")
                # Excluir archivos de la fecha actual
                if fecha.date() != fecha_actual.date():
                    archivos_encontrados.append((fecha, ANALISIS_CATEGORIA_CD_DIR / archivo.name))
            except:
                pass
    
    # Buscar en user_input_files (múltiples ubicaciones posibles)
    posibles_dirs = ["user_input_files", "../user_input_files", "/workspace/user_input_files"]
    
    for user_input_dir in posibles_dirs:
        if os.path.exists(user_input_dir):
            for archivo in os.listdir(user_input_dir):
                match = re.match(patron, archivo)
                if match:
                    fecha_str = match.group(1)
                    try:
                        fecha = datetime.strptime(fecha_str, "%d%m%Y")
                        # Excluir archivos de la fecha actual
                        if fecha.date() != fecha_actual.date():
                            archivos_encontrados.append((fecha, os.path.join(user_input_dir, archivo)))
                    except:
                        pass
    
    if not archivos_encontrados:
        return None
    
    # Ordenar por fecha (más reciente primero)
    archivos_encontrados.sort(reverse=True)
    
    # Obtener la fecha actual
    fecha_actual = datetime.now()
    
    # Buscar el archivo más reciente que sea anterior a la fecha actual
    for fecha, ruta in archivos_encontrados:
        if fecha < fecha_actual:
            return ruta
    
    return None


def comparar_con_semana_anterior(ruta_archivo_actual):
    """
    Compara el archivo actual con el de la semana anterior.
    """
    print("\n" + "=" * 60)
    print("BUSCANDO ARCHIVO DE LA SEMANA ANTERIOR")
    print("=" * 60)
    
    # Pasar la fecha actual para excluir archivos del mismo día
    archivo_anterior = buscar_archivo_semana_anterior(datetime.now())
    
    if archivo_anterior is None:
        print("⚠️ No se encontró archivo de la semana anterior para comparar")
        return None
    
    print(f"📊 Archivo anterior encontrado: {archivo_anterior}")
    
    # Importar y ejecutar la comparación
    try:
        from comparar_analisis_cd import comparar_archivos
        resultado = comparar_archivos(ruta_archivo_actual, archivo_anterior)
        
        # También enviar email con la comparación
        if resultado:
            try:
                from comparar_analisis_cd import enviar_email_informe as enviar_email_comparacion
                periodo = obtener_periodo_año()
                print("\n" + "=" * 60)
                print("ENVIANDO EMAIL DE COMPARACIÓN")
                print("=" * 60)
                email_enviado = enviar_email_comparacion(resultado, periodo)
                if email_enviado:
                    print(f"\n📧 Email de comparación enviado a los destinatarios: Ivan y Sandra")
            except Exception as e:
                print(f"  AVISO: No se pudo enviar el email de comparación: {e}")
        
        return resultado
    except Exception as e:
        print(f"⚠️ Error al comparar: {e}")
        return None


# ============================================================================
# FUNCIONES AUXILIARES PARA EMAIL
# ============================================================================

def obtener_periodo_año() -> str:
    """
    Obtiene el período y año del análisis actual.
    
    Returns:
        str: Período y año (ej: 'P2_2025')
    """
    try:
        datos_dinamicos = get_periodo_y_año_dinamico(tipo_calculo="actual")
        periodo = datos_dinamicos['periodo']
        año = datos_dinamicos['año']
        return f"{periodo}_{año}"
    except Exception as e:
        print(f"  AVISO: No se pudo obtener el período: {e}")
        return 'desconocido'


def enviar_email_informe(archivo_informe: str, periodo: str = 'desconocido') -> bool:
    """
    Envía un email a Ivan y Sandra con el informe de categoría C y D adjunto.
    
    Args:
        archivo_informe: Ruta del archivo Excel generado
        periodo: Período del análisis (ej: 'P2_2025')
    
    Returns:
        bool: True si el email fue enviado exitosamente, False en caso contrario
    """
    if not archivo_informe:
        print("  AVISO: No hay informe para enviar. No se enviará email.")
        return False
    
    # Verificar que el archivo existe
    if not Path(archivo_informe).exists():
        print(f"  AVISO: El archivo '{archivo_informe}' no existe. No se enviará email.")
        return False
    
    # Verificar contraseña en variable de entorno
    password = os.environ.get('EMAIL_PASSWORD')
    if not password:
        print(f"  AVISO: Variable de entorno 'EMAIL_PASSWORD' no configurada. No se enviará email.")
        print(f"  DEBUG: Variables de entorno disponibles: {list(os.environ.keys())}")
        return False
    else:
        print(f"  DEBUG: EMAIL_PASSWORD detectada (longitud: {len(password)})")
    
    # Enviar email a cada destinatario
    emails_enviados = 0
    
    for destinatario in DESTINATARIOS:
        nombre_destinatario = destinatario['nombre']
        email_destinatario = destinatario['email']
        
        try:
            # Crear mensaje MIME
            msg = MIMEMultipart()
            msg['From'] = f"{SMTP_CONFIG['remitente_nombre']} <{SMTP_CONFIG['remitente_email']}>"
            msg['To'] = email_destinatario
            msg['Subject'] = f"Viveverde: Análisis Categoría C y D - Período {periodo}"
            msg['Date'] = formatdate(localtime=True)
            
            # Cuerpo del email
            cuerpo = f"""Buenos días {nombre_destinatario},

Te adjunto en este correo el análisis de artículos de categoría C y D del período {periodo}.

Este informe muestra los artículos que deberían eliminarse del stock pero todavía están presentes.

Este informe te permitirá tomar decisiones sobre el inventario.

Atentamente,

Sistema de Pedidos Viveverde."""
            
            msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
            
            # Adjuntar archivo Excel
            filename = Path(archivo_informe).name
            with open(archivo_informe, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename= "{filename}"')
            part.add_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            msg.attach(part)
            
            # Enviar email mediante SSL
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(SMTP_CONFIG['servidor'], SMTP_CONFIG['puerto'], context=context) as server:
                server.login(SMTP_CONFIG['remitente_email'], password)
                server.sendmail(SMTP_CONFIG['remitente_email'], email_destinatario, msg.as_string())
            
            print(f"  Email enviado a {nombre_destinatario} ({email_destinatario})")
            emails_enviados += 1
            
        except smtplib.SMTPException as e:
            print(f"  ERROR SMTP al enviar email a {nombre_destinatario}: {e}")
        except Exception as e:
            print(f"  ERROR al enviar email a {nombre_destinatario}: {e}")
    
    return emails_enviados > 0


if __name__ == "__main__":
    try:
        archivo_salida = generar_informe()
        
        # Después de generar el informe, comparar con la semana anterior
        print("\n" + "=" * 60)
        print("INICIANDO COMPARACIÓN SEMANAL")
        print("=" * 60)
        resultado_comparacion = comparar_con_semana_anterior(archivo_salida)
        
        if resultado_comparacion:
            print(f"\n✅ Comparación generada: {resultado_comparacion}")
        
        # Obtener período para el email
        periodo = obtener_periodo_año()
        
        # Enviar email con el informe adjunto
        print("\n" + "=" * 60)
        print("ENVIANDO EMAIL")
        print("=" * 60)
        
        email_enviado = enviar_email_informe(archivo_salida, periodo)
        
        print(f"\n🎉 Proceso completado exitosamente!")
        print(f"📄 Archivo principal: {archivo_salida}")
        if email_enviado:
            print(f"Email enviado a los destinatarios: Ivan y Sandra")
    except Exception as e:
        print(f"\n❌ Error durante la generación: {str(e)}")
        import traceback
        traceback.print_exc()
