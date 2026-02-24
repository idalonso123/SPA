#!/usr/bin/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para comparar dos archivos de análisis de categoría C y D.
Compara las métricas de resumen entre el archivo actual y el de la semana pasada.
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import glob
import smtplib
import ssl
import os
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import formatdate

# Importar rutas centralizadas
from src.paths import OUTPUT_DIR, ANALISIS_CATEGORIA_CD_DIR, COMPARACION_CATEGORIA_CD_DIR

# ============================================================================
# CONFIGURACIÓN DE EMAIL
# ============================================================================

# Destinatarios: Ivan y Sandra
DESTINATARIOS = [
    {'nombre': 'Ivan', 'email': 'ivan.delgado@viveverde.es'},
    {'nombre': 'Sandra', 'email': 'ivan.delgado@viveverde.es'}
]

# Configuración del servidor SMTP
SMTP_CONFIG = {
    'servidor': 'smtp.serviciodecorreo.es',
    'puerto': 465,
    'remitente_email': 'ivan.delgado@viveverde.es',
    'remitente_nombre': 'Sistema de Pedidos Viveverde'
}


# ============================================================================
# FUNCIONES AUXILIARES PARA EMAIL
# ============================================================================

def enviar_email_informe(archivo_informe: str, periodo: str = 'desconocido') -> bool:
    """
    Envía un email a Ivan y Sandra con el informe de comparación de categoría C y D adjunto.
    
    Args:
        archivo_informe: Ruta del archivo Excel generado
        periodo: Período del análisis (ej: 'P2_2025')
    
    Returns:
        bool: True si el email fue enviado exitosamente, False en caso contrario
    """
    if not archivo_informe:
        print("  AVISO: No hay informe para enviar. No se enviará email.")
        return False
    
    # Verificar que el archivo existe
    if not Path(archivo_informe).exists():
        print(f"  AVISO: El archivo '{archivo_informe}' no existe. No se enviará email.")
        return False
    
    # Verificar contraseña en variable de entorno
    password = os.environ.get('EMAIL_PASSWORD')
    if not password:
        print(f"  AVISO: Variable de entorno 'EMAIL_PASSWORD' no configurada. No se enviará email.")
        print(f"  DEBUG: Variables de entorno disponibles: {list(os.environ.keys())}")
        return False
    else:
        print(f"  DEBUG: EMAIL_PASSWORD detectada (longitud: {len(password)})")
    
    # Enviar email a cada destinatario
    emails_enviados = 0
    
    for destinatario in DESTINATARIOS:
        nombre_destinatario = destinatario['nombre']
        email_destinatario = destinatario['email']
        
        try:
            # Crear mensaje MIME
            msg = MIMEMultipart()
            msg['From'] = f"{SMTP_CONFIG['remitente_nombre']} <{SMTP_CONFIG['remitente_email']}>"
            msg['To'] = email_destinatario
            msg['Subject'] = f"Viveverde: Comparación Análisis Categoría C y D - Período {periodo}"
            msg['Date'] = formatdate(localtime=True)
            
            # Cuerpo del email
            cuerpo = f"""Buenos días {nombre_destinatario},

Te adjunto en este correo la comparación de artículos de categoría C y D del período {periodo}.

Este informe muestra la evolución semanal de los artículos que deberían eliminarse del stock.

Este informe te permitirá identificar si las acciones de eliminación de stock están siendo efectivas.

Atentamente,

Sistema de Pedidos Viveverde."""
            
            msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
            
            # Adjuntar archivo Excel
            filename = Path(archivo_informe).name
            with open(archivo_informe, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename= "{filename}"')
            part.add_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            msg.attach(part)
            
            # Enviar email mediante SSL
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(SMTP_CONFIG['servidor'], SMTP_CONFIG['puerto'], context=context) as server:
                server.login(SMTP_CONFIG['remitente_email'], password)
                server.sendmail(SMTP_CONFIG['remitente_email'], email_destinatario, msg.as_string())
            
            print(f"  Email enviado a {nombre_destinatario} ({email_destinatario})")
            emails_enviados += 1
            
        except smtplib.SMTPException as e:
            print(f"  ERROR SMTP al enviar email a {nombre_destinatario}: {e}")
        except Exception as e:
            print(f"  ERROR al enviar email a {nombre_destinatario}: {e}")
    
    return emails_enviados > 0


def extraer_metricas_de_excel(ruta_archivo):
    """
    Extrae las métricas de resumen de un archivo Excel de análisis.
    """
    from openpyxl import load_workbook
    
    metricas = {}
    
    try:
        wb = load_workbook(ruta_archivo, data_only=True)
        
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            
            # Buscar la sección de métricas
            metricas_seccion = {}
            
            # Buscar la fila donde dice "MÉTRICAS DE RESUMEN"
            fila_metricas = None
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row[0] and "MÉTRICAS DE RESUMEN" in str(row[0]):
                    fila_metricas = i
                    break
            
            if fila_metricas:
                # Leer las métricas
                for i in range(fila_metricas + 1, fila_metricas + 10):
                    row = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
                    if row[0] and row[1] is not None:
                        clave = str(row[0]).strip()
                        valor = row[1]
                        metricas_seccion[clave] = valor
                
                metricas[nombre_hoja] = metricas_seccion
        
        return metricas
    
    except Exception as e:
        print(f"Error al leer {ruta_archivo}: {e}")
        return None


def comparar_metricas(metricas_actual, metricas_anterior):
    """
    Compara las métricas entre dos semanas.
    """
    comparacion = {}
    
    # Obtener todas las secciones
    todas_secciones = set(metricas_actual.keys()) | set(metricas_anterior.keys())
    
    for seccion in todas_secciones:
        metricas_seccion_actual = metricas_actual.get(seccion, {})
        metricas_seccion_anterior = metricas_anterior.get(seccion, {})
        
        comparacion_seccion = {}
        
        # Obtener todas las métricas
        todas_claves = set(metricas_seccion_actual.keys()) | set(metricas_seccion_anterior.keys())
        
        for clave in todas_claves:
            valor_actual = metricas_seccion_actual.get(clave, 0)
            valor_anterior = metricas_seccion_anterior.get(clave, 0)
            
            # Convertir a número
            try:
                if isinstance(valor_actual, str):
                    valor_actual = float(valor_actual.replace('%', '').replace(',', '.'))
                else:
                    valor_actual = float(valor_actual) if valor_actual is not None else 0
                
                if isinstance(valor_anterior, str):
                    valor_anterior = float(valor_anterior.replace('%', '').replace(',', '.'))
                else:
                    valor_anterior = float(valor_anterior) if valor_anterior is not None else 0
                
                diferencia = valor_actual - valor_anterior
                
                # Determinar si es positivo o negativo
                if diferencia > 0:
                    tendencia = "↑"
                elif diferencia < 0:
                    tendencia = "↓"
                else:
                    tendencia = "="
                
                comparacion_seccion[clave] = {
                    'actual': valor_actual,
                    'anterior': valor_anterior,
                    'diferencia': diferencia,
                    'tendencia': tendencia
                }
            except:
                comparacion_seccion[clave] = {
                    'actual': valor_actual,
                    'anterior': valor_anterior,
                    'diferencia': 0,
                    'tendencia': '='
                }
        
        comparacion[seccion] = comparacion_seccion
    
    return comparacion


def generar_excel_comparacion(comparacion, archivo_salida):
    """
    Genera un archivo Excel con la comparación.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FF008000", end_color="FF008000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    section_font = Font(bold=True, size=12, color="FF008000")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Crear hoja de resumen
    ws = wb.create_sheet(title="RESUMEN COMPARATIVO")
    
    # Título
    ws['A1'] = "COMPARACIÓN DE ARTÍCULOS C Y D - EVOLUCIÓN SEMANAL"
    ws['A1'].font = Font(bold=True, size=14, color="FF008000")
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws['A2'] = f"Fecha de comparación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:F2')
    
    # Encabezados
    headers = ['Sección', 'Métrica', 'Semana Anterior', 'Semana Actual', 'Diferencia', 'Tendencia']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Escribir datos
    fila_actual = 5
    
    for seccion, metricas in comparacion.items():
        # Título de sección
        ws.cell(row=fila_actual, column=1, value=seccion)
        ws.cell(row=fila_actual, column=1).font = section_font
        ws.merge_cells(f'A{fila_actual}:F{fila_actual}')
        ws.cell(row=fila_actual, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        fila_actual += 1
        
        for metrica, datos in metricas.items():
            # Convertir diferencia a formato legible
            diferencia = datos['diferencia']
            tendencia = datos['tendencia']
            
            # Formatear valores
            if 'unidades' in metrica.lower() or 'artículos' in metrica.lower():
                actual_str = f"{datos['actual']:.0f}"
                anterior_str = f"{datos['anterior']:.0f}"
                diferencia_str = f"{diferencia:+.0f}"
            elif '%' in metrica:
                actual_str = f"{datos['actual']:.1f}%"
                anterior_str = f"{datos['anterior']:.1f}%"
                diferencia_str = f"{diferencia:+.1f}%"
            else:
                actual_str = str(datos['actual'])
                anterior_str = str(datos['anterior'])
                diferencia_str = str(diferencia)
            
            # Color de tendencia
            if tendencia == "↓":
                color_tendencia = "FF008000"  # Verde (descenso)
            elif tendencia == "↑":
                color_tendencia = "FFFF0000"  # Rojo (aumento)
            else:
                color_tendencia = "FF808080"  # Gris (sin cambio)
            
            ws.cell(row=fila_actual, column=1, value="").border = thin_border
            ws.cell(row=fila_actual, column=2, value=metrica).border = thin_border
            ws.cell(row=fila_actual, column=3, value=anterior_str).border = thin_border
            ws.cell(row=fila_actual, column=4, value=actual_str).border = thin_border
            ws.cell(row=fila_actual, column=5, value=diferencia_str).border = thin_border
            
            cell_tendencia = ws.cell(row=fila_actual, column=6, value=tendencia)
            cell_tendencia.border = thin_border
            cell_tendencia.font = Font(bold=True, color=color_tendencia)
            
            fila_actual += 1
        
        fila_actual += 1  # Espacio entre secciones
    
    # Calcular totales
    ws.cell(row=fila_actual, column=1, value="TOTALES")
    ws.cell(row=fila_actual, column=1).font = section_font
    ws.merge_cells(f'A{fila_actual}:B{fila_actual}')
    ws.cell(row=fila_actual, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Calcular totales
    total_anterior = 0
    total_actual = 0
    
    for seccion, metricas in comparacion.items():
        for metrica, datos in metricas.items():
            if 'artículos' in metrica.lower() and 'en stock' in metrica.lower():
                total_anterior += datos['anterior']
                total_actual += datos['actual']
            elif 'unidades totales' in metrica.lower():
                total_anterior += datos['anterior']
                total_actual += datos['actual']
    
    diferencia_total = total_actual - total_anterior
    tendencia_total = "↓" if diferencia_total < 0 else ("↑" if diferencia_total > 0 else "=")
    
    ws.cell(row=fila_actual, column=3, value=f"{total_anterior:.0f}").border = thin_border
    ws.cell(row=fila_actual, column=4, value=f"{total_actual:.0f}").border = thin_border
    ws.cell(row=fila_actual, column=5, value=f"{diferencia_total:+.0f}").border = thin_border
    
    cell_tendencia_total = ws.cell(row=fila_actual, column=6, value=tendencia_total)
    cell_tendencia_total.font = Font(bold=True, size=12)
    if tendencia_total == "↓":
        cell_tendencia_total.font = Font(bold=True, size=12, color="FF008000")
    elif tendencia_total == "↑":
        cell_tendencia_total.font = Font(bold=True, size=12, color="FFFF0000")
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    
    # Guardar archivo
    wb.save(archivo_salida)
    return archivo_salida


def buscar_archivos_analisis():
    """
    Busca automáticamente el archivo más reciente y el anterior.
    """
    import re
    from datetime import datetime
    
    # Buscar todos los archivos de análisis
    patron = str(ANALISIS_CATEGORIA_CD_DIR / "Analisis_Categorias_C_y_D_*.xlsx")
    archivos = glob.glob(patron)
    
    if not archivos:
        print("❌ No se encontraron archivos de análisis en:", ANALISIS_CATEGORIA_CD_DIR)
        return None, None
    
    # Convertir a objetos Path
    from pathlib import Path
    archivos = [Path(f) for f in archivos]
    
    # Ordenar por nombre (que contiene la fecha en formato DDMMYYYY)
    # Ejemplo: Analisis_Categorias_C_y_D_21022026.xlsx
    def extraer_fecha(nombre):
        match = re.search(r'(\d{8})\.xlsx$', nombre)
        if match:
            fecha_str = match.group(1)
            try:
                return datetime.strptime(fecha_str, '%d%m%Y')
            except:
                pass
        return datetime(1900, 1, 1)  # Fecha muy antigua si no se puede parsear
    
    archivos.sort(key=lambda x: extraer_fecha(x.name), reverse=True)
    
    if len(archivos) < 2:
        print("⚠️ Solo se encontró un archivo. Se necesita al menos 2 para comparar.")
        return archivos[0] if archivos else None, None
    
    archivo_actual = archivos[0]
    archivo_anterior = archivos[1]
    
    print(f"📁 Archivo actual (más reciente): {archivo_actual.name}")
    print(f"📁 Archivo anterior (semana pasada): {archivo_anterior.name}")
    
    return archivo_actual, archivo_anterior


def obtener_fecha_semana_anterior(fecha_actual):
    """
    Obtiene la fecha de la semana anterior en formato DDMMYYYY.
    """
    from datetime import timedelta
    fecha_anterior = fecha_actual - timedelta(days=7)
    return fecha_actual.strftime('%d%m%Y')


def comparar_archivos(archivo_actual=None, archivo_anterior=None, archivo_salida=None):
    """
    Compara dos archivos de análisis de categoría C y D.
    Si no se especifican archivos, los busca automáticamente.
    """
    print("=" * 60)
    print("COMPARANDO ANÁLISIS DE ARTÍCULOS C Y D")
    print("=" * 60)
    
    # Si no se proporcionan archivos, buscarlos automáticamente
    if archivo_actual is None or archivo_anterior is None:
        print("\n🔍 Buscando archivos automáticamente...")
        archivo_actual, archivo_anterior = buscar_archivos_analisis()
        
        if archivo_actual is None or archivo_anterior is None:
            print("❌ No se pueden comparar los archivos. Se necesitan al menos 2 archivos.")
            return None
    
    print(f"\n📊 Archivo actual: {archivo_actual if isinstance(archivo_actual, str) else archivo_actual.name}")
    print(f"📊 Archivo anterior: {archivo_anterior if isinstance(archivo_anterior, str) else archivo_anterior.name}")
    
    # Extraer métricas
    print("\n📥 Extrayendo métricas...")
    metricas_actual = extraer_metricas_de_excel(archivo_actual)
    metricas_anterior = extraer_metricas_de_excel(archivo_anterior)
    
    if metricas_actual is None or metricas_anterior is None:
        print("❌ Error al leer los archivos")
        return None
    
    # Comparar métricas
    print("📊 Comparando métricas...")
    comparacion = comparar_metricas(metricas_actual, metricas_anterior)
    
    # Generar archivo Excel
    if archivo_salida is None:
        fecha = datetime.now().strftime("%d%m%Y")
        archivo_salida = COMPARACION_CATEGORIA_CD_DIR / f"Comparacion_Categorias_C_y_D_{fecha}.xlsx"
    
    print(f"\n📝 Generando archivo de comparación: {archivo_salida}")
    generar_excel_comparacion(comparacion, archivo_salida)
    
    # Mostrar resumen en consola
    print("\n" + "=" * 60)
    print("RESUMEN DE COMPARACIÓN")
    print("=" * 60)
    
    for seccion, metricas in comparacion.items():
        print(f"\n📦 {seccion}:")
        
        for metrica, datos in metricas.items():
            tendencia = datos['tendencia']
            diferencia = datos['diferencia']
            
            if 'unidades' in metrica.lower() or 'artículos' in metrica.lower():
                print(f"   {metrica}: {datos['anterior']:.0f} → {datos['actual']:.0f} ({diferencia:+.0f}) {tendencia}")
            elif '%' in metrica:
                print(f"   {metrica}: {datos['anterior']:.1f}% → {datos['actual']:.1f}% ({diferencia:+.1f}%) {tendencia}")
    
    # Calcular totales
    print("\n" + "=" * 60)
    print("EVOLUCIÓN TOTAL")
    print("=" * 60)
    
    total_articulos_anterior = 0
    total_articulos_actual = 0
    total_unidades_anterior = 0
    total_unidades_actual = 0
    
    for seccion, metricas in comparacion.items():
        for metrica, datos in metricas.items():
            if 'artículos' in metrica.lower() and 'en stock' in metrica.lower():
                total_articulos_anterior += datos['anterior']
                total_articulos_actual += datos['actual']
            elif 'unidades totales' in metrica.lower():
                total_unidades_anterior += datos['anterior']
                total_unidades_actual += datos['actual']
    
    print(f"\n📊 Total artículos en stock:")
    print(f"   Anterior: {total_articulos_anterior:.0f}")
    print(f"   Actual: {total_articulos_actual:.0f}")
    print(f"   Diferencia: {total_articulos_actual - total_articulos_anterior:+.0f}")
    
    print(f"\n📊 Total unidades en stock:")
    print(f"   Anterior: {total_unidades_anterior:.0f}")
    print(f"   Actual: {total_unidades_actual:.0f}")
    print(f"   Diferencia: {total_unidades_actual - total_unidades_anterior:+.0f}")
    
    if total_articulos_actual < total_articulos_anterior:
        print("\n✅ EVOLUCIÓN POSITIVA:减少 artículos de categoría C y D")
    elif total_articulos_actual > total_articulos_anterior:
        print("\n⚠️ EVOLUCIÓN NEGATIVA: Aumentan artículos de categoría C y D")
    else:
        print("\n➡️ SIN CAMBIOS: Mismo número de artículos")
    
    return archivo_salida


if __name__ == "__main__":
    import sys
    from src.date_utils import get_periodo_y_año_dinamico
    
    # Verificar si se proporcionan argumentos
    if len(sys.argv) >= 3:
        # Modo manual: usar los argumentos proporcionados
        archivo_actual = sys.argv[1]
        archivo_anterior = sys.argv[2]
        archivo_salida = sys.argv[3] if len(sys.argv) > 3 else None
    else:
        # Modo automático: buscar los archivos más recientes
        print("🔍 Modo automático: buscando archivos...")
        archivo_actual, archivo_anterior = buscar_archivos_analisis()
        archivo_salida = None
        
        if archivo_actual is None or archivo_anterior is None:
            print("\nUso (manual):")
            print("  python comparar_analisis_cd.py <archivo_actual> <archivo_anterior> [archivo_salida]")
            print("\nEjemplo:")
            print("  python comparar_analisis_cd.py Analisis_Categorias_C_y_D_21022026.xlsx Analisis_Categorias_C_y_D_13022026.xlsx")
            sys.exit(1)
    
    try:
        resultado = comparar_archivos(archivo_actual, archivo_anterior, archivo_salida)
        if resultado:
            print(f"\n✅ Comparación completada: {resultado}")
            
            # Obtener período para el email
            try:
                datos_dinamicos = get_periodo_y_año_dinamico(tipo_calculo="actual")
                periodo = f"{datos_dinamicos['periodo']}_{datos_dinamicos['año']}"
            except:
                periodo = datetime.now().strftime("%m%Y")
            
            # Enviar email con el informe adjunto
            print("\n" + "=" * 60)
            print("ENVIANDO EMAIL")
            print("=" * 60)
            
            email_enviado = enviar_email_informe(resultado, periodo)
            
            if email_enviado:
                print(f"\n📧 Email enviado a los destinatarios: Ivan y Sandra")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
