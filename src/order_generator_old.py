#!/usr/bin/env python3
"""
Módulo OrderGenerator - Generación de archivos de salida

Este módulo es responsable de crear los archivos Excel con formato profesional
que contienen los pedidos de compra generados. Implementa toda la lógica de
formateo, estilos, y estructuración de la información para facilitar su uso
por parte de los responsables de compras.

Autor: Sistema de Pedidos Vivero V2
Fecha: 2026-01-31
"""

import pandas as pd
import os
import logging
from datetime import datetime
from typing import Optional, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración del logger
logger = logging.getLogger(__name__)


class OrderGenerator:
    """
    Generador de archivos de salida para pedidos de compra.
    
    Esta clase encapsula toda la lógica de creación de archivos Excel,
    incluyendo el formateo de celdas, aplicación de estilos, creación de
    resúmenes y métricas, y organización de la información por semanas
    y secciones.
    
    Attributes:
        config (dict): Configuración del sistema
        rutas (dict): Rutas de archivos configuradas
    """
    
    def __init__(self, config: dict):
        """
        Inicializa el OrderGenerator con la configuración proporcionada.
        
        Args:
            config (dict): Diccionario con la configuración del sistema
        """
        self.config = config
        self.rutas = config.get('rutas', {})
        self.formato = config.get('formato_salida', {})
        
        # Estilos predefinidos
        self.HEADER_FILL = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
        self.HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
        
        self.THIN_BLACK_BORDER = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        
        logger.info("OrderGenerator inicializado correctamente")
    
    def obtener_directorio_salida(self) -> str:
        """
        Obtiene el directorio de salida configurado.
        
        Returns:
            str: Ruta del directorio de salida
        """
        base = self.rutas.get('directorio_base', '.')
        salida = self.rutas.get('directorio_salida', './data/output')
        
        # Si es ruta relativa, combinar con base
        if not os.path.isabs(salida):
            salida = os.path.join(base, salida)
        
        # Crear directorio si no existe
        os.makedirs(salida, exist_ok=True)
        
        return salida
    
    def generar_nombre_archivo(self, semana: int, seccion: Optional[str] = None,
                                incluir_fecha: bool = True) -> str:
        """
        Genera el nombre del archivo según el formato configurado.
        
        Args:
            semana (int): Número de semana
            seccion (Optional[str]): Nombre de la sección (si aplica)
            incluir_fecha (bool): Si True, incluye la fecha actual en el nombre
        
        Returns:
            str: Nombre del archivo generado
        """
        prefijo = self.formato.get('prefijo_archivo', 'Pedido_Semana')
        
        if incluir_fecha:
            fecha_actual = datetime.now().strftime('%d%m%Y')
            nombre = f"{prefijo}_{semana:02d}_{fecha_actual}"
        else:
            nombre = f"{prefijo}_{semana:02d}"
        
        if seccion:
            nombre += f"_{seccion}"
        
        nombre += ".xlsx"
        
        return nombre
    
    def generar_archivo_pedido(self, pedidos_df: pd.DataFrame, semana: int,
                                seccion: str, parametros: dict) -> Optional[str]:
        """
        Genera el archivo Excel con el pedido de una semana específica.
        
        Args:
            pedidos_df (pd.DataFrame): DataFrame con los artículos y cantidades
            semana (int): Número de semana
            seccion (str): Nombre de la sección
            parametros (dict): Parámetros utilizados en el cálculo
        
        Returns:
            Optional[str]: Ruta del archivo generado o None si hay error
        """
        logger.debug(f"[DEBUG generar_archivo_pedido] Recibido DataFrame con {len(pedidos_df)} registros")
        if len(pedidos_df) > 0:
            logger.debug(f"[DEBUG] Columnas disponibles: {list(pedidos_df.columns)}")
            if 'Pedido_Corregido_Stock' in pedidos_df.columns:
                logger.debug(f"[DEBUG] Distribución de Pedido_Corregido_Stock:")
                logger.debug(f"  Con valor 0: {len(pedidos_df[pedidos_df['Pedido_Corregido_Stock'] == 0])}")
                logger.debug(f"  Con valor > 0: {len(pedidos_df[pedidos_df['Pedido_Corregido_Stock'] > 0])}")
                if 'Ventas_Objetivo' in pedidos_df.columns:
                    logger.debug(f"  Suma Ventas_Objetivo: {pedidos_df['Ventas_Objetivo'].sum():.2f}€")
        
        if len(pedidos_df) == 0:
            logger.warning(f"No hay pedidos para generar en semana {semana}")
            return None
        
        # Filtrar artículos con Pedido_Corregido_Stock > 0
        pedidos_filtrados = pedidos_df[pedidos_df['Pedido_Corregido_Stock'] > 0].copy()
        
        logger.debug(f"[DEBUG] Tras filtrar Pedido_Corregido_Stock > 0: {len(pedidos_filtrados)} registros")
        
        if len(pedidos_filtrados) == 0:
            logger.warning(f"Tras filtrar, no hay artículos con pedido > 0 para semana {semana}")
            logger.warning(f"[DEBUG] Posible causa: Todos los artículos tienen Pedido_Corregido_Stock = 0")
            return None
        
        # Ordenar por proveedor y código
        pedidos_filtrados = pedidos_filtrados.sort_values(
            ['Proveedor', 'Codigo_Articulo', 'Ventas_Objetivo'],
            ascending=[True, True, False]
        )
        
        # Generar nombre y ruta del archivo
        nombre_archivo = self.generar_nombre_archivo(semana, seccion)
        dir_salida = self.obtener_directorio_salida()
        ruta_completa = os.path.join(dir_salida, nombre_archivo)
        
        logger.info(f"Generando archivo: {ruta_completa}")
        
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Semana_{semana}"
            
            # Definir anchuras de columna
            COLUMN_WIDTHS = {
                'A': 11.25,  # Código artículo
                'B': 50.00,  # Nombre Artículo
                'C': 8.50,   # Talla
                'D': 6.50,   # Color
                'E': 11.00,  # Sección
                'F': 11.00,  # Unidades Calculadas
                'G': 11.50,  # PVP
                'H': 11.50,  # Coste Pedido
                'I': 10.00,  # Categoría
                'J': 15.00,  # Acción Aplicada
                'K': 12.50,  # Stock Mínimo Objetivo
                'L': 10.50,  # Diferencia Stock (era M)
                'M': 9.70,   # Ventas Objetivo (era N)
                'N': 11.50,  # Beneficio Objetivo (era O)
                'O': 27.00,  # Proveedor (era P)
                'P': 11.00,  # Pedido Corregido Stock (FASE 2) (era Q)
                'Q': 11.00,  # Ventas Reales (FASE 2) (era R)
                'R': 11.00,  # Tendencia Consumo (FASE 2) (era S)
                'S': 11.00   # Pedido Final (FASE 2) (era T)
            }
            
            COLUMN_HEADERS = [
                'Código artículo',      # A
                'Nombre Artículo',      # B
                'Talla',                # C
                'Color',                # D
                'Sección',              # E
                'Unidades Calculadas',  # F
                'PVP',                  # G
                'Coste Pedido',         # H
                'Categoría',            # I
                'Acción Aplicada',      # J
                'Stock Mínimo Objetivo',# K
                'Diferencia Stock',     # L (era M)
                'Ventas Objetivo',      # M (era N)
                'Beneficio Objetivo',   # N (era O)
                'Proveedor',            # O (era P)
                'Pedido Corregido Stock',# P (FASE 2 - Corrección 1) (era Q)
                'Ventas Reales',        # Q (FASE 2 - Datos) (era R)
                'Tendencia Consumo',    # R (FASE 2 - Corrección 2) (era S)
                'Pedido Final'          # S (FASE 2 - Resultado) (era T)
            ]
            
            COLUMN_MAPPING = {
                'Codigo_Articulo': 'Código artículo',
                'Nombre_Articulo': 'Nombre Artículo',
                'Talla': 'Talla',
                'Color': 'Color',
                'Seccion': 'Sección',
                'Unidades_Finales': 'Unidades Calculadas',
                'PVP': 'PVP',
                'Coste_Pedido': 'Coste Pedido',
                'Categoria': 'Categoría',
                'Accion_Aplicada': 'Acción Aplicada',
                'Stock_Minimo_Objetivo': 'Stock Mínimo Objetivo',
                'Diferencia_Stock': 'Diferencia Stock',
                'Ventas_Objetivo': 'Ventas Objetivo',
                'Beneficio_Objetivo': 'Beneficio Objetivo',
                'Proveedor': 'Proveedor',
                'Pedido_Corregido_Stock': 'Pedido Corregido Stock',
                'Ventas_Reales': 'Ventas Reales',
                'Tendencia_Consumo': 'Tendencia Consumo',
                'Pedido_Final': 'Pedido Final'
            }
            
            # Escribir cabeceras
            for col_idx, header in enumerate(COLUMN_HEADERS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.THIN_BLACK_BORDER
            
            # Renombrar y reordenar columnas
            pedidos_renamed = pedidos_filtrados.rename(columns=COLUMN_MAPPING)
            pedidos_renamed = pedidos_renamed[COLUMN_HEADERS]
            
            # Escribir datos
            for r_idx, row in enumerate(dataframe_to_rows(pedidos_renamed, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = self.THIN_BLACK_BORDER
                    
                    header_name = COLUMN_HEADERS[c_idx - 1]
                    
                    if header_name in ['PVP', 'Coste Pedido', 'Ventas Objetivo', 'Beneficio Objetivo']:
                        if isinstance(value, (int, float)):
                            cell.value = round(value, 2)
                        cell.number_format = '#,##0.00'
                        
                    elif header_name in ['Unidades Calculadas', 'Stock Mínimo Objetivo',
                                         'Diferencia Stock', 
                                         'Pedido Corregido Stock', 'Ventas Reales', 
                                         'Tendencia Consumo', 'Pedido Final']:
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0'
                    
                    elif header_name == 'Proveedor':
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Aplicar anchuras de columna
            for col_letter, width in COLUMN_WIDTHS.items():
                ws.column_dimensions[col_letter].width = width
            
            # Añadir métricas de resumen
            summary_row = len(pedidos_filtrados) + 3
            
            # Fusionar celdas para el título
            ws.merge_cells(f'B{summary_row}:C{summary_row}')
            
            # Título del resumen
            title_cell = ws.cell(row=summary_row, column=2, value="METRICAS DE RESUMEN")
            title_cell.font = Font(bold=True, size=12, color="FFFFFF")
            title_cell.fill = self.HEADER_FILL
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Métricas
            metricas_labels = [
                ("Total_Unidades:", int(pedidos_filtrados['Pedido_Final'].sum())),
                ("Total_Articulos:", len(pedidos_filtrados)),
                ("Total_Importe:", f"{pedidos_filtrados['Ventas_Objetivo'].sum():.2f}€"),
                ("Objetivo_Semana:", f"{parametros.get('objetivos_semanales', {}).get(str(semana), 0)}€"),
                ("Factor_Crecimiento:", f"{parametros.get('objetivo_crecimiento', 0.05)*100:.0f}%"),
                ("Factor_Festivo:", f"{parametros.get('festivos', {}).get(str(semana), 0)*100:.0f}%"),
                ("Articulos_A:", len(pedidos_filtrados[pedidos_filtrados['Categoria'] == 'A'])),
                ("Articulos_B:", len(pedidos_filtrados[pedidos_filtrados['Categoria'] == 'B'])),
                ("Articulos_C:", len(pedidos_filtrados[pedidos_filtrados['Categoria'] == 'C'])),
                ("Stock_Minimo_%:", f"{parametros.get('stock_minimo_porcentaje', 0.30)*100:.0f}%"),
                ("Stock_Minimo_Objetivo:", int(pedidos_filtrados['Stock_Minimo_Objetivo'].sum())),
                ("Total_Ajuste_Stock:", int(pedidos_filtrados['Diferencia_Stock'].sum()))
            ]
            
            for i, (label, value) in enumerate(metricas_labels):
                # Columna B (etiquetas)
                cell_label = ws.cell(row=summary_row + 1 + i, column=2, value=label)
                cell_label.font = Font(bold=True, size=10)
                cell_label.border = self.THIN_BLACK_BORDER
                cell_label.alignment = Alignment(horizontal='left', vertical='center')
                
                # Columna C (valores)
                cell_value = ws.cell(row=summary_row + 1 + i, column=3, value=value)
                cell_value.border = self.THIN_BLACK_BORDER
                cell_value.alignment = Alignment(horizontal='center', vertical='center')
            
            # Guardar archivo
            wb.save(ruta_completa)
            
            logger.info(f"Archivo guardado: {ruta_completa}")
            return ruta_completa
            
        except Exception as e:
            logger.error(f"Error al generar archivo: {str(e)}")
            return None
    
    def generar_resumen_excel(self, resumen_df: pd.DataFrame, seccion: str) -> Optional[str]:
        """
        Genera un archivo Excel con el resumen consolidado de pedidos.
        
        Args:
            resumen_df (pd.DataFrame): DataFrame con el resumen de pedidos
            seccion (str): Nombre de la sección
        
        Returns:
            Optional[str]: Ruta del archivo generado o None si hay error
        """
        if len(resumen_df) == 0:
            logger.warning("No hay datos para generar resumen")
            return None
        
        # Generar nombre del archivo
        dir_salida = self.obtener_directorio_salida()
        nombre_archivo = f"Resumen_Pedidos_{seccion}_{datetime.now().strftime('%d%m%Y')}.xlsx"
        ruta_completa = os.path.join(dir_salida, nombre_archivo)
        
        logger.info(f"Generando resumen: {ruta_completa}")
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumen"
            
            # Configurar columnas
            COLUMN_WIDTHS = {
                'A': 8, 'B': 18, 'C': 16, 'D': 22, 'E': 18, 'F': 12, 'G': 15, 'H': 16, 'I': 14,
                'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 14, 'O': 14, 'P': 14
            }
            
            # Cabeceras
            COLUMN_HEADERS = [
                'Sección', 'Semana', 'Vtas. semana año pasado', 'Objetivo semana', 'Obj. semana + % crec. anual', 
                'Obj. semana + % crec. + Festivos', '% Obj. crecim. + Festivos', 'Total Unidades', 'Total Articulos', 
                'Total Importe', 'Alcance %', 'Articulos A', 'Articulos B', 'Articulos C',
                '% Festivo', '% Stock Min', 'Stock Min Obj'
            ]
            
            COLUMN_MAPPING = {
                'Seccion': 'Sección',
                'Semana': 'Semana',
                'Vtas. semana año pasado': 'Vtas. semana año pasado',
                'Objetivo_Semana': 'Objetivo semana',
                'Obj. semana + % crec. anual': 'Obj. semana + % crec. anual',
                'Obj. semana + % crec. + Festivos': 'Obj. semana + % crec. + Festivos',
                '% Obj. crecim. + Festivos': '% Obj. crecim. + Festivos',
                'Total_Unidades': 'Total Unidades',
                'Total_Articulos': 'Total Articulos',
                'Total_Importe': 'Total Importe',
                'Alcance_Objetivo_%': 'Alcance %',
                'Articulos_A': 'Articulos A',
                'Articulos_B': 'Articulos B',
                'Articulos_C': 'Articulos C',
                'Incremento_Festivo_%': '% Festivo',
                'Stock_Minimo_%': '% Stock Min',
                'Stock_Minimo_Objetivo': 'Stock Min Obj'
            }
            
            # Escribir cabeceras (fila 2)
            for col_idx, header in enumerate(COLUMN_HEADERS, 1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.THIN_BLACK_BORDER
            
            # Renombrar y reordenar columnas
            resumen_renamed = resumen_df.rename(columns=COLUMN_MAPPING)
            resumen_renamed = resumen_renamed[COLUMN_HEADERS]
            
            # Escribir datos (empezar en fila 3)
            for r_idx, row in enumerate(dataframe_to_rows(resumen_renamed, index=False, header=False), 3):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = self.THIN_BLACK_BORDER
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    header_name = COLUMN_HEADERS[c_idx - 1]
                    
                    # Formatear según tipo de dato
                    if header_name in ['Vtas. semana año pasado', 'Objetivo semana', 
                                      'Obj. semana + % crec. anual', 
                                      'Obj. semana + % crec. + Festivos', 
                                      'Total Importe', 'Stock Min Obj']:
                        if isinstance(value, (int, float)):
                            cell.value = int(round(value)) if value == int(value) else int(value)
                        cell.number_format = '#,##0'
                    
                    elif header_name in ['% Obj. crecim. + Festivos', 'Alcance %', '% Festivo', '% Stock Min']:
                        if isinstance(value, (int, float)):
                            cell.value = round(value, 1)
                        cell.number_format = '0.0'
            
            # Aplicar anchuras de columna
            for col_letter, width in COLUMN_WIDTHS.items():
                ws.column_dimensions[col_letter].width = width
            
            # Añadir título en fila 1
            ws.merge_cells('A1:P1')
            titulo = f"RESUMEN DE PEDIDOS DE COMPRA - {seccion.upper()} - VIVERO ARANJUEZ 2026"
            ws['A1'] = titulo
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = self.HEADER_FILL
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Guardar archivo
            wb.save(ruta_completa)
            
            logger.info(f"Resumen guardado: {ruta_completa}")
            return ruta_completa
            
        except Exception as e:
            logger.error(f"Error al generar resumen: {str(e)}")
            return None
    
    def generar_archivo_csv(self, pedidos_df: pd.DataFrame, semana: int,
                            seccion: str) -> Optional[str]:
        """
        Genera un archivo CSV con los datos del pedido.
        
        Args:
            pedidos_df (pd.DataFrame): DataFrame con los artículos y cantidades
            semana (int): Número de semana
            seccion (str): Nombre de la sección
        
        Returns:
            Optional[str]: Ruta del archivo generado o None si hay error
        """
        if len(pedidos_df) == 0:
            return None
        
        # Filtrar artículos con Pedido_Corregido_Stock > 0
        pedidos_filtrados = pedidos_df[pedidos_df['Pedido_Corregido_Stock'] > 0].copy()
        
        if len(pedidos_filtrados) == 0:
            return None
        
        # Generar nombre del archivo
        dir_salida = self.obtener_directorio_salida()
        nombre_archivo = f"Pedido_Semana_{semana:02d}_{seccion}_{datetime.now().strftime('%d%m%Y')}.csv"
        ruta_completa = os.path.join(dir_salida, nombre_archivo)
        
        try:
            # Seleccionar columnas relevantes para CSV
            columnas_csv = [
                'Codigo_Articulo', 'Nombre_Articulo', 'Talla', 'Color',
                'Pedido_Corregido_Stock', 'PVP', 'Coste_Pedido', 'Proveedor', 'Categoria'
            ]
            
            pedidos_csv = pedidos_filtrados[columnas_csv].copy()
            pedidos_csv.to_csv(ruta_completa, index=False, encoding='utf-8-sig')
            
            logger.info(f"CSV guardado: {ruta_completa}")
            return ruta_completa
            
        except Exception as e:
            logger.error(f"Error al generar CSV: {str(e)}")
            return None


# Funciones de utilidad para uso directo
def crear_order_generator(config: dict) -> OrderGenerator:
    """
    Crea una instancia del OrderGenerator.
    
    Args:
        config (dict): Configuración del sistema
    
    Returns:
        OrderGenerator: Instancia inicializada del generador de pedidos
    """
    return OrderGenerator(config)


if __name__ == "__main__":
    # Ejemplo de uso
    print("OrderGenerator - Módulo de generación de archivos de salida")
    print("=" * 50)
    
    # Configurar logging básico
    logging.basicConfig(level=logging.INFO)
    
    # Ejemplo de configuración
    config_ejemplo = {
        'rutas': {
            'directorio_base': '.',
            'directorio_salida': './data/output'
        },
        'formato_salida': {
            'prefijo_archivo': 'Pedido_Semana',
            'incluir_fecha_en_nombre': True
        }
    }
    
    # Crear OrderGenerator
    generator = crear_order_generator(config_ejemplo)
    
    # Ejemplo de generación de nombre
    nombre = generator.generar_nombre_archivo(10, 'vivero')
    print(f"Nombre de archivo generado: {nombre}")
    
    print("\nOrderGenerator listo para usar.")
