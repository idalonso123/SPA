#!/usr/bin/env python3
"""
Módulo OrderGenerator - Generación de archivos de salida

Este módulo es responsable de crear los archivos Excel con formato profesional
que contienen los pedidos de compra generados. Implementa toda la lógica de
formateo, estilos, y estructuración de la información para facilitar su uso
por parte de los responsables de compras.

Autor: Sistema de Pedidos Viveverde V2
Fecha: 2026-01-31
"""

import pandas as pd
import os
import logging
from datetime import datetime
from typing import Optional, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from src.paths import INPUT_DIR, OUTPUT_DIR, PEDIDOS_SEMANALES_DIR, PEDIDOS_SEMANALES_RESUMEN_DIR, RESUMENES_DIR
from openpyxl.worksheet.page import PageMargins

# Configuración del logger
logger = logging.getLogger(__name__)

# Intentamos importar el servicio de alertas (si está disponible)
try:
    from src.alert_service import crear_alert_service
    ALERT_SERVICE = None
    def get_alert_service():
        global ALERT_SERVICE
        if ALERT_SERVICE is None:
            try:
                from src.config_loader import cargar_configuracion
                config = cargar_configuracion()
                if config:
                    ALERT_SERVICE = crear_alert_service(config)
            except:
                pass
        return ALERT_SERVICE
except ImportError:
    def get_alert_service():
        return None


class OrderGenerator:
    """
    Generador de archivos de salida para pedidos de compra.
    
    Esta clase encapsula toda la lógica de creación de archivos Excel,
    incluyendo el formateo de celdas, aplicación de estilos, creación de
    resúmenes y métricas, y organización de la información por semanas
    y secciones.
    
    Attributes:
        config (dict): Configuración del sistema
        rutas (dict): Rutas de archivos configuradas
    """
    
    def __init__(self, config: dict):
        """
        Inicializa el OrderGenerator con la configuración proporcionada.
        
        Args:
            config (dict): Diccionario con la configuración del sistema
        """
        self.config = config
        self.rutas = config.get('rutas', {})
        self.formato = config.get('formato_salida', {})
        
        # Estilos predefinidos
        self.HEADER_FILL = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
        self.HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
        
        self.THIN_BLACK_BORDER = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        
        logger.info("OrderGenerator inicializado correctamente")
    
    def obtener_directorio_salida(self) -> str:
        """
        Obtiene el directorio de salida configurado.
        
        Returns:
            str: Ruta del directorio de salida
        """
        # Primero verificar si hay una configuración personalizada en config
        base = self.rutas.get('directorio_base', '.')
        salida = self.rutas.get('directorio_salida')
        
        # Si no hay configuración personalizada, usar la ruta centralizada
        if salida is None:
            salida = str(PEDIDOS_SEMANALES_DIR)
        
        # Si es ruta relativa, combinar con base
        if not os.path.isabs(salida):
            salida = os.path.join(base, salida)
        
        # Crear directorio si no existe
        os.makedirs(salida, exist_ok=True)
        
        return salida
    
    def obtener_directorio_resumenes(self) -> str:
        """
        Obtiene el directorio de salida para resúmenes consolidados.
        
        Returns:
            str: Ruta del directorio de resúmenes
        """
        # Primero verificar si hay una configuración personalizada en config
        base = self.rutas.get('directorio_base', '.')
        salida = self.rutas.get('directorio_resumenes')
        
        # Si no hay configuración personalizada, usar la ruta centralizada
        if salida is None:
            salida = str(RESUMENES_DIR)
        
        # Si es ruta relativa, combinar con base
        if not os.path.isabs(salida):
            salida = os.path.join(base, salida)
        
        # Crear directorio si no existe
        os.makedirs(salida, exist_ok=True)
        
        return salida
    
    def generar_nombre_archivo(self, semana: int, seccion: Optional[str] = None,
                                incluir_fecha: bool = True) -> str:
        """
        Genera el nombre del archivo según el formato configurado.
        
        Args:
            semana (int): Número de semana
            seccion (Optional[str]): Nombre de la sección (si aplica)
            incluir_fecha (bool): Si True, incluye la fecha actual en el nombre
        
        Returns:
            str: Nombre del archivo generado
        """
        prefijo = self.formato.get('prefijo_archivo', 'Pedido_Semana')
        
        if incluir_fecha:
            fecha_actual = datetime.now().strftime('%d%m%Y')
            nombre = f"{prefijo}_{semana:02d}_{fecha_actual}"
        else:
            nombre = f"{prefijo}_{semana:02d}"
        
        if seccion:
            nombre += f"_{seccion}"
        
        nombre += ".xlsx"
        
        return nombre
    
    def generar_archivo_pedido(self, pedidos_df: pd.DataFrame, semana: int,
                                seccion: str, parametros: dict) -> Optional[str]:
        """
        Genera el archivo Excel con el pedido de una semana específica.
        
        Args:
            pedidos_df (pd.DataFrame): DataFrame con los artículos y cantidades
            semana (int): Número de semana
            seccion (str): Nombre de la sección
            parametros (dict): Parámetros utilizados en el cálculo
        
        Returns:
            Optional[str]: Ruta del archivo generado o None si hay error
        """
        logger.debug(f"[DEBUG generar_archivo_pedido] Recibido DataFrame con {len(pedidos_df)} registros")
        if len(pedidos_df) > 0:
            logger.debug(f"[DEBUG] Columnas disponibles: {list(pedidos_df.columns)}")
            if 'Pedido_Corregido_Stock' in pedidos_df.columns:
                logger.debug(f"[DEBUG] Distribución de Pedido_Corregido_Stock:")
                logger.debug(f"  Con valor 0: {len(pedidos_df[pedidos_df['Pedido_Corregido_Stock'] == 0])}")
                logger.debug(f"  Con valor > 0: {len(pedidos_df[pedidos_df['Pedido_Corregido_Stock'] > 0])}")
                if 'Ventas_Objetivo' in pedidos_df.columns:
                    logger.debug(f"  Suma Ventas_Objetivo: {pedidos_df['Ventas_Objetivo'].sum():.2f}€")
        
        if len(pedidos_df) == 0:
            logger.warning(f"No hay pedidos para generar en semana {semana}")
            # Enviar alerta específica
            alert_svc = get_alert_service()
            if alert_svc:
                alert_svc.alerta_pedido_error(seccion, semana, Exception("No hay pedidos para generar"))
            return None
        
        # Filtrar artículos con Unidades_Finales > 0 (artículos calculados por el forecast)
        # Nota: Cambiado de Pedido_Corregido_Stock > 0 a Unidades_Finales > 0 para preservar
        # artículos con datos válidos (incluyendo Stock Real) aunque Pedido_Corregido_Stock sea 0
        # debido a sobrestock (Stock_Real > Unidades_Finales + Stock_Minimo)
        pedidos_filtrados = pedidos_df[pedidos_df['Unidades_Finales'] > 0].copy()

        logger.debug(f"[DEBUG] Tras filtrar Unidades_Finales > 0: {len(pedidos_filtrados)} registros")

        if len(pedidos_filtrados) == 0:
            logger.warning(f"Tras filtrar, no hay artículos con Unidades_Finales > 0 para semana {semana}")
            logger.warning(f"[DEBUG] Posible causa: Todos los artículos tienen Unidades_Finales = 0")
            return None
        
        # Ordenar por proveedor y código
        pedidos_filtrados = pedidos_filtrados.sort_values(
            ['Proveedor', 'Codigo_Articulo', 'Ventas_Objetivo'],
            ascending=[True, True, False]
        )
        
        # Generar nombre y ruta del archivo
        nombre_archivo = self.generar_nombre_archivo(semana, seccion)
        dir_salida = self.obtener_directorio_salida()
        ruta_completa = os.path.join(dir_salida, nombre_archivo)
        
        logger.info(f"Generando archivo: {ruta_completa}")
        
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Semana_{semana}"
            
            # Definir orden de columnas con Stock Real entre Proveedor y Pedido Corregido Stock
            # Formato: Nombre en primera línea, fórmula en segunda línea
            COLUMN_HEADERS = [
                'Código artículo',      # A - 1
                'Nombre Artículo',      # B - 2
                'Talla',                # C - 3
                'Color',                # D - 4
                'Sección',              # E - 5
                'Unidades Calculadas',  # F - 6
                'PVP',                  # G - 7
                'Coste Pedido',         # H - 8
                'Categoría',            # I - 9
                'Acción Aplicada',      # J - 10
                'Stock Mínimo Objetivo',# K - 11
                'Diferencia Stock',     # L - 12
                'Ventas Objetivo\n13=7x21',      # M - 13
                'Beneficio Objetivo',   # N - 14
                'Proveedor',            # O - 15
                'Stock Real',           # P - 16
                'Pedido Corregido Stock\n17=6+11-16',# Q - 17
                'uds. Objetivo semana pasada',# R - 18
                'Uds. Vtas. reales semana pasada',  # S - 19
                'Tendencia Consumo\n20=19-18',     # T - 20
                'Pedido Final\n21=6+11-16+20'          # U - 21
            ]
            
            # Mapeo de nombres de columnas del DataFrame a nombres en Excel
            COLUMN_MAPPING = {
                'Codigo_Articulo': 'Código artículo',
                'Nombre_Articulo': 'Nombre Artículo',
                'Talla': 'Talla',
                'Color': 'Color',
                'Seccion': 'Sección',
                'Unidades_Finales': 'Unidades Calculadas',
                'PVP': 'PVP',
                'Coste_Pedido': 'Coste Pedido',
                'Categoria': 'Categoría',
                'Accion_Aplicada': 'Acción Aplicada',
                'Stock_Minimo_Objetivo': 'Stock Mínimo Objetivo',
                'Diferencia_Stock': 'Diferencia Stock',
                'Ventas_Objetivo': 'Ventas Objetivo\n13=7x21',
                'Beneficio_Objetivo': 'Beneficio Objetivo',
                'Proveedor': 'Proveedor',
                'Stock_Real': 'Stock Real',
                'Pedido_Corregido_Stock': 'Pedido Corregido Stock\n17=6+11-16',
                'Unidades_Calculadas_Semana_Pasada': 'uds. Objetivo semana pasada',
                'Ventas_Reales': 'Uds. Vtas. reales semana pasada',
                'Tendencia_Consumo': 'Tendencia Consumo\n20=19-18',
                'Pedido_Final': 'Pedido Final\n21=6+11-16+20'
            }
            
            # Definir anchuras de columna
            COLUMN_WIDTHS = {
                'A': 11.25,  # Código artículo
                'B': 50.00,  # Nombre Artículo
                'C': 8.50,   # Talla
                'D': 6.50,   # Color
                'E': 11.00,  # Sección
                'F': 11.00,  # Unidades Calculadas
                'G': 11.50,  # PVP
                'H': 11.50,  # Coste Pedido
                'I': 10.00,  # Categoría
                'J': 15.00,  # Acción Aplicada
                'K': 12.50,  # Stock Mínimo Objetivo
                'L': 10.50,  # Diferencia Stock
                'M': 9.70,   # Ventas Objetivo
                'N': 11.50,  # Beneficio Objetivo
                'O': 27.00,  # Proveedor
                'P': 11.00,  # Stock Real (MOVED here)
                'Q': 18.00,  # Pedido Corregido Stock (with formula)
                'R': 14.00,  # uds. Objetivo semana pasada
                'S': 16.00,  # Uds. Vtas. reales semana pasada
                'T': 18.00,  # Tendencia Consumo (with formula)
                'U': 22.00   # Pedido Final (with formula)
            }
            
            # Fila 1: Números de columna (índice)
            for col_idx in range(1, len(COLUMN_HEADERS) + 1):
                cell = ws.cell(row=1, column=col_idx, value=col_idx)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.THIN_BLACK_BORDER
            
            # Fila 2: Cabeceras
            for col_idx, header in enumerate(COLUMN_HEADERS, 1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.THIN_BLACK_BORDER
            
            # Renombrar y reordenar columnas
            pedidos_renamed = pedidos_filtrados.rename(columns=COLUMN_MAPPING)
            pedidos_renamed = pedidos_renamed[COLUMN_HEADERS]
            
            # Escribir datos (empezar en fila 3, ya que fila 1=números, fila 2=cabeceras)
            for r_idx, row in enumerate(dataframe_to_rows(pedidos_renamed, index=False, header=False), 3):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = self.THIN_BLACK_BORDER
                    
                    header_name = COLUMN_HEADERS[c_idx - 1]
                    
                    if header_name in ['PVP', 'Coste Pedido', 'Ventas Objetivo\n13=7x21', 'Beneficio Objetivo']:
                        if isinstance(value, (int, float)):
                            cell.value = round(value, 2)
                        cell.number_format = '#,##0.00'
                        
                    elif header_name in ['Unidades Calculadas', 'Stock Mínimo Objetivo',
                                         'Diferencia Stock', 
                                         'Pedido Corregido Stock\n17=6+11-16', 
                                         'Ventas Obj. Semana Pasada',
                                         'Ventas Reales', 
                                         'Stock Real',
                                         'Tendencia Consumo\n20=19-18', 'Pedido Final\n21=6+11-16+20']:
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0'
                    
                    elif header_name == 'Proveedor':
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Centrar columnas 9 (Categoría), 13 (Ventas Objetivo) y 21 (Pedido Final)
                    # Este centrado se aplica siempre, independientemente del formato de número
                    if header_name in ['Categoría', 'Ventas Objetivo\n13=7x21', 'Pedido Final\n21=6+11-16+20']:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar anchuras de columna
            for col_letter, width in COLUMN_WIDTHS.items():
                ws.column_dimensions[col_letter].width = width
            
            # OCULTAR COLUMNAS ESPECÍFICAS
            # Columnas a ocultar: 5, 6, 7, 8, 10, 11, 12, 14, 17, 18, 19, 20
            columns_to_hide = ['E', 'F', 'G', 'H', 'J', 'K', 'L', 'N', 'Q', 'R', 'S', 'T']
            for col_letter in columns_to_hide:
                ws.column_dimensions[col_letter].hidden = True
            
            # OCULTAR FILAS DONDE Pedido Final = 0 (sin eliminar los datos)
            pedido_final_col_idx = 21  # Columna U (Pedido Final)
            for row_idx in range(3, len(pedidos_filtrados) + 3):
                cell_value = ws.cell(row=row_idx, column=pedido_final_col_idx).value
                if cell_value is not None and cell_value == 0:
                    ws.row_dimensions[row_idx].hidden = True
            
            # OCULTAR COLUMNA 16 (Stock Real - Columna P)
            ws.column_dimensions['P'].hidden = True
            
            # CONFIGURACIÓN DE PÁGINA
            # Márgenes mínimos
            ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0.0, footer=0.0)
            
            # Orientación horizontal (landscape)
            ws.page_setup.orientation = 'landscape'
            
            # Ajustar todas las columnas en una página
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = False
            
            # Añadir métricas de resumen (los datos empiezan en fila 3, después de 2 filas de cabecera)
            summary_row = len(pedidos_filtrados) + 4
            
            # Fusionar celdas para el título
            ws.merge_cells(f'B{summary_row}:C{summary_row}')
            
            # Título del resumen
            title_cell = ws.cell(row=summary_row, column=2, value="METRICAS DE RESUMEN")
            title_cell.font = Font(bold=True, size=12, color="FFFFFF")
            title_cell.fill = self.HEADER_FILL
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Métricas
            metricas_labels = [
                ("Total_Unidades:", int(pedidos_filtrados['Pedido_Final'].sum())),
                ("Total_Articulos:", len(pedidos_filtrados)),
                ("Total_Importe:", f"{pedidos_filtrados['Ventas_Objetivo'].sum():.2f}€"),
                ("Objetivo_Semana:", f"{parametros.get('objetivos_semanales', {}).get(str(semana), 0)}€"),
                ("Factor_Crecimiento:", f"{parametros.get('objetivo_crecimiento', 0.05)*100:.0f}%"),
                ("Factor_Festivo:", f"{parametros.get('festivos', {}).get(str(semana), 0)*100:.0f}%"),
                ("Articulos_A:", len(pedidos_filtrados[pedidos_filtrados['Categoria'] == 'A'])),
                ("Articulos_B:", len(pedidos_filtrados[pedidos_filtrados['Categoria'] == 'B'])),
                ("Articulos_C:", len(pedidos_filtrados[pedidos_filtrados['Categoria'] == 'C'])),
                ("Stock_Minimo_%:", f"{parametros.get('stock_minimo_porcentaje', 0.30)*100:.0f}%"),
                ("Stock_Minimo_Objetivo:", int(pedidos_filtrados['Stock_Minimo_Objetivo'].sum())),
                ("Total_Ajuste_Stock:", int(pedidos_filtrados['Diferencia_Stock'].sum()))
            ]
            
            for i, (label, value) in enumerate(metricas_labels):
                # Columna B (etiquetas)
                cell_label = ws.cell(row=summary_row + 1 + i, column=2, value=label)
                cell_label.font = Font(bold=True, size=10)
                cell_label.border = self.THIN_BLACK_BORDER
                cell_label.alignment = Alignment(horizontal='left', vertical='center')
                
                # Columna C (valores)
                cell_value = ws.cell(row=summary_row + 1 + i, column=3, value=value)
                cell_value.border = self.THIN_BLACK_BORDER
                cell_value.alignment = Alignment(horizontal='center', vertical='center')
            
            # AUTO-AJUSTAR ANCHO DE COLUMNA 21 (Pedido Final)
            # Calcular el ancho máximo necesario basado en el contenido
            max_length_col_u = 0
            for row in range(3, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=21).value
                if cell_value is not None:
                    cell_str = str(cell_value)
                    # Contar caracteres considerando saltos de línea
                    lines = cell_str.split('\n')
                    for line in lines:
                        max_length_col_u = max(max_length_col_u, len(line))
            
            # Añadir un pequeño margen y establecer el ancho
            # El factor 1.2 proporciona un pequeño margen para legibilidad
            optimal_width = min(max(max_length_col_u * 1.2, 8), 30)  # Mínimo 8, máximo 30
            ws.column_dimensions['U'].width = optimal_width
            
            logger.info(f"Ancho de columna U (Pedido Final) ajustado a: {optimal_width:.1f}")

            # Guardar archivo
            wb.save(ruta_completa)
            
            logger.info(f"Archivo guardado: {ruta_completa}")
            return ruta_completa
            
        except Exception as e:
            logger.error(f"Error al generar archivo: {str(e)}")
            # Enviar alerta específica
            alert_svc = get_alert_service()
            if alert_svc:
                alert_svc.alerta_excel_error("archivo_pedido.xlsx", str(e), seccion)
            return None
    
    def generar_resumen_excel(self, resumen_df: pd.DataFrame, seccion: str) -> Optional[str]:
        """
        Genera un archivo Excel con el resumen consolidado de pedidos.
        
        Args:
            resumen_df (pd.DataFrame): DataFrame con el resumen de pedidos
            seccion (str): Nombre de la sección
        
        Returns:
            Optional[str]: Ruta del archivo generado o None si hay error
        """
        if len(resumen_df) == 0:
            logger.warning("No hay datos para generar resumen")
            return None
        
        # Generar nombre del archivo - usar directorio específico para resúmenes
        dir_salida = str(PEDIDOS_SEMANALES_RESUMEN_DIR)
        nombre_archivo = f"Resumen_Pedidos_{seccion}_{datetime.now().strftime('%d%m%Y')}.xlsx"
        ruta_completa = os.path.join(dir_salida, nombre_archivo)
        
        logger.info(f"Generando resumen: {ruta_completa}")
        
        # Crear directorio si no existe
        os.makedirs(dir_salida, exist_ok=True)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumen"
            
            # Configurar columnas
            COLUMN_WIDTHS = {
                'A': 8, 'B': 18, 'C': 16, 'D': 22, 'E': 18, 'F': 12, 'G': 15, 'H': 16, 'I': 14,
                'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 14, 'O': 14, 'P': 14
            }
            
            # Cabeceras
            COLUMN_HEADERS = [
                'Sección', 'Semana', 'Vtas. semana año pasado', 'Objetivo semana', 'Obj. semana + % crec. anual', 
                'Obj. semana + % crec. + Festivos', '% Obj. crecim. + Festivos', 'Total Unidades', 'Total Articulos', 
                'Total Importe', 'Alcance %', 'Articulos A', 'Articulos B', 'Articulos C',
                '% Festivo', '% Stock Min', 'Stock Min Obj'
            ]
            
            COLUMN_MAPPING = {
                'Seccion': 'Sección',
                'Semana': 'Semana',
                'Vtas. semana año pasado': 'Vtas. semana año pasado',
                'Objetivo_Semana': 'Objetivo semana',
                'Obj. semana + % crec. anual': 'Obj. semana + % crec. anual',
                'Obj. semana + % crec. + Festivos': 'Obj. semana + % crec. + Festivos',
                '% Obj. crecim. + Festivos': '% Obj. crecim. + Festivos',
                'Total_Unidades': 'Total Unidades',
                'Total_Articulos': 'Total Articulos',
                'Total_Importe': 'Total Importe',
                'Alcance_Objetivo_%': 'Alcance %',
                'Articulos_A': 'Articulos A',
                'Articulos_B': 'Articulos B',
                'Articulos_C': 'Articulos C',
                'Incremento_Festivo_%': '% Festivo',
                'Stock_Minimo_%': '% Stock Min',
                'Stock_Minimo_Objetivo': 'Stock Min Obj'
            }
            
            # Escribir cabeceras (fila 2)
            for col_idx, header in enumerate(COLUMN_HEADERS, 1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.THIN_BLACK_BORDER
            
            # Renombrar y reordenar columnas
            resumen_renamed = resumen_df.rename(columns=COLUMN_MAPPING)
            resumen_renamed = resumen_renamed[COLUMN_HEADERS]
            
            # Escribir datos (empezar en fila 3)
            for r_idx, row in enumerate(dataframe_to_rows(resumen_renamed, index=False, header=False), 3):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = self.THIN_BLACK_BORDER
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    header_name = COLUMN_HEADERS[c_idx - 1]
                    
                    # Formatear según tipo de dato
                    if header_name in ['Vtas. semana año pasado', 'Objetivo semana', 
                                      'Obj. semana + % crec. anual', 
                                      'Obj. semana + % crec. + Festivos', 
                                      'Total Importe', 'Stock Min Obj']:
                        if isinstance(value, (int, float)):
                            cell.value = int(round(value)) if value == int(value) else int(value)
                        cell.number_format = '#,##0'
                    
                    elif header_name in ['% Obj. crecim. + Festivos', 'Alcance %', '% Festivo', '% Stock Min']:
                        if isinstance(value, (int, float)):
                            cell.value = round(value, 1)
                        cell.number_format = '0.0'
            
            # Aplicar anchuras de columna
            for col_letter, width in COLUMN_WIDTHS.items():
                ws.column_dimensions[col_letter].width = width
            
            # Añadir título en fila 1
            ws.merge_cells('A1:P1')
            titulo = f"RESUMEN DE PEDIDOS DE COMPRA - {seccion.upper()} - VIVEVERDE 2026"
            ws['A1'] = titulo
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = self.HEADER_FILL
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Guardar archivo
            wb.save(ruta_completa)
            
            logger.info(f"Resumen guardado: {ruta_completa}")
            return ruta_completa
            
        except Exception as e:
            logger.error(f"Error al generar resumen: {str(e)}")
            return None
    
    def generar_archivo_csv(self, pedidos_df: pd.DataFrame, semana: int,
                            seccion: str) -> Optional[str]:
        """
        Genera un archivo CSV con los datos del pedido.
        
        Args:
            pedidos_df (pd.DataFrame): DataFrame con los artículos y cantidades
            semana (int): Número de semana
            seccion (str): Nombre de la sección
        
        Returns:
            Optional[str]: Ruta del archivo generado o None si hay error
        """
        if len(pedidos_df) == 0:
            return None
        
        # Filtrar artículos con Pedido_Corregido_Stock > 0
        pedidos_filtrados = pedidos_df[pedidos_df['Pedido_Corregido_Stock'] > 0].copy()
        
        if len(pedidos_filtrados) == 0:
            return None
        
        # Generar nombre del archivo
        dir_salida = self.obtener_directorio_salida()
        nombre_archivo = f"Pedido_Semana_{semana:02d}_{seccion}_{datetime.now().strftime('%d%m%Y')}.csv"
        ruta_completa = os.path.join(dir_salida, nombre_archivo)
        
        try:
            # Seleccionar columnas relevantes para CSV
            columnas_csv = [
                'Codigo_Articulo', 'Nombre_Articulo', 'Talla', 'Color',
                'Pedido_Corregido_Stock', 'PVP', 'Coste_Pedido', 'Proveedor', 'Categoria'
            ]
            
            pedidos_csv = pedidos_filtrados[columnas_csv].copy()
            pedidos_csv.to_csv(ruta_completa, index=False, encoding='utf-8-sig')
            
            logger.info(f"CSV guardado: {ruta_completa}")
            return ruta_completa
            
        except Exception as e:
            logger.error(f"Error al generar CSV: {str(e)}")
            return None


# Funciones de utilidad para uso directo
def crear_order_generator(config: dict) -> OrderGenerator:
    """
    Crea una instancia del OrderGenerator.
    
    Args:
        config (dict): Configuración del sistema
    
    Returns:
        OrderGenerator: Instancia inicializada del generador de pedidos
    """
    return OrderGenerator(config)


if __name__ == "__main__":
    # Ejemplo de uso
    print("OrderGenerator - Módulo de generación de archivos de salida")
    print("=" * 50)
    
    # Configurar logging básico
    logging.basicConfig(level=logging.INFO)
    
    # Ejemplo de configuración
    config_ejemplo = {
        'rutas': {
            'directorio_base': '.',
            'directorio_salida': str(PEDIDOS_SEMANALES_DIR)
        },
        'formato_salida': {
            'prefijo_archivo': 'Pedido_Semana',
            'incluir_fecha_en_nombre': True
        }
    }
    
    # Crear OrderGenerator
    generator = crear_order_generator(config_ejemplo)
    
    # Ejemplo de generación de nombre
    nombre = generator.generar_nombre_archivo(10, 'vivero')
    print(f"Nombre de archivo generado: {nombre}")
    
    print("\nOrderGenerator listo para usar.")
