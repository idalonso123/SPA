#!/usr/bin/env python3
"""
Motor de Cálculo ABC+D para Gestión de Inventarios - VERSIÓN DINÁMICA
Jardinería Aranjuez - Período y año configurables automáticamente o mediante argumentos

Este script combina:
- Cálculo de clasificación ABC+D
- Aplicación de formatos Excel
- Lógica corregida de riesgo
- Período y año configurables mediante argumentos
- Filtrado automático de datos según el período indicado
- Procesamiento de múltiples secciones
- Envío automático de emails a los encargados de cada sección

MODO DE USO:
- Sin parámetros: Procesa automáticamente según la fecha actual
  * Detecta el período y año actual
  * Analiza los datos del año anterior
  * Genera archivos para el período SIGUIENTE al actual
  * Ejemplo: Si hoy es febrero 2026 (P1), analiza datos de 2025 y genera archivos para P2_2025

- Con parámetros:
  * -P <periodo>: Período específico (P1, P2, P3, P4)
  * -Y <año>: Año de los datos a analizar
  * -S <sección>: Sección específica a procesar
  
  Ejemplo: python clasificacionABC.py -P P3 -Y 2025 -S maf

Ejecutar: 
    python clasificacionABC.py                              # Modo automático
    python clasificacionABC.py -P P1 -Y 2025               # Período P1 del año 2025
    python clasificacionABC.py -S maf                       # Solo sección maf (modo automático)
    python clasificacionABC.py -P P2 -Y 2025 -S vivero     # Período P2 de 2025, solo vivero

Los datos se leen de archivos con datos de TODO el año:
- SPA_compras.xlsx: Datos de compras de todo el año
- SPA_ventas.xlsx: Datos de ventas de todo el año
- SPA_stock_{periodo}.xlsx: Datos de stock actual
- SPA_coste.xlsx: Costes unitarios de artículos (para calcular beneficio real)

El script filtra automáticamente los datos según las fechas del período indicado.
Al generar cada archivo de clasificación, se envía automáticamente un email
al encargado de la sección con el archivo adjunto.
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
import sys
import argparse
import warnings
import smtplib
import ssl
import os
import json
import unicodedata
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from pathlib import Path
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURACIÓN DE RUTAS - DIRECTORIO BASE DEL SCRIPT
# ============================================================================

# Determinar el directorio base del script (donde está clasificacionABC.py)
DIRECTORIO_BASE = os.path.dirname(os.path.abspath(__file__))
DIRECTORIO_DATA = os.path.join(DIRECTORIO_BASE, 'data', 'input')
DIRECTORIO_CONFIG = os.path.join(DIRECTORIO_BASE, 'config')

# ============================================================================
# FUNCIONES DE NORMALIZACIÓN PARA BÚSQUEDAS INTELIGENTES
# ============================================================================

def normalizar_texto(texto):
    """
    Normaliza un texto para comparación:
    - Convierte a minúsculas
    - Elimina acentos
    - Elimina puntuación (puntos, guiones, espacios, paréntesis, etc.)
    
    Esta función es la base para todas las búsquedas normalizadas en el script.
    
    Ejemplos:
    - 'Cóste' -> 'coste'
    - 'Últ. Comp' -> 'ultcomp'
    - 'ÚLTIMA COMPRA' -> 'ultimacompra'
    - 'Coste Unitario' -> 'costeunitario'
    
    Args:
        texto: Texto a normalizar
    
    Returns:
        str: Texto normalizado (minúsculas, sin acentos, sin puntuación) o cadena vacía si es None/NaN
    """
    if pd.isna(texto):
        return ''
    texto = str(texto)
    # Convertir a minúsculas
    texto = texto.lower()
    # Normalizar unicode: á → a, é → e, ñ → n, etc.
    texto = unicodedata.normalize('NFD', texto)
    texto = texto.encode('ascii', 'ignore').decode('utf-8')
    # Eliminar puntuación: puntos, guiones, espacios, paréntesis, etc.
    texto = ''.join(c for c in texto if c.isalnum())
    return texto

def normalizar_con_espacios(texto):
    """
    Normaliza un texto conservando los espacios pero eliminando:
    - Mayúsculas/minúsculas
    - Acentos
    - Puntuación (excepto espacios)
    
    Útil cuando quieres mantener la estructura de palabras pero ignorar detalles.
    
    Ejemplos:
    - 'Cóste' -> 'coste'
    - 'Últ. Comp' -> 'ult comp'
    - 'ÚLTIMA COMPRA' -> 'ultima compra'
    
    Args:
        texto: Texto a normalizar
    
    Returns:
        str: Texto normalizado (minúsculas, sin acentos,保留 espacios, sin otra puntuación)
    """
    if pd.isna(texto):
        return ''
    texto = str(texto)
    # Convertir a minúsculas
    texto = texto.lower()
    # Normalizar unicode: á → a, é → e, ñ → n, etc.
    texto = unicodedata.normalize('NFD', texto)
    texto = texto.encode('ascii', 'ignore').decode('utf-8')
    # Reemplazar puntuación (excepto espacios) por nada
    texto = ''.join(c if c.isalnum() or c == ' ' else '' for c in texto)
    # Normalizar espacios múltiples
    texto = ' '.join(texto.split())
    return texto

def encontrar_columna(columnas, nombre_buscado):
    """
    Busca una columna por nombre, ignorando mayúsculas, acentos y puntuación.
    
    Esta función permite encontrar columnas aunque sus nombres tengan variaciones:
    - 'Coste', 'COSTE', 'cósté', 'Cóste' → todas dan positivo para 'coste'
    - 'Últ. compra', 'ULTIMA COMPRA', 'ultima compra' → todas dan positivo para 'ultimacompra'
    - 'Últ. Comp', 'Ult. Comp', 'últ comp' → todas dan positivo para 'ultcomp'
    
    Args:
        columnas: Lista de nombres de columnas del DataFrame
        nombre_buscado: Nombre base a buscar (puede tener acentos, mayúsculas, etc.)
    
    Returns:
        str: El nombre real de la columna encontrada, o None si no existe
    """
    nombre_normalizado = normalizar_texto(nombre_buscado)
    
    for columna in columnas:
        if normalizar_texto(columna) == nombre_normalizado:
            return columna
    
    return None

def obtener_columna_segura(df, nombre_buscado):
    """
    Obtiene una columna del DataFrame buscando por nombre normalizado.
    
    Args:
        df: DataFrame con los datos
        nombre_buscado: Nombre base a buscar (ej: 'coste', 'fecha', 'últ. comp')
    
    Returns:
        Series: La columna encontrada, o una serie vacía si no existe
    """
    nombre_real = encontrar_columna(list(df.columns), nombre_buscado)
    if nombre_real:
        return df[nombre_real]
    else:
        print(f"ADVERTENCIA: No se encontró columna '{nombre_buscado}' en el DataFrame")
        return pd.Series([], dtype='object')

def filtrar_por_valor_normalizado(df, nombre_columna, valor_buscado):
    """
    Filtra un DataFrame buscando un valor en una columna específica,
    ignorando mayúsculas, acentos y puntuación.
    
    Esta función es ideal para búsquedas en columnas de texto donde los valores
    pueden tener variaciones en su escritura.
    
    Ejemplos de uso:
    - Buscar 'coste' en columna 'Nombre': encuentra 'Coste', 'COSTE', 'cósté', etc.
    - Buscar 'últ. comp' en columna 'Concepto': encuentra 'Últ. Comp', 'Ult Comp', etc.
    
    Args:
        df: DataFrame con los datos
        nombre_columna: Nombre de la columna donde buscar (se normaliza la búsqueda)
        valor_buscado: Valor a buscar (puede tener variaciones de caso, acentos, etc.)
    
    Returns:
        DataFrame: Filas que coinciden con el valor buscado
    """
    # Normalizar el valor buscado
    valor_normalizado = normalizar_texto(valor_buscado)
    
    # Encontrar la columna (si no existe, retornar DataFrame vacío)
    columna_real = encontrar_columna(list(df.columns), nombre_columna)
    if columna_real is None:
        print(f"ADVERTENCIA: No se encontró la columna '{nombre_columna}' para filtrar")
        return df.iloc[0:0]
    
    # Filtrar comparando valores normalizados
    mask = df[columna_real].apply(lambda x: normalizar_texto(x) == valor_normalizado)
    return df[mask]

def filtrar_por_valor_parcial(df, nombre_columna, fragmento_buscado):
    """
    Filtra un DataFrame buscando un fragmento de texto en una columna,
    ignorando mayúsculas, acentos y puntuación.
    
    Útil cuando quieres encontrar valores que CONTENGAN el texto buscado,
    no solo valores que sean EXACTAMENTE iguales.
    
    Ejemplos de uso:
    - Buscar 'coste' en columna 'Nombre': encuentra cualquier nombre que contenga 'coste'
    - Buscar 'últ' en columna 'Concepto': encuentra 'Últ. Comp', 'Última', 'Último', etc.
    
    Args:
        df: DataFrame con los datos
        nombre_columna: Nombre de la columna donde buscar
        fragmento_buscado: Fragmento de texto a buscar
    
    Returns:
        DataFrame: Filas que contienen el fragmento buscado
    """
    # Normalizar el fragmento buscado
    fragmento_normalizado = normalizar_texto(fragmento_buscado)
    
    # Encontrar la columna
    columna_real = encontrar_columna(list(df.columns), nombre_columna)
    if columna_real is None:
        print(f"ADVERTENCIA: No se encontró la columna '{nombre_columna}' para filtrar")
        return df.iloc[0:0]
    
    # Filtrar buscando el fragmento en valores normalizados
    mask = df[columna_real].apply(lambda x: fragmento_normalizado in normalizar_texto(x))
    return df[mask]

def buscar_valores_unicos_normalizados(df, nombre_columna):
    """
    Obtiene los valores únicos de una columna, junto con su versión normalizada.
    
    Útil para analizar qué valores únicos existen en una columna y poder
    hacer búsquedas normalizadas posteriormente.
    
    Args:
        df: DataFrame con los datos
        nombre_columna: Nombre de la columna a analizar
    
    Returns:
        dict: Diccionario {valor_normalizado: [lista de valores originales]}
    """
    columna_real = encontrar_columna(list(df.columns), nombre_columna)
    if columna_real is None:
        print(f"ADVERTENCIA: No se encontró la columna '{nombre_columna}'")
        return {}
    
    resultado = {}
    for valor in df[columna_real].unique():
        clave = normalizar_texto(valor)
        if clave not in resultado:
            resultado[clave] = []
        if valor not in resultado[clave]:
            resultado[clave].append(valor)
    
    return resultado

# ============================================================================
# CONFIGURACIÓN DE SECCIONES
# ============================================================================

# Códigos de animales vivos - Se cargan desde config_comun.json
# La variable se inicializa después de CONFIG
CODIGOS_MASCOTAS_VIVO = ['2104', '2204', '2305', '2405', '2504', '2606', '2705', '2707', '2708', '2805', '2806', '2906']  # Valor por defecto

def obtener_codigos_mascotas_vivo():
    """
    Obtiene los códigos de mascotas vivos desde config_comun.json
    
    Returns:
        list: Lista de códigos de mascotas vivos
    """
    global CODIGOS_MASCOTAS_VIVO
    
    # Intentar cargar desde CONFIG (definida más adelante en el archivo)
    try:
        if 'CONFIG' in globals() and CONFIG and 'configuracion_mascotas' in CONFIG:
            codigos = CONFIG['configuracion_mascotas'].get('codigos_mascotas_vivo', [])
            if codigos:
                CODIGOS_MASCOTAS_VIVO = codigos
                print(f"INFO: Códigos de mascotas vivos cargados desde config: {len(CODIGOS_MASCOTAS_VIVO)} códigos")
    except:
        pass
    
    return CODIGOS_MASCOTAS_VIVO

# Definición de todas las secciones - Se cargará desde config_comun.json
SECCIONES = {
    'interior': {
        'descripcion': 'Plantas de interior',
        'rangos': [{'tipo': 'prefijos', 'valores': ['1']}]
    },
    'utiles_jardin': {
        'descripcion': 'Útiles de jardín',
        'rangos': [{'tipo': 'prefijos', 'valores': ['4']}]
    },
    'semillas': {
        'descripcion': 'Semillas y bulbos',
        'rangos': [{'tipo': 'prefijos', 'valores': ['5']}]
    },
    'deco_interior': {
        'descripcion': 'Decoración interior',
        'rangos': [{'tipo': 'prefijos', 'valores': ['6']}]
    },
    'maf': {
        'descripcion': 'Planta de temporada y floristería',
        'rangos': [{'tipo': 'prefijos', 'valores': ['7']}]
    },
    'vivero': {
        'descripcion': 'Vivero y plantas exterior',
        'rangos': [{'tipo': 'prefijos', 'valores': ['8']}]
    },
    'deco_exterior': {
        'descripcion': 'Decoración exterior',
        'rangos': [{'tipo': 'prefijos', 'valores': ['9']}]
    },
    'mascotas_manufacturado': {
        'descripcion': 'Mascotas (productos manufacturados)',
        'rangos': [
            {'tipo': 'prefijos', 'valores': ['2'], 'excluir': CODIGOS_MASCOTAS_VIVO}
        ]
    },
    'mascotas_vivo': {
        'descripcion': 'Mascotas (animales vivos)',
        'rangos': [
            {'tipo': 'codigos_exactos', 'valores': CODIGOS_MASCOTAS_VIVO}
        ]
    },
    'tierra_aridos': {
        'descripcion': 'Tierras y áridos',
        'rangos': [
            {'tipo': 'prefijos', 'valores': ['31', '32']}
        ]
    },
    'fitos': {
        'descripcion': 'Fitosanitarios y abonos',
        'rangos': [
            {'tipo': 'rango', 'valores': ['33', '34', '35', '36', '37', '38', '39']}
        ]
    }
}

# ============================================================================
# CONFIGURACIÓN DE FECHAS - PERÍODO CONFIGURABLE MEDIANTE ARGUMENTOS
# ============================================================================

# Valores por defecto (se sobrescribirán si se especifica un argumento de período)
FECHA_INICIO = None  # Se configurará según el argumento --P1, --P2, --P3, --P4
FECHA_FIN = None     # Se configurará según el argumento --P1, --P2, --P3, --P4
DIAS_PERIODO = 0
PERIODO = "PERIODO"

# ============================================================================
# CARGA DE CONFIGURACIÓN DESDE JSON
# ============================================================================

def cargar_configuracion():
    """
    Carga la configuración desde config/config_comun.json
    
    Returns:
        dict: Configuración cargada o None si hay error
    """
    try:
        ruta_config = os.path.join(DIRECTORIO_CONFIG, 'config_comun.json')
        if os.path.exists(ruta_config):
            with open(ruta_config, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            print(f"ADVERTENCIA: No se encontró {ruta_config}. Usando configuración por defecto.")
            return None
    except Exception as e:
        print(f"ERROR al cargar configuración: {e}. Usando valores por defecto.")
        return None

def obtener_fechas_periodo(periodo_nombre, config, año_actual=None):
    """
    Obtiene las fechas de inicio y fin de un período desde la configuración.
    
    Args:
        periodo_nombre: Nombre del período (P1, P2, P3, P4)
        config: Configuración cargada desde JSON
        año_actual: Año a utilizar (por defecto el año actual)
    
    Returns:
        tuple: (fecha_inicio, fecha_fin) como objetos datetime, o (None, None) si no se encuentra
    """
    if año_actual is None:
        año_actual = datetime.now().year
    
    if config and 'configuracion_periodo_clasificacion' in config and 'periodos' in config['configuracion_periodo_clasificacion']:
        periodos = config['configuracion_periodo_clasificacion']['periodos']
        
        if periodo_nombre in periodos:
            periodo = periodos[periodo_nombre]
            mes_inicio = periodo['mes_inicio']
            dia_inicio = periodo['dia_inicio']
            mes_fin = periodo['mes_fin']
            dia_fin = periodo['dia_fin']
            
            fecha_inicio = datetime(año_actual, mes_inicio, dia_inicio)
            fecha_fin = datetime(año_actual, mes_fin, dia_fin)
            
            return fecha_inicio, fecha_fin
    
    return None, None

def obtener_periodo_desde_fecha(fecha, config):
    """
    Determina el período (P1, P2, P3, P4) basándose en una fecha usando la configuración.
    
    Períodos definidos en config:
    - P1: 1 enero a 28 de febrero
    - P2: 1 marzo a 31 mayo
    - P3: 1 junio a 31 de agosto
    - P4: 1 de septiembre a 31 de diciembre
    
    Args:
        fecha: Objeto datetime con la fecha a verificar
        config: Configuración cargada desde JSON
    
    Returns:
        str: Período (P1, P2, P3 o P4)
    """
    if config and 'configuracion_periodo_clasificacion' in config and 'periodos' in config['configuracion_periodo_clasificacion']:
        periodos = config['configuracion_periodo_clasificacion']['periodos']
        
        mes = fecha.month
        dia = fecha.day
        
        for key, periodo in periodos.items():
            mes_inicio = periodo['mes_inicio']
            dia_inicio = periodo['dia_inicio']
            mes_fin = periodo['mes_fin']
            dia_fin = periodo['dia_fin']
            
            # Comprobar si el mes está dentro del rango del período
            if mes_inicio <= mes <= mes_fin:
                # Si es el mes de inicio, verificar que el día sea >= día de inicio
                if mes == mes_inicio and dia < dia_inicio:
                    continue
                # Si es el mes de fin, verificar que el día sea <= día de fin
                if mes == mes_fin and dia > dia_fin:
                    continue
                return periodo['nombre']
        
        # Si no coincide ningún período completo, buscar el más cercano
        if mes < 3:
            return "P1"
        elif mes < 6:
            return "P2"
        elif mes < 9:
            return "P3"
        else:
            return "P4"
    
    # Fallback: cálculo por defecto si no hay configuración
    mes = fecha.month
    if mes <= 2:
        return "P1"
    elif mes <= 5:
        return "P2"
    elif mes <= 8:
        return "P3"
    else:
        return "P4"


def obtener_periodo_siguiente(periodo_actual):
    """
    Obtiene el período siguiente al actual.
    
    Args:
        periodo_actual: Período actual (P1, P2, P3, P4)
    
    Returns:
        str: Período siguiente (P2->P3, P3->P4, P4->P1, P1->P2)
    """
    orden_periodos = ["P1", "P2", "P3", "P4"]
    idx = orden_periodos.index(periodo_actual)
    return orden_periodos[(idx + 1) % 4]


def configurar_periodo(periodo_seleccionado, config, año_datos=None):
    """
    Configura las variables globales de período basándose en el argumento proporcionado.
    
    Args:
        periodo_seleccionado: Período seleccionado (P1, P2, P3, P4) o None
        config: Configuración cargada desde JSON
        año_datos: Año de los datos (si no se especifica, usa el año actual)
    
    Returns:
        tuple: (FECHA_INICIO, FECHA_FIN, DIAS_PERIODO, PERIODO, año_datos)
    """
    if año_datos is None:
        año_datos = datetime.now().year
    
    if periodo_seleccionado:
        # Usar el período especificado como argumento
        fecha_inicio, fecha_fin = obtener_fechas_periodo(periodo_seleccionado, config, año_datos)
        if fecha_inicio and fecha_fin:
            dias_periodo = (fecha_fin - fecha_inicio).days + 1
            return fecha_inicio, fecha_fin, dias_periodo, periodo_seleccionado, año_datos
        else:
            print(f"ADVERTENCIA: Período '{periodo_seleccionado}' no encontrado en configuración. Usando valores por defecto.")
    
    # Valores por defecto si no se especifica período o no se encuentra
    fecha_inicio = datetime(año_datos, 1, 1)
    fecha_fin = datetime(año_datos, 12, 31)
    dias_periodo = (fecha_fin - fecha_inicio).days + 1
    return fecha_inicio, fecha_fin, dias_periodo, "ANUAL", año_datos

def detectar_año_datos(compras_df, ventas_df):
    """
    Detecta automáticamente el año de los datos basándose en las fechas de compras y ventas.
    
    Args:
        compras_df: DataFrame de compras
        ventas_df: DataFrame de ventas
    
    Returns:
        int: Año detectado o año actual si no se puede detectar
    """
    try:
        # Intentar detectar desde ventas
        if len(ventas_df) > 0:
            ventas_df['Fecha'] = pd.to_datetime(ventas_df['Fecha'], errors='coerce')
            fechas_validas = ventas_df['Fecha'].dropna()
            if len(fechas_validas) > 0:
                año_detectado = fechas_validas.max().year
                print(f"INFO: Año detectado desde VENTAS: {año_detectado}")
                return año_detectado
        
        # Intentar detectar desde compras
        if len(compras_df) > 0:
            compras_df['Fecha'] = pd.to_datetime(compras_df['Fecha'], errors='coerce')
            fechas_validas = compras_df['Fecha'].dropna()
            if len(fechas_validas) > 0:
                año_detectado = fechas_validas.max().year
                print(f"INFO: Año detectado desde COMPRAS: {año_detectado}")
                return año_detectado
    except Exception as e:
        print(f"ADVERTENCIA: No se pudo detectar el año de los datos: {e}")
    
    # Fallback: usar el año actual
    año_actual = datetime.now().year
    print(f"INFO: Usando año actual por defecto: {año_actual}")
    return año_actual

# Cargar configuración al inicio
CONFIG = cargar_configuracion()

# Actualizar códigos de mascotas vivos desde CONFIG
if CONFIG and 'configuracion_mascotas' in CONFIG:
    codigos = CONFIG['configuracion_mascotas'].get('codigos_mascotas_vivo', [])
    if codigos:
        CODIGOS_MASCOTAS_VIVO = codigos
        print(f"INFO: Códigos de mascotas vivos cargados desde config: {len(CODIGOS_MASCOTAS_VIVO)} códigos")

# ============================================================================
# CONFIGURACIÓN DE FORMATOS EXCEL
# ============================================================================

COLORES_RIESGO = {
    'Bajo': '90EE90',      # Verde claro
    'Medio': 'FFFF00',     # Amarillo
    'Alto': 'FFA500',      # Naranja
    'Crítico': 'FF6B6B',   # Rojo claro
    'Cero': '90EE90',      # Verde para riesgo cero
}

COLOR_CABECERA = '008000'
COLOR_TEXTO_CABECERA = 'FFFFFF'

# ============================================================================
# TABLA DE ROTACIONES POR FAMILIA
# ============================================================================

ROTACIONES_FAMILIA = {
    # Plantas (2 dígitos)
    '11': ('PLANTAS VERDES', 30),
    '12': ('ORQUIDEAS', 15),
    '13': ('PLANTAS DE FLOR', 15),
    '14': ('FLOR CORTADA', 7),
    '15': ('CACTUS', 30),
    '16': ('COMPOSICIONES', 30),
    '17': ('BONSAIS', 30),
    
    # Animales - familias con 4 dígitos (empiezan por 2)
    '2101': ('ALIMENTACION PAJARO', 60),
    '2102': ('JAULAS PAJARO', 60),
    '2103': ('HIGIENE/CUIDADO PAJARO', 60),
    '2104': ('ANIMAL VIVO PAJARO', 30),
    '2201': ('ALIMENTACION PEQUEÑOS MAMIFEROS', 60),
    '2202': ('JAULAS PEQUEÑOS MAMIFEROS', 60),
    '2203': ('HIGIENE/CUIDADO PEQUEÑOS MAMIFEROS', 60),
    '2204': ('ANIMAL VIVO PEQUEÑOS MAMIFEROS', 30),
    '2301': ('ALIMENTACION PERRO', 60),
    '2302': ('CONFORT PERRO', 60),
    '2303': ('CORREAS Y COLLARES PERRO', 60),
    '2304': ('HIGIENE/CUIDADO PERRO', 60),
    '2305': ('ANIMAL VIVO PERRO', 30),
    '2401': ('ALIMENTACION GATO', 60),
    '2402': ('CONFORT GATO', 60),
    '2403': ('CORREAS Y COLLARES GATO', 60),
    '2404': ('HIGIENE/CUIDADO GATO', 60),
    '2405': ('ANIMAL VIVO GATO', 30),
    '2501': ('ALIMENTACION ANIMALES DE GRANJA', 60),
    '2502': ('HABITAT / ACCES ANIMALES DE GRANJA', 60),
    '2503': ('HIGIENE/CUIDADO ANIMALES DE GRANJA', 60),
    '2504': ('ANIMAL VIVO GRANJA', 30),
    '2601': ('ALIMENTACION REPTILES', 60),
    '2602': ('TERRARIO REPTILES', 60),
    '2603': ('ACCESORIOS REPTILES', 60),
    '2604': ('DECO INERTE REPTILES', 60),
    '2605': ('HIGIENE/CUIDADO REPTILES', 60),
    '2606': ('ANIMAL VIVO REPTILES', 30),
    '2701': ('ALIMENTACION ACUARIOFILIA', 60),
    '2702': ('ACUARIOS', 60),
    '2703': ('ACCESORIOS ACUARIOFILIA', 60),
    '2704': ('DECO INERTE ACUARIOFILIA', 60),
    '2705': ('PLANTA ACUATICA DECORACION ACUARIOFILIA', 15),
    '2706': ('HIGIENE/CUIDADO ACUARIOFILIA', 60),
    '2707': ('PECES AGUA CALIENTE ACUARIOFILIA', 15),
    '2708': ('PECES AGUA FRIA ACUARIOFILIA', 15),
    '2709': ('AGUA OSMOSIS ACUARIOFILIA', 60),
    '2801': ('ALIMENTACION JARDIN ACUATICO', 60),
    '2802': ('ACCESORIOS JARDIN ACUATICO', 60),
    '2803': ('HIGIENE/CUIDADO JARDIN ACUATICO', 60),
    '2804': ('DECORACION JARDIN ACUATICO', 60),
    '2805': ('PLANTAS JARDIN ACUATICO', 30),
    '2806': ('PECES JARDIN ACUATICO', 15),
    '2906': ('INSECTO VIVO', 15),
    
    # Mantenimiento/tratamiento/cuidados (2 dígitos)
    '31': ('TIERRAS', 90),
    '32': ('MANTENIMIENTO', 90),
    '33': ('ABONOS', 90),
    '34': ('ABONO NATURAL', 90),
    '35': ('FITOSANITARIOS', 90),
    '36': ('FITOSANITARIO NATURAL', 90),
    '37': ('HERBICIDAS', 90),
    '39': ('ANTI-PLAGAS', 90),

    # Utiles jardin (2 dígitos)
    '41': ('UTILES JARDIN', 90),
    '42': ('PODA', 90),
    '43': ('PULVERIZACION', 90),
    '44': ('PROTECCION CULTIVO', 90),
    '45': ('PROTECCION PERSONAL', 90),
    '46': ('RIEGO', 90),
    '48': ('OTRAS MAQUINAS MOTOR', 90),
    '49': ('ACCESS/PIEZAS', 90),

    # Semillas (2 dígitos)
    '51': ('BULBOS DE FLOR', 60),
    '53': ('CESPED', 60),
    '54': ('SEMILLAS', 60),

    # Decoracion casa (2 dígitos)
    '61': ('DECORACION NAVIDAD', 90),
    '62': ('DECORACION FLORAL', 90),
    '63': ('RECIPIENTES', 90),
    '64': ('DECORACION AMBIENTE', 90),
    '65': ('LIB/PAP/SON/IMAG.', 90),

    # Planta de temporada (2 dígitos)
    '71': ('PLANTAS PARA MACIZOS EN BDJA.', 15),
    '72': ('PLANTAS PARA MACIZOS EN MAC.', 15),
    '74': ('VIVACES EN MACETA', 15),
    '75': ('PLANTAS TRADICIONALES', 15),
    '77': ('PELARGONIUM EN MACETA', 15),
    '78': ('AROMATICAS', 15),
    '79': ('FRESALES/HORTICOLAS', 15),

    # Vivero (2 dígitos)
    '81': ('ARBOLES/ARBUSTOS DECO', 30),
    '82': ('CONIFERAS', 30),
    '83': ('ROSALES', 30),
    '84': ('FRUTALES', 30),
    '85': ('PLANTAS TIERRA DE BREZO', 30),
    '86': ('PLANTAS PARA SETOS', 30),
    '87': ('PLANTAS TREPADORAS', 30),
    '88': ('PLANTAS CLIMA MEDITERRANEO', 30),
    '89': ('ABETOS NAVIDAD', 30),

    # Decoracion exterior (2 dígitos)
    '91': ('MOBILIARIO JARDIN', 90),
    '92': ('EQUIP. JARDIN', 90),
    '93': ('AIRE LIBRE', 90),
    '94': ('MACETERIA/SOPORTES', 90),
    '95': ('DECORACION', 90),
    '96': ('COBERTIZOS', 90),
    '97': ('CERRAMIENTOS/SOMBREO', 90),
}

# ============================================================================
# TABLA DE IVA POR FAMILIA (2 dígitos)
# ============================================================================

IVA_FAMILIA = {
    # Plantas (IVA 10%)
    '11': 10,  # PLANTAS VERDES
    '12': 10,  # ORQUIDEAS
    '13': 10,  # PLANTAS DE FLOR
    '14': 10,  # FLOR CORTADA
    '15': 10,  # CACTUS
    '16': 10,  # COMPOSICIONES
    '17': 10,  # BONSAIS
    '18': 10,  # MUGUET
    
    # Mantenimiento/tratamiento/cuidados (IVA 21%)
    '31': 21,  # TIERRAS
    '32': 21,  # MANTENIMIENTO
    '33': 21,  # ABONOS
    '34': 21,  # ABONO NATURAL
    '35': 21,  # FITOSANITARIOS
    '36': 21,  # FITOSANITARIO NATURAL
    '37': 21,  # HERBICIDAS
    '38': 21,  # HERBICIDA NATURAL
    '39': 21,  # ANTI-PLAGAS

    # Utiles jardin (IVA 21%)
    '41': 21,  # UTILES JARDIN
    '42': 21,  # PODA
    '43': 21,  # PULVERIZACION
    '44': 21,  # PROTECCION CULTIVO
    '45': 21,  # PROTECCION PERSONAL
    '46': 21,  # RIEGO
    '47': 21,  # CORTACESPEDES
    '48': 21,  # OTRAS MAQUINAS MOTOR
    '49': 21,  # ACCESS/PIEZAS

    # Semillas (IVA 10%)
    '51': 10,  # BULBOS DE FLOR
    '52': 10,  # BULBOS DE HORTICOLAS
    '53': 10,  # CESPED
    '54': 10,  # SEMILLAS
    '55': 10,  # PATATAS

    # Decoracion casa (IVA 21%)
    '61': 21,  # DECORACION NAVIDAD
    '62': 21,  # DECORACION FLORAL
    '63': 21,  # RECIPIENTES
    '64': 21,  # DECORACION AMBIENTE
    '65': 21,  # LIB/PAP/SON/IMAG.
    '66': 21,  # CHIMENEAS
    '67': 21,  # ALIMENTACION

    # Planta de temporada (IVA 10%)
    '71': 10,  # PLANTAS PARA MACIZOS EN BDJA.
    '72': 10,  # PLANTAS PARA MACIZOS EN MAC.
    '73': 10,  # VIVACES EN BANDEJA
    '74': 10,  # VIVACES EN MACETA
    '75': 10,  # PLANTAS TRADICIONALES
    '76': 10,  # PELARGONIUM EN BANDEJA
    '77': 10,  # PELARGONIUM EN MACETA
    '78': 10,  # AROMATICAS
    '79': 10,  # FRESALES/HORTICOLAS

    # Vivero (IVA 10%)
    '81': 10,  # ARBOLES/ARBUSTOS DECO
    '82': 10,  # CONIFERAS
    '83': 10,  # ROSALES
    '84': 10,  # FRUTALES
    '85': 10,  # PLANTAS TIERRA DE BREZO
    '86': 10,  # PLANTAS PARA SETOS
    '87': 10,  # PLANTAS TREPADORAS
    '88': 10,  # PLANTAS CLIMA MEDITERRANEO
    '89': 10,  # ABETOS NAVIDAD

    # Decoracion exterior (IVA 21%)
    '91': 21,  # MOBILIARIO JARDIN
    '92': 21,  # EQUIP. JARDIN
    '93': 21,  # AIRE LIBRE
    '94': 21,  # MACETERIA/SOPORTES
    '95': 21,  # DECORACION
    '96': 21,  # COBERTIZOS
    '97': 21,  # CERRAMIENTOS/SOMBREO
}

# ============================================================================
# TABLA DE IVA POR SUBFAMILIA (4 dígitos, empiezan por 2)
# ============================================================================

IVA_SUBFAMILIA = {
    # Familia 21 - MUGUET
    '2101': 10,  # ALIMENTACION
    '2102': 21,  # JAULAS
    '2103': 21,  # HIGIENE/CUIDADO
    '2104': 21,  # ANIMAL VIVO

    # Familia 22 - PEQUEÑOS MAMÍFEROS
    '2201': 10,  # ALIMENTACION
    '2202': 21,  # JAULAS
    '2203': 21,  # HIGIENE/CUIDADO
    '2204': 21,  # ANIMAL VIVO

    # Familia 23 - PERROS
    '2301': 10,  # ALIMENTACION
    '2302': 21,  # CONFORT
    '2303': 21,  # CORREAS Y COLLARES
    '2304': 21,  # HIGIENE/CUIDADO
    '2305': 21,  # ANIMAL VIVO PERRO
    '2306': 21,  # PELUQUERIA

    # Familia 24 - GATOS
    '2401': 10,  # ALIMENTACION
    '2402': 21,  # CONFORT
    '2403': 21,  # CORREAS Y COLLARES
    '2404': 21,  # HIGIENE/CUIDADO
    '2405': 21,  # ANIMAL VIVO GATO

    # Familia 25 - ANIMALES DE GRANJA
    '2501': 10,  # ALIMENTACION
    '2502': 21,  # HABITAT / ACCES
    '2503': 21,  # HIGIENE/CUIDADO
    '2504': 21,  # ANIMAL VIVO GRANJA

    # Familia 26 - REPTILES
    '2601': 10,  # ALIMENTACION
    '2602': 21,  # TERRARIO
    '2603': 21,  # ACCESORIOS
    '2604': 21,  # DECO INERTE
    '2605': 21,  # HIGIENE/CUIDADO
    '2606': 21,  # ANIMAL VIVO REPTILES

    # Familia 27 - ACUARIOFILIA
    '2701': 10,  # ALIMENTACION
    '2702': 21,  # ACUARIOS
    '2703': 21,  # ACCESORIOS
    '2704': 21,  # DECO INERTE
    '2705': 10,  # PLANTA ACUATICA DECORACION
    '2706': 21,  # HIGIENE/CUIDADO
    '2707': 21,  # PECES AGUA CALIENTE
    '2708': 21,  # PECES AGUA FRIA
    '2709': 21,  # AGUA OSMOSIS

    # Familia 28 - JARDÍN ACUÁTICO
    '2801': 10,  # ALIMENTACION
    '2802': 21,  # ACCESORIOS
    '2803': 21,  # HIGIENE/CUIDADO
    '2804': 21,  # DECORACION
    '2805': 10,  # PLANTAS
    '2806': 21,  # PECES

    # Familia 29 - JARDÍN ACUÁTICO
    '2906': 21,  # INSECTO VIVO
}

# ============================================================================
# CARGA DE ENCARGADOS DESDE ARCHIVO JSON
# ============================================================================

def cargar_encargados():
    """
    Carga los encargados desde el archivo config/encargados.json
    
    Returns:
        dict: Diccionario de encargados cargado desde JSON
    """
    try:
        ruta_encargados = os.path.join(DIRECTORIO_CONFIG, 'encargados.json')
        if os.path.exists(ruta_encargados):
            with open(ruta_encargados, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('encargados', {})
        else:
            print(f"ADVERTENCIA: No se encontró {ruta_encargados}. Usando diccionario vacío.")
            return {}
    except Exception as e:
        print(f"ERROR al cargar encargados: {e}. Usando diccionario vacío.")
        return {}

# Cargar encargados desde JSON al inicio
ENCARGADOS = cargar_encargados()

# ============================================================================
# CONFIGURACIÓN DEL SERVIDOR SMTP (HARDCODED - SE ACTUALIZARÁ EN PUNTO 4)
# ============================================================================

SMTP_CONFIG = {
    'servidor': 'smtp.serviciodecorreo.es',
    'puerto': 465,
    'remitente_email': 'ivan.delgado@viveverde.es',
    'remitente_nombre': 'Sistema de Pedidos automáticos VIVEVERDE'
}

# ============================================================================
# FUNCIÓN PARA ENVIAR EMAIL CON ARCHIVO ADJUNTO
# ============================================================================

def enviar_email_clasificacion(seccion: str, archivo: str, periodo: str) -> bool:
    """
    Envía un email con el archivo de clasificación ABC+D adjunto al encargado de la sección.
    Soporta tanto un único encargado (objeto) como varios encargados (array).
    
    Args:
        seccion: Nombre de la sección procesada
        archivo: Ruta completa del archivo Excel generado
        periodo: Período del análisis (formato: "dd/mm/yyyy - dd/mm/yyyy")
    
    Returns:
        bool: True si el email fue enviado exitosamente a al menos un destinatario, False en caso contrario
    """
    # Obtener información del encargado
    encargado = ENCARGADOS.get(seccion.lower())
    
    if not encargado:
        print(f"  AVISO: No hay encargado configurado para la sección '{seccion}'. No se enviará email.")
        return False
    
    # Normalizar: puede ser un objeto único o un array de objetos
    if isinstance(encargado, list):
        lista_encargados = encargado
    else:
        lista_encargados = [encargado]
    
    # Verificar que el archivo existe
    if not Path(archivo).exists():
        print(f"  AVISO: El archivo '{archivo}' no existe. No se enviará email.")
        return False
    
    # Verificar contraseña en variable de entorno
    password = os.environ.get('EMAIL_PASSWORD')
    if not password:
        print(f"  AVISO: Variable de entorno 'EMAIL_PASSWORD' no configurada. No se enviará email.")
        return False
    
    # Enviar email a cada encargado en la lista
    emails_enviados = 0
    
    for encargado_item in lista_encargados:
        nombre_encargado = encargado_item.get('nombre', 'Encargado')
        email_destinatario = encargado_item.get('email', '')
        
        if not email_destinatario:
            print(f"  AVISO: Email no configurado para {nombre_encargado}. Saltando...")
            continue
        
        try:
            # Crear mensaje MIME
            msg = MIMEMultipart()
            msg['From'] = f"{SMTP_CONFIG['remitente_nombre']} <{SMTP_CONFIG['remitente_email']}>"
            msg['To'] = email_destinatario
            msg['Subject'] = f"VIVEVERDE: listado ClasificacionABC+D de {seccion} del periodo {periodo}"
            
            # Cuerpo del email
            cuerpo = f"""Buenos días {nombre_encargado},

Te adjunto en este correo el listado Clasificación ABC+D de {seccion} para que lo analices y te aprendas cuales son los artículos de cada categoría:

- Artículos que no te deben faltar nunca (Categoria A).
- Artículos que confeccionan el complemento de gama (Categoría B).
- Artículos que tienen una presencia mínima en las ventas de tu sección (Categoría C).
- Artículos que no debemos tener en tienda (Categoria D).

Pon en práctica el listado.

Atentamente,

Sistema de Pedidos automáticos VIVEVERDE."""
            
            msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
            
            # Adjuntar archivo Excel
            filename = Path(archivo).name
            with open(archivo, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename= "{filename}"')
            part.add_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            msg.attach(part)
            
            # Enviar email mediante SSL
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(SMTP_CONFIG['servidor'], SMTP_CONFIG['puerto'], context=context) as server:
                server.login(SMTP_CONFIG['remitente_email'], password)
                server.sendmail(SMTP_CONFIG['remitente_email'], email_destinatario, msg.as_string())
            
            print(f"  Email enviado a {nombre_encargado} ({email_destinatario})")
            emails_enviados += 1
            
        except smtplib.SMTPException as e:
            print(f"  ERROR SMTP al enviar email a {nombre_encargado}: {e}")
        except Exception as e:
            print(f"  ERROR al enviar email a {nombre_encargado}: {e}")
    
    return emails_enviados > 0

# ============================================================================
# FUNCIÓN PARA OBTENER IVA DE UN ARTÍCULO
# ============================================================================

def obtener_iva_articulo(codigo_articulo):
    """
    Obtiene el IVA correspondiente a un artículo según su familia o subfamilia.
    
    Args:
        codigo_articulo: Código del artículo (puede ser string o número)
    
    Returns:
        float: Porcentaje de IVA (10 o 21), o 21 por defecto si no se encuentra
    """
    if codigo_articulo is None or str(codigo_articulo) == 'nan':
        return 21  # IVA por defecto
    
    codigo_str = str(codigo_articulo).strip()
    
    # Eliminar decimales si viene como float
    if codigo_str.endswith('.0'):
        codigo_str = codigo_str[:-2]
    
    # Si empieza por 2, buscar en subfamilia (4 dígitos)
    if codigo_str.startswith('2'):
        if len(codigo_str) >= 4:
            subfamilia = codigo_str[:4]
            if subfamilia in IVA_SUBFAMILIA:
                return IVA_SUBFAMILIA[subfamilia]
    else:
        # Buscar en familia (2 dígitos)
        familia = codigo_str[:2]
        if familia in IVA_FAMILIA:
            return IVA_FAMILIA[familia]
    
    # IVA por defecto si no se encuentra
    return 21

# ============================================================================
# FUNCIÓN PARA DETERMINAR LA SECCIÓN DE UN ARTÍCULO
# ============================================================================

def determinar_seccion(codigo_articulo):
    """
    Determina la sección de un artículo según su código.
    
    Args:
        codigo_articulo: Código del artículo (puede ser string o número)
    
    Returns:
        str: Nombre de la sección o None si no se puede clasificar
    """
    if codigo_articulo is None:
        return None
    
    codigo_str = str(codigo_articulo).strip()
    
    # Eliminar decimales si viene como float
    if codigo_str.endswith('.0'):
        codigo_str = codigo_str[:-2]
    
    if not codigo_str or codigo_str == 'nan':
        return None
    
    # REGLA CRÍTICA: Filtrar artículos con menos de 10 dígitos
    # Esta regla tiene prioridad sobre todas las demás
    if len(codigo_str) < 10:
        return None
    
    # 1. Verificar códigos de mascotas vivos (primero, tienen prioridad)
    # Los códigos de mascotas vivos son códigos de 4 dígitos (2104, 2204, etc.)
    #，所以我们需要检查代码的前4位是否在这个列表中
    if codigo_str.startswith('2') and codigo_str[:4] in CODIGOS_MASCOTAS_VIVO:
        return 'mascotas_vivo'
    
    # 2. Sección 2: Mascotas manufacturadas (empieza por 2 y no está en vivos)
    if codigo_str.startswith('2'):
        return 'mascotas_manufacturado'
    
    # 3. Sección 3: Tierra/Áridos (31 o 32)
    if codigo_str.startswith('31') or codigo_str.startswith('32'):
        return 'tierra_aridos'
    
    # 4. Sección 3: Fitosanitarios (33-39)
    if codigo_str.startswith('3'):
        if len(codigo_str) >= 2:
            try:
                segundo_digito = int(codigo_str[1])
                if 3 <= segundo_digito <= 9:
                    return 'fitos'
            except (ValueError, IndexError):
                pass
    
    # 5. Secciones por primer dígito
    if codigo_str.startswith('1'):
        return 'interior'
    elif codigo_str.startswith('4'):
        return 'utiles_jardin'
    elif codigo_str.startswith('5'):
        return 'semillas'
    elif codigo_str.startswith('6'):
        return 'deco_interior'
    elif codigo_str.startswith('7'):
        return 'maf'
    elif codigo_str.startswith('8'):
        return 'vivero'
    elif codigo_str.startswith('9'):
        return 'deco_exterior'
    
    return None

# ============================================================================
# FUNCIÓN PARA PROCESAR UNA SECCIÓN ESPECÍFICA
# ============================================================================

def procesar_seccion(compras_df, ventas_df, stock_df, coste_df, nombre_seccion, seccion_info):
    """
    Procesa los datos de una sección específica y genera su archivo Excel.
    
    Args:
        compras_df: DataFrame de compras
        ventas_df: DataFrame de ventas
        stock_df: DataFrame de stock
        coste_df: DataFrame de costes
        nombre_seccion: Nombre de la sección a procesar
        seccion_info: Información de la sección (diccionario con descripción)
    
    Returns:
        dict: Estadísticas del procesamiento o None si no hay datos
    """
    print(f"\n{'='*80}")
    print(f"PROCESANDO SECCIÓN: {nombre_seccion.upper()}")
    print(f"Descripción: {seccion_info['descripcion']}")
    print(f"{'='*80}")
    
    # Filtrar datos por sección
    def filtrar_por_seccion(df, columna_codigo='codigo_str'):
        """Filtra un DataFrame para incluir solo artículos de la sección"""
        if columna_codigo not in df.columns:
            return df[df[columna_codigo].apply(lambda x: determinar_seccion(x) == nombre_seccion)]
        
        # Optimizado: aplicar determinar_seccion solo a códigos únicos
        codigos_unicos = df[columna_codigo].unique()
        codigos_seccion = set()
        for codigo in codigos_unicos:
            if determinar_seccion(codigo) == nombre_seccion:
                codigos_seccion.add(codigo)
        
        return df[df[columna_codigo].isin(codigos_seccion)]
    
    # Crear copias filtradas
    compras_seccion = filtrar_por_seccion(compras_df.copy(), 'codigo_str')
    ventas_seccion = filtrar_por_seccion(ventas_df.copy(), 'codigo_str')
    stock_seccion = filtrar_por_seccion(stock_df.copy(), 'codigo_str')
    
    print(f"Datos filtrados:")
    print(f"  - Compras: {len(compras_seccion)} registros")
    print(f"  - Ventas: {len(ventas_seccion)} registros")
    print(f"  - Stock: {len(stock_seccion)} registros")
    
    # Si no hay datos en ninguna tabla, avisar y continuar
    if len(compras_seccion) == 0 and len(ventas_seccion) == 0 and len(stock_seccion) == 0:
        print(f"  AVISO: No hay datos para la sección '{nombre_seccion}'. Saltando...")
        return None
    
    # =========================================================================
    # IDENTIFICACIÓN DE ARTÍCULOS ÚNICOS
    # =========================================================================
    
    def crear_clave(row):
        return (row['codigo_str'], row['nombre_str'], row['talla_str'], row['color_str'])
    
    articulos_compras = set(compras_seccion.apply(crear_clave, axis=1))
    articulos_ventas = set(ventas_seccion.apply(crear_clave, axis=1))
    articulos_stock = set(stock_seccion.apply(crear_clave, axis=1))
    
    articulos_unicos = articulos_compras.union(articulos_ventas).union(articulos_stock)
    print(f"\nTotal artículos únicos en sección: {len(articulos_unicos)}")
    
    if len(articulos_unicos) == 0:
        print(f"  AVISO: No hay artículos únicos en la sección '{nombre_seccion}'. Saltando...")
        return None
    
    # =========================================================================
    # PROCESAMIENTO DE DATOS POR ARTÍCULO
    # =========================================================================
    
    resultados = []
    
    for clave in articulos_unicos:
        codigo, nombre, talla, color = clave
        
        # Extraer familia del código
        codigo_str = str(codigo)
        if codigo_str.startswith('2'):
            familia_codigo = codigo_str[:4]  # 4 dígitos para familias de animales
        else:
            familia_codigo = codigo_str[:2]  # 2 dígitos para el resto
        
        if familia_codigo in ROTACIONES_FAMILIA:
            nombre_familia, rotacion_familia = ROTACIONES_FAMILIA[familia_codigo]
        else:
            nombre_familia = 'OTROS'
            rotacion_familia = 90
        
        # Datos de COMPRAS
        mask_compra = (compras_seccion['codigo_str'] == codigo) & \
                      (compras_seccion['nombre_str'] == nombre) & \
                      (compras_seccion['talla_str'] == talla) & \
                      (compras_seccion['color_str'] == color)
        compras_articulo = compras_seccion[mask_compra]
        total_compras = compras_articulo['Unidades'].sum() if len(compras_articulo) > 0 else 0
        
        # Datos de VENTAS
        mask_venta = (ventas_seccion['codigo_str'] == codigo) & \
                     (ventas_seccion['nombre_str'] == nombre) & \
                     (ventas_seccion['talla_str'] == talla) & \
                     (ventas_seccion['color_str'] == color)
        ventas_articulo = ventas_seccion[mask_venta]
        unidades_vendidas = ventas_articulo['Unidades'].sum() if len(ventas_articulo) > 0 else 0
        importe_ventas = ventas_articulo['Importe'].sum() if len(ventas_articulo) > 0 else 0
        beneficio = ventas_articulo['Beneficio'].sum() if len(ventas_articulo) > 0 else 0
        coste_ventas = ventas_articulo['Coste'].sum() if len(ventas_articulo) > 0 else 0
        
        # Fecha última venta
        if len(ventas_articulo) > 0:
            ultima_venta = ventas_articulo['Fecha'].max()
            antiguedad_ultima_venta = (FECHA_FIN - ultima_venta).days
        else:
            antiguedad_ultima_venta = DIAS_PERIODO
        
        # Datos de STOCK
        mask_stock = (stock_seccion['codigo_str'] == codigo) & \
                     (stock_seccion['nombre_str'] == nombre) & \
                     (stock_seccion['talla_str'] == talla) & \
                     (stock_seccion['color_str'] == color)
        stock_articulo = stock_seccion[mask_stock]
        stock_inicial = stock_articulo['Unidades'].sum() if len(stock_articulo) > 0 else 0
        precio_coste_stock = stock_articulo['Precio'].iloc[0] if len(stock_articulo) > 0 else 0
        
        # Métricas
        stock_disponible_total = stock_inicial + total_compras
        stock_final = stock_inicial + total_compras - unidades_vendidas
        
        # Tasa de Venta
        if stock_disponible_total > 0:
            tasa_venta = (unidades_vendidas / stock_disponible_total) * 100
        else:
            tasa_venta = 0
        
        # Antigüedad Stock
        if stock_final > 0:
            if stock_inicial - unidades_vendidas > 0:
                antiguedad_stock = DIAS_PERIODO
                origen_stock = 'Stock inicial'
            else:
                ventas_acumuladas = 0
                compras_ordenadas = compras_articulo.sort_values('Fecha')
                for idx, compra in compras_ordenadas.iterrows():
                    ventas_acumuladas += compra['Unidades']
                    if ventas_acumuladas >= (stock_inicial + total_compras - stock_final):
                        antiguedad_stock = (FECHA_FIN - compra['Fecha']).days
                        origen_stock = f'Compra {compra["Fecha"].strftime("%d/%m/%Y")}'
                        break
                else:
                    if len(compras_ordenadas) > 0:
                        ultima_compra = compras_ordenadas.iloc[-1]
                        antiguedad_stock = (FECHA_FIN - ultima_compra['Fecha']).days
                        origen_stock = f'Compra {ultima_compra["Fecha"].strftime("%d/%m/%Y")}'
                    else:
                        antiguedad_stock = DIAS_PERIODO
                        origen_stock = 'Stock inicial'
        else:
            antiguedad_stock = 0
            origen_stock = 'Sin stock'
        
        # % Rotación Consumida
        if stock_final > 0 and rotacion_familia > 0:
            pct_rotacion_consumida = (antiguedad_stock / rotacion_familia) * 100
        else:
            pct_rotacion_consumida = 0
        
        # Descuento Sugerido
        if pct_rotacion_consumida <= 65:
            descuento_sugerido = 0
        elif pct_rotacion_consumida <= 100:
            descuento_sugerido = 10
        elif pct_rotacion_consumida <= 150:
            descuento_sugerido = 20
        else:
            descuento_sugerido = 30
        
        # Riesgo de Merma/Inmovilizado
        es_categoria_d = (unidades_vendidas == 0)
        
        if es_categoria_d:
            if stock_final == 0:
                riesgo = 'Cero'
            else:
                riesgo = 'Crítico'
        else:
            if stock_final == 0:
                riesgo = 'Cero'
            elif pct_rotacion_consumida <= 65:
                riesgo = 'Bajo'
            elif pct_rotacion_consumida <= 100:
                riesgo = 'Medio'
            elif pct_rotacion_consumida <= 150:
                riesgo = 'Alto'
            else:
                riesgo = 'Crítico'
        
        # Rotación Excedida
        if antiguedad_ultima_venta > rotacion_familia and stock_final > 0:
            rotacion_excedida = stock_final
        else:
            rotacion_excedida = 0
        
        # Clasificación por Stock Final
        demanda_mensual_promedio = unidades_vendidas / 2
        if stock_final == 0:
            nivel_stock = 'Cero'
        elif stock_final <= demanda_mensual_promedio * 0.5:
            nivel_stock = 'Bajo'
        elif stock_final <= demanda_mensual_promedio:
            nivel_stock = 'Normal'
        else:
            nivel_stock = 'Elevado'
        
        # Ventas media diaria
        ventas_media_diaria = unidades_vendidas / DIAS_PERIODO if DIAS_PERIODO > 0 else 0
        
        # Stock Mínimo
        if rotacion_familia == 7:
            stock_minimo = ventas_media_diaria * 3.5
        elif rotacion_familia == 15:
            stock_minimo = ventas_media_diaria * 7.5
        elif rotacion_familia == 30:
            stock_minimo = ventas_media_diaria * 15
        elif rotacion_familia == 60:
            stock_minimo = ventas_media_diaria * 30
        elif rotacion_familia == 90:
            stock_minimo = ventas_media_diaria * 45
        else:
            stock_minimo = ventas_media_diaria * 45
        
        # Stock Máximo
        if rotacion_familia == 7:
            stock_maximo = ventas_media_diaria * 10.5
        elif rotacion_familia == 15:
            stock_maximo = ventas_media_diaria * 22.5
        elif rotacion_familia == 30:
            stock_maximo = ventas_media_diaria * 45
        elif rotacion_familia == 60:
            stock_maximo = ventas_media_diaria * 90
        elif rotacion_familia == 90:
            stock_maximo = ventas_media_diaria * 135
        else:
            stock_maximo = ventas_media_diaria * 135
        
        # Días de cobertura
        if ventas_media_diaria > 0:
            dias_cobertura = stock_final / ventas_media_diaria
        else:
            dias_cobertura = 0
        
        resultados.append({
            'Artículo': codigo,
            'Nombre artículo': nombre,
            'Talla': talla if talla else '',
            'Color': color if color else '',
            'Familia': familia_codigo,
            'Nombre Familia': nombre_familia,
            'Rotación Familia (días)': rotacion_familia,
            'Stock Inicial (unidades)': stock_inicial,
            'Compras Período (unidades)': total_compras,
            'Ventas (unidades)': unidades_vendidas,
            'Importe ventas (€)': round(importe_ventas, 2),
            'Beneficio (importe €)': round(beneficio, 2),
            'Coste Ventas Real (€)': round(coste_ventas, 2),
            'Stock Disponible Total': stock_disponible_total,
            'Tasa de venta (%)': round(tasa_venta, 2),
            'Rotación excedida (unidades)': rotacion_excedida,
            'Stock mínimo (unidades)': round(stock_minimo, 1),
            'Stock máximo (unidades)': round(stock_maximo, 1),
            'Stock Final (unidades)': stock_final,
            'Antigüedad Última Venta (días)': antiguedad_ultima_venta,
            'Antigüedad Stock (días)': antiguedad_stock,
            '% Rotación Consumido': round(pct_rotacion_consumida, 2),
            'Descuento Sugerido (%)': descuento_sugerido,
            'Riesgo de Merma/ inmovilizado': riesgo,
            'Nivel Stock Final': nivel_stock,
            'Días de cobertura': round(dias_cobertura, 1),
            'Origen Stock Final': origen_stock,
            'Precio Coste Unitario (€)': precio_coste_stock,
        })
    
    df_resultados = pd.DataFrame(resultados)
    print(f"\nTotal artículos procesados: {len(df_resultados)}")
    
    # =========================================================================
    # CLASIFICACIÓN ABC+D
    # =========================================================================
    
    df_con_ventas = df_resultados[df_resultados['Coste Ventas Real (€)'] > 0].copy()
    df_sin_ventas = df_resultados[df_resultados['Coste Ventas Real (€)'] == 0].copy()
    
    print(f"Artículos con ventas: {len(df_con_ventas)}")
    print(f"Artículos sin ventas: {len(df_sin_ventas)}")
    
    if len(df_con_ventas) > 0:
        df_con_ventas = df_con_ventas.sort_values('Coste Ventas Real (€)', ascending=False)
        
        total_coste = df_con_ventas['Coste Ventas Real (€)'].sum()
        df_con_ventas['% Individual'] = (df_con_ventas['Coste Ventas Real (€)'] / total_coste) * 100
        df_con_ventas['% Acumulado'] = df_con_ventas['% Individual'].cumsum()
        
        def asignar_categoria(pct_acumulado):
            if pct_acumulado <= 80:
                return 'A'
            elif pct_acumulado <= 95:
                return 'B'
            else:
                return 'C'
        
        df_con_ventas['Categoria ABC'] = df_con_ventas['% Acumulado'].apply(asignar_categoria)
        
        print(f"\n  Categoría A: {len(df_con_ventas[df_con_ventas['Categoria ABC'] == 'A'])} artículos")
        print(f"  Categoría B: {len(df_con_ventas[df_con_ventas['Categoria ABC'] == 'B'])} artículos")
        print(f"  Categoría C: {len(df_con_ventas[df_con_ventas['Categoria ABC'] == 'C'])} artículos")
    
    df_sin_ventas['Categoria ABC'] = 'D'
    
    df_clasificado = pd.concat([df_con_ventas, df_sin_ventas], ignore_index=True)
    
    print(f"\n  Categoría D: {len(df_clasificado[df_clasificado['Categoria ABC'] == 'D'])} artículos")
    
    # =========================================================================
    # ASIGNACIÓN DE ESCENARIOS
    # =========================================================================
    
    def asignar_escenario(row):
        stock_final = row['Stock Final (unidades)']
        pct_rotacion = row['% Rotación Consumido']
        antiguedad_venta = row['Antigüedad Última Venta (días)']
        rotacion = row['Rotación Familia (días)']
        categoria = row['Categoria ABC']
        riesgo = row['Riesgo de Merma/ inmovilizado']
        nivel_stock = row['Nivel Stock Final']
        
        if stock_final == 0:
            if rotacion > 0:
                pct_rotacion_venta = (antiguedad_venta / rotacion) * 100
            else:
                pct_rotacion_venta = 0
            
            if categoria in ['A', 'B']:
                if pct_rotacion_venta <= 24: return '13A'
                elif pct_rotacion_venta <= 50: return '13B'
                elif pct_rotacion_venta <= 100: return '13C'
                else: return '13D'
            else:
                if pct_rotacion_venta <= 24: return '26A'
                elif pct_rotacion_venta <= 50: return '26B'
                elif pct_rotacion_venta <= 100: return '26C'
                else: return '26D'
        else:
            if nivel_stock == 'Elevado':
                if riesgo == 'Crítico': return '1' if categoria in ['A', 'B'] else '14'
                elif riesgo == 'Alto': return '2' if categoria in ['A', 'B'] else '15'
                elif riesgo == 'Medio': return '3' if categoria in ['A', 'B'] else '16'
                else: return '4' if categoria in ['A', 'B'] else '17'
            elif nivel_stock == 'Normal':
                if riesgo == 'Crítico': return '5' if categoria in ['A', 'B'] else '18'
                elif riesgo == 'Alto': return '6' if categoria in ['A', 'B'] else '19'
                elif riesgo == 'Medio': return '7' if categoria in ['A', 'B'] else '20'
                else: return '8' if categoria in ['A', 'B'] else '21'
            else:
                if riesgo == 'Crítico': return '9' if categoria in ['A', 'B'] else '22'
                elif riesgo == 'Alto': return '10' if categoria in ['A', 'B'] else '23'
                elif riesgo == 'Medio': return '11' if categoria in ['A', 'B'] else '24'
                else: return '12' if categoria in ['A', 'B'] else '25'
    
    df_clasificado['Escenario'] = df_clasificado.apply(asignar_escenario, axis=1)
    
    # =========================================================================
    # ACCIONES SUGERIDAS
    # =========================================================================
    
    TEXTOS_ESCENARIOS = {
        '1': "DESCUENTO MÁXIMO + REDUCCIÓN COMPRAS: Aplicar descuento [descuento]% inmediato. Reducir compras 50% próxima temporada. Stock objetivo: [unidades] unidades. Prioridad alta.",
        '2': "DESCUENTO MODERADO + REDUCCIÓN COMPRAS: Aplicar descuento [descuento]% para dinamizar ventas. Reducir compras 35% próxima temporada. Stock objetivo: [unidades] unidades. Monitorear.",
        '3': "DESCUENTO PREVENTIVO + AJUSTE COMPRAS: Aplicar descuento [descuento]% para anticipar venta. Reducir compras 20% próxima temporada. Mantener bajo observación semanal.",
        '4': "MANTENER + GESTIÓN ACTIVA: Stock fresco de calidad. Reducir compras 15% próxima temporada. Stock actual suficiente para [X días] días.",
        '5': "DESCUENTO CORRECTIVO + MONITOREO: Aplicar descuento [descuento]% a stock actual para renovar inventario. Mantener nivel de compras actual.",
        '6': "DESCUENTO LEVE + OPTIMIZACIÓN: Aplicar descuento [descuento]% para renovar inventario. Reducir compras 15% próxima temporada.",
        '7': "OPTIMIZAR PREVENTIVO: Aplicar descuento [descuento]% preventivo. Mantener nivel de compras actual. Stock bien gestionado.",
        '8': "MANTENER ESTRATEGIA ACTUAL: Gestión excelente. Stock óptimo y fresco. Mantener nivel de compras actual. Producto clave del catálogo.",
        '9': "INVESTIGAR + REDISEÑAR: Analizar causa de baja rotación. Mantener stock mínimo. Implementar acciones de venta. Reducir compras 25%.",
        '10': "PROMOCIÓN ACTIVA + AJUSTE: Implementar promoción del 15% para estimula demanda. Aumentar visibilidad en punto de venta.",
        '11': "REPOSICIÓN SELECTIVA: Aumentar compras 15% para evitar ruptura de stock. Aplicar descuento 5% para consolidar demanda.",
        '12': "AUMENTAR STOCK: Producto de alto interés. Incrementar compras 20% próxima temporada. Stock actual: [unidades] unidades. Maximizar disponibilidad.",
        '13A': "URGENTE - REPOSICIÓN INMEDIATA: Producto de alta demanda agotado. Recompra prioritaria inmediata. Aumentar compras 40%. Stock objetivo: [unidades] unidades.",
        '13B': "REPOSICIÓN PRIORITARIA: Producto agotado con demanda reciente. Aumentar compras 25%. Stock objetivo: [unidades] unidades.",
        '13C': "REPOSICIÓN PROGRAMADA: Stock agotado con rotación moderada. Mantener nivel de compras anterior. Stock objetivo: [unidades] unidades.",
        '13D': "EVALUAR CONTINUIDAD: Producto agotado con demanda decreciente. Reducir compras 30% próxima temporada. Evaluar continuidad en catálogo.",
        '14': "LIQUIDACIÓN URGENTE: Aplicar descuento [descuento]% inmediato. Eliminar del catálogo próxima temporada. Capital liberado: [importe]€. Prioridad máxima.",
        '15': "REDUCCIÓN AGRESIVA: Aplicar descuento [descuento]% inmediato. Reducir compras 70% próxima temporada. Stock objetivo: [unidades] unidades. Riesgo alto de merma.",
        '16': "DESCUENTO PREVENTIVO: Aplicar descuento [descuento]% para acelerar rotación. Reducir compras 40% próxima temporada. Monitorear evolución semanal.",
        '17': "MANTENER SIN DESCUENTO: Stock fresco de calidad. Reducir compras 25% próxima temporada. Stock actual suficiente para [X días] días.",
        '18': "LIQUIDACIÓN PARCIAL: Aplicar descuento [descuento]% a stock actual. Reducir compras 50% próxima temporada. Producto de baja rotación confirmada.",
        '19': "DESCUENTO MODERADO: Aplicar descuento [descuento]% para renovar inventario. Reducir compras 30% próxima temporada. Stock actual en rango aceptable pero envejecido.",
        '20': "OPTIMIZAR: Aplicar descuento [descuento]% preventivo. Mantener nivel de compras actual. Stock bien gestionado. Continuar monitoreo.",
        '21': "MANTENER ESTRATEGIA ACTUAL: Gestión excelente. Stock óptimo y fresco. Mantener nivel de compras. Producto bien equilibrado.",
        '22': "ELIMINAR DEL CATÁLOGO: Aplicar descuento [descuento]% para liquidar stock residual. NO recomprar. Bajo interés confirmado del cliente.",
        '23': "LIQUIDAR Y DESCATALOGAR: Aplicar descuento [descuento]% para agotar stock. NO recomprar próxima temporada. Producto sin demanda suficiente.",
        '24': "COMPRAS CONSERVADORAS: Aplicar descuento [descuento]% al stock actual. Reducir compras 50% próxima temporada. Demanda limitada confirmada.",
        '25': "AUMENTAR STOCK: Producto de alto interés. Incrementar compras 30% próxima temporada. Stock actual: [unidades] unidades. Alta rotación confirmada.",
        '26A': "URGENTE - RUPTURA DE STOCK: Producto de alta demanda agotado. Recompra inmediata prioritaria. Aumentar compras 50%. Stock objetivo: [unidades] unidades.",
        '26B': "RECOMPRA PRIORITARIA: Producto agotado con demanda reciente. Aumentar compras 30%. Stock objetivo: [unidades] unidades. Monitorear demanda.",
        '26C': "RECOMPRA MODERADA: Stock agotado con rotación moderada. Mantener nivel de compras anterior. Stock objetivo: [unidades] unidades. Demanda estable.",
        '26D': "RECOMPRA CONSERVADORA: Producto agotado de baja rotación. Reducir compras 40% próxima temporada. Stock objetivo mínimo: [unidades] unidades.",
    }
    
    def generar_accion_sugerida(row):
        escenario = row['Escenario']
        if escenario not in TEXTOS_ESCENARIOS:
            return "Sin acción asignada"
        
        texto = TEXTOS_ESCENARIOS[escenario]
        descuento = row['Descuento Sugerido (%)']
        stock_final = row['Stock Final (unidades)']
        stock_minimo = row['Stock mínimo (unidades)']
        dias_cobertura = row['Días de cobertura']
        
        stock_objetivo_14_dias = max(1, round(stock_minimo * 0.5, 0))
        stock_objetivo_aumentar = max(1, round(stock_final * 1.5, 0))
        stock_objetivo_doble = max(1, round(stock_final * 2, 0))
        capital_liberado = round(stock_final * row['Precio Coste Unitario (€)'] * 0.7, 2)
        
        texto = texto.replace('[descuento]', str(descuento))
        texto = texto.replace('[unidades]', str(int(stock_objetivo_14_dias)))
        texto = texto.replace('[X días]', str(int(dias_cobertura)))
        texto = texto.replace('[importe]', str(capital_liberado))
        
        # Casos especiales
        if escenario == '12':
            texto = f"AUMENTAR STOCK: Producto de alto interés. Incrementar compras 20% próxima temporada. Stock actual: {int(stock_final)} unidades. Stock objetivo: {int(stock_objetivo_aumentar)} unidades. Maximizar disponibilidad."
        elif escenario == '13A':
            texto = f"URGENTE - REPOSICIÓN INMEDIATA: Producto de alta demanda agotado. Recompra prioritaria inmediata. Aumentar compras 40%. Stock objetivo: {int(stock_objetivo_doble)} unidades. Evitar futura ruptura."
        elif escenario == '13B':
            texto = f"REPOSICIÓN PRIORITARIA: Producto agotado con demanda reciente. Aumentar compras 25%. Stock objetivo: {int(stock_objetivo_aumentar)} unidades. Programar reposición para próxima semana."
        elif escenario == '25':
            texto = f"AUMENTAR STOCK: Producto de alto interés. Incrementar compras 30% próxima temporada. Stock actual: {int(stock_final)} unidades. Stock objetivo: {int(stock_objetivo_aumentar)} unidades. Alta rotación confirmada."
        elif escenario == '26A':
            texto = f"URGENTE - RUPTURA DE STOCK: Producto de alta demanda agotado. Recompra inmediata prioritaria. Aumentar compras 50%. Stock objetivo: {int(stock_objetivo_doble)} unidades. Pérdida de ventas estimada."
        elif escenario == '26B':
            texto = f"RECOMPRA PRIORITARIA: Producto agotado con demanda reciente. Aumentar compras 30%. Stock objetivo: {int(stock_objetivo_aumentar)} unidades. Monitorear demanda próximas semanas."
        
        return texto
    
    df_clasificado['Acción Sugerida'] = df_clasificado.apply(generar_accion_sugerida, axis=1)
    
    # =========================================================================
    # SEPARACIÓN POR CATEGORÍAS
    # =========================================================================
    
    columnas_salida = [
        'Artículo', 'Nombre artículo', 'Talla', 'Color',
        'Familia', 'Nombre Familia', 'Rotación Familia (días)',
        'Ventas (unidades)', 'Importe ventas (€)', 'Beneficio (importe €)',
        'Tasa de venta (%)', 'Rotación excedida (unidades)',
        'Stock mínimo (unidades)', 'Stock máximo (unidades)',
        'Stock Final (unidades)', 'Antigüedad Última Venta (días)',
        'Antigüedad Stock (días)', '% Rotación Consumido',
        'Descuento Sugerido (%)', 'Riesgo de Merma/ inmovilizado',
        'Acción Sugerida', 'Origen Stock Final', 'Escenario'
    ]
    
    df_categoria_a = df_clasificado[df_clasificado['Categoria ABC'] == 'A'][columnas_salida].copy()
    df_categoria_b = df_clasificado[df_clasificado['Categoria ABC'] == 'B'][columnas_salida].copy()
    df_categoria_c = df_clasificado[df_clasificado['Categoria ABC'] == 'C'][columnas_salida].copy()
    df_categoria_d = df_clasificado[df_clasificado['Categoria ABC'] == 'D'][columnas_salida].copy()
    
    # =========================================================================
    # GUARDAR ARCHIVO EXCEL
    # =========================================================================
    
    nombre_archivo = os.path.join(DIRECTORIO_DATA, f"CLASIFICACION_ABC+D_{nombre_seccion.upper()}_{PERIODO}_{AÑO}.xlsx")
    
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df_categoria_a.to_excel(writer, sheet_name='CATEGORIA A – BASICOS', index=False)
        df_categoria_b.to_excel(writer, sheet_name='CATEGORIA B – COMPLEMENTO', index=False)
        df_categoria_c.to_excel(writer, sheet_name='CATEGORIA C – BAJO IMPACTO', index=False)
        df_categoria_d.to_excel(writer, sheet_name='CATEGORIA D – SIN VENTAS', index=False)
    
    print(f"\nArchivo generado: {nombre_archivo}")
    
    # =========================================================================
    # APLICAR FORMATOS EXCEL
    # =========================================================================
    
    def aplicar_formato_hoja(worksheet, df):
        worksheet.column_dimensions['A'].width = 18
        worksheet.column_dimensions['B'].width = 45
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['M'].width = 18
        worksheet.column_dimensions['N'].width = 18
        worksheet.column_dimensions['S'].width = 22
        worksheet.column_dimensions['U'].width = 32
        worksheet.column_dimensions['X'].width = 15
        
        columnas_ocultar = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'O', 'P', 'Q', 'R', 'V', 'W']
        for col in columnas_ocultar:
            worksheet.column_dimensions[col].hidden = True
        
        fill_cabecera = PatternFill(start_color=COLOR_CABECERA, end_color=COLOR_CABECERA, fill_type='solid')
        font_cabecera = Font(color=COLOR_TEXTO_CABECERA, bold=True, size=10)
        
        for cell in worksheet[1]:
            cell.fill = fill_cabecera
            cell.font = font_cabecera
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        worksheet.row_dimensions[1].height = 45
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        columnas_centradas = ['M', 'N', 'S', 'T']
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = thin_border
                
                col_letter = chr(64 + col_idx)
                if col_letter in columnas_centradas:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_letter in ['A', 'B', 'C', 'D', 'U']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                if col_letter == 'T' and cell.value:
                    riesgo = str(cell.value)
                    if riesgo in COLORES_RIESGO:
                        cell.fill = PatternFill(start_color=COLORES_RIESGO[riesgo], 
                                               end_color=COLORES_RIESGO[riesgo], 
                                               fill_type='solid')
        
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.margin_left = 0
        worksheet.page_setup.margin_right = 0
        worksheet.page_setup.margin_top = 0
        worksheet.page_setup.margin_bottom = 0
        
        worksheet.auto_filter.ref = worksheet.dimensions
        
        return worksheet
    
    wb = load_workbook(nombre_archivo)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        df = pd.read_excel(nombre_archivo, sheet_name=sheet_name)
        aplicar_formato_hoja(ws, df)
    
    wb.save(nombre_archivo)
    
    print(f"\nEnviando email al encargado de la sección...")
    
    # Formatear período para el email
    periodo_str = f"{FECHA_INICIO.strftime('%d/%m/%Y')} - {FECHA_FIN.strftime('%d/%m/%Y')}"
    
    # Enviar email con el archivo adjunto
    email_enviado = enviar_email_clasificacion(nombre_seccion, nombre_archivo, periodo_str)
    
    # Retornar estadísticas
    return {
        'seccion': nombre_seccion,
        'archivo': nombre_archivo,
        'total_articulos': len(df_clasificado),
        'categoria_a': len(df_categoria_a),
        'categoria_b': len(df_categoria_b),
        'categoria_c': len(df_categoria_c),
        'categoria_d': len(df_categoria_d),
        'email_enviado': email_enviado,
    }

# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

def main():
    """Función principal del script"""
    
    # =========================================================================
    # PARSEO DE ARGUMENTOS DE LÍNEA DE COMANDOS
    # =========================================================================
    
    # Primeiro, verificar se hai argumentos sin前缀 para detectar o modo
    import sys
    
    # Analizar os argumentos de forma flexible
    argumentos_sin_prefijo = []
    argumentos_con_prefijo = {}
    
    i = 1
    while i < len(sys.argv):
        arg = sys.argv[i]
        
        # Argumentos con prefijo corto (-P, -Y, -S)
        if arg.startswith('-') and not arg.startswith('--'):
            if i + 1 < len(sys.argv) and not sys.argv[i + 1].startswith('-'):
                argumentos_con_prefijo[arg] = sys.argv[i + 1]
                i += 2
            else:
                argumentos_con_prefijo[arg] = True
                i += 1
        # Argumentos con prefijo longo (--P3, --2025, --maf)
        elif arg.startswith('--'):
            arg_sin_doble_guion = arg[2:]
            # Detectar se é un período (P1, P2, P3, P4)
            if arg_sin_doble_guion.upper() in ['P1', 'P2', 'P3', 'P4']:
                argumentos_con_prefijo['--periodo'] = arg_sin_doble_guion.upper()
                i += 1
            # Detectar se é un ano (número de 4 díxitos)
            elif arg_sin_doble_guion.isdigit() and len(arg_sin_doble_guion) == 4:
                argumentos_con_prefijo['--año'] = int(arg_sin_doble_guion)
                i += 1
            # Detectar se é unha sección
            elif arg_sin_doble_guion.lower() in SECCIONES:
                argumentos_con_prefijo['--seccion'] = arg_sin_doble_guion.lower()
                i += 1
            else:
                i += 1
        # Argumentos sen prefijo (posicionais)
        else:
            argumentos_sin_prefijo.append(arg)
            i += 1
    
    # Determinar período, ano e sección desde os argumentos
    periodo_especificado = None
    año_especificado = None
    seccion_especificada = None
    
    # Analizar argumentos sin prefijo (posicionais)
    for idx, arg in enumerate(argumentos_sin_prefijo):
        arg_upper = arg.upper()
        arg_lower = arg.lower()
        
        # Se é un período (P1, P2, P3, P4)
        if arg_upper in ['P1', 'P2', 'P3', 'P4']:
            periodo_especificado = arg_upper
        # Se é un ano (número de 4 díxitos)
        elif arg.isdigit() and len(arg) == 4:
            año_especificado = int(arg)
        # Se é unha sección
        elif arg_lower in SECCIONES:
            seccion_especificada = arg_lower
    
    # Sobrescribir cos argumentos con prefijo se existen
    if '--periodo' in argumentos_con_prefijo:
        periodo_especificado = argumentos_con_prefijo['--periodo']
    if '--año' in argumentos_con_prefijo:
        año_especificado = argumentos_con_prefijo['--año']
    if '--seccion' in argumentos_con_prefijo:
        seccion_especificada = argumentos_con_prefijo['--seccion']
    
    # Ver os argumentos -P, -Y, -S tamén
    if '-P' in argumentos_con_prefijo:
        periodo_especificado = argumentos_con_prefijo['-P']
    if '-Y' in argumentos_con_prefijo:
        año_especificado = argumentos_con_prefijo['-Y']
    if '-S' in argumentos_con_prefijo:
        seccion_especificada = argumentos_con_prefijo['-S']
    
    # Determinar o modo de operación
    # Modo automático: solo sin parámetros, o solo con sección (sin período ni año)
    # Modo manual: cualquier combinación que incluya período o año, o solo período, o solo año
    modo_automatico = (periodo_especificado is None and año_especificado is None)
    
    # Obter data actual
    fecha_actual = datetime.now()
    año_actual = fecha_actual.year
    periodo_actual = obtener_periodo_desde_fecha(fecha_actual, CONFIG)
    
    if modo_automatico:
        # MODO AUTOMÁTICO: Sin parámetros especificados
        # Analiza datos do ano anterior e xera arquivos para o período seguinte
        periodo_seleccionado = obtener_periodo_siguiente(periodo_actual)
        año_datos = año_actual - 1
        print("=" * 80)
        print("MODO: AUTOMÁTICO (período e ano calculados desde data do sistema)")
        print("=" * 80)
        print(f"\nData actual do sistema: {fecha_actual.strftime('%d de %B de %Y')}")
        print(f"Período actual detectado: {periodo_actual}")
        print(f"Ano actual: {año_actual}")
        print(f"\n>>> O script analizará os datos do ano {año_datos}")
        print(f">>> e xerará arquivos para o período {periodo_seleccionado}")
    else:
        # MODO MANUAL: Con parámetros específicos
        periodo_seleccionado = periodo_especificado if periodo_especificado else obtener_periodo_siguiente(periodo_actual)
        año_datos = año_especificado if año_especificado else año_actual - 1
        
        print("=" * 80)
        print("MODO: MANUAL (parámetros especificados polo usuario)")
        print("=" * 80)
        print(f"\nPeríodo especificado: {periodo_seleccionado}")
        print(f"Ano especificado: {año_datos}")
    
    seccion_especifica = seccion_especificada
    
    # Validar sección se se especificou
    if seccion_especifica and seccion_especifica not in SECCIONES:
        print(f"ERROR: Sección '{seccion_especifica}' non válida.")
        print(f"Seccións dispoñibles: {', '.join(sorted(SECCIONES.keys()))}")
        sys.exit(1)
    
    # Obtener fecha actual
    fecha_actual = datetime.now()
    año_actual = fecha_actual.year
    periodo_actual = obtener_periodo_desde_fecha(fecha_actual, CONFIG)
    
    if modo_automatico:
        # MODO AUTOMÁTICO: Sin parámetros especificados
        # Analiza datos del año anterior y genera archivos para el período siguiente
        periodo_seleccionado = obtener_periodo_siguiente(periodo_actual)
        año_datos = año_actual - 1
        print("=" * 80)
        print("MODO: AUTOMÁTICO (período y año calculados desde fecha del sistema)")
        print("=" * 80)
        print(f"\nFecha actual del sistema: {fecha_actual.strftime('%d de %B de %Y')}")
        print(f"Período actual detectado: {periodo_actual}")
        print(f"Año actual: {año_actual}")
        print(f"\n>>> El script analizará los datos del año {año_datos}")
        print(f">>> y generará archivos para el período {periodo_seleccionado}")
    else:
        # MODO MANUAL: Con parámetros específicos
        periodo_seleccionado = periodo_especificado if periodo_especificado else obtener_periodo_siguiente(periodo_actual)
        año_datos = año_especificado if año_especificado else año_actual - 1
        
        print("=" * 80)
        print("MODO: MANUAL (parámetros especificados por el usuario)")
        print("=" * 80)
        print(f"\nPeríodo especificado: {periodo_seleccionado}")
        print(f"Año especificado: {año_datos}")
    
    # La variable seccion_especifica ya está definida en el parser personalizado
    
    # Validar sección si se especificó
    if seccion_especifica and seccion_especifica not in SECCIONES:
        print(f"ERROR: Sección '{seccion_especifica}' no válida.")
        print(f"Secciones disponibles: {', '.join(sorted(SECCIONES.keys()))}")
        sys.exit(1)
    
    print("=" * 80)
    print("MOTOR DE CÁLCULO ABC+D PARA GESTIÓN DE INVENTARIOS")
    print("=" * 80)
    
    if periodo_seleccionado:
        print(f"\nMODO: Período específico")
        print(f"Período seleccionado: {periodo_seleccionado}")
    else:
        print(f"\nMODO: Período por defecto (año completo)")
    
    if seccion_especifica:
        print(f"Sección seleccionada: {seccion_especifica}")
    else:
        print(f"MODO: Multi-sección (todas las secciones)")
    
    # =========================================================================
    # CARGA DE DATOS DESDE ARCHIVOS CON DATOS DEL AÑO COMPLETO
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("FASE 1: CARGA Y EXTRACCIÓN DE DATOS")
    print("=" * 80)
    
    try:
        # Cargar archivos con datos de TODO el año
        compras_df = pd.read_excel(os.path.join(DIRECTORIO_DATA, 'SPA_compras.xlsx'))
        ventas_df = pd.read_excel(os.path.join(DIRECTORIO_DATA, 'SPA_ventas.xlsx'))
        # El archivo de stock se cargará después de detectar el año
        # El archivo de costes puede llamarse SPA_Coste.xlsx o SPA_coste.xlsx
        if os.path.exists(os.path.join(DIRECTORIO_DATA, 'SPA_Coste.xlsx')):
            coste_df = pd.read_excel(os.path.join(DIRECTORIO_DATA, 'SPA_Coste.xlsx'))
        elif os.path.exists(os.path.join(DIRECTORIO_DATA, 'SPA_coste.xlsx')):
            coste_df = pd.read_excel(os.path.join(DIRECTORIO_DATA, 'SPA_coste.xlsx'))
        else:
            raise FileNotFoundError("No se encontró SPA_Coste.xlsx ni SPA_coste.xlsx")
    except FileNotFoundError as e:
        print(f"ERROR: No se encontró el archivo: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR al cargar archivos: {e}")
        sys.exit(1)
    
    print(f"COMPRAS (año completo): {len(compras_df)} registros cargados")
    print(f"VENTAS (año completo): {len(ventas_df)} registros cargados")
    print(f"COSTE: {len(coste_df)} registros cargados")
    
    # =========================================================================
    # CARGAR STOCK Y CONFIGURAR PERÍODO USANDO AÑO Y PERIODO SELECCIONADOS
    # =========================================================================
    
    # El período y año ya fueron determinados al inicio del main()
    # periodo_seleccionado contiene el período a procesar
    # año_datos contiene el año de los datos a analizar
    
    # Determinar el nombre del archivo de stock según el período seleccionado
    # Siempre hay un período seleccionado (ya sea automático o manual)
    nombre_stock = f'SPA_stock_{periodo_seleccionado}.xlsx'
    
    try:
        stock_df = pd.read_excel(os.path.join(DIRECTORIO_DATA, nombre_stock))
    except FileNotFoundError:
        print(f"ADVERTENCIA: No se encontró {nombre_stock}, buscando archivo alternativo...")
        # Buscar cualquier archivo de stock disponible
        archivos_stock = [f for f in os.listdir(DIRECTORIO_DATA) if f.startswith('SPA_stock') and f.endswith('.xlsx')]
        if archivos_stock:
            nombre_stock = archivos_stock[0]
            stock_df = pd.read_excel(os.path.join(DIRECTORIO_DATA, nombre_stock))
        else:
            print("ERROR: No se encontró ningún archivo de stock")
            sys.exit(1)
    
    print(f"STOCK: {len(stock_df)} registros cargados ({nombre_stock})")
    
    # Configurar el período usando el año de datos seleccionado
    global FECHA_INICIO, FECHA_FIN, DIAS_PERIODO, PERIODO, AÑO
    FECHA_INICIO, FECHA_FIN, DIAS_PERIODO, PERIODO, AÑO = configurar_periodo(periodo_seleccionado, CONFIG, año_datos)
    
    print(f"\nPeríodo de análisis:")
    print(f"   Período a generar: {periodo_seleccionado}")
    print(f"   Año de los datos: {año_datos}")
    print(f"   Fechas: {FECHA_INICIO.strftime('%d de %B de %Y')} - {FECHA_FIN.strftime('%d de %B de %Y')}")
    print(f"   Días: {DIAS_PERIODO}")
    
    # =========================================================================
    # FILTRAR DATOS POR PERÍODO (SOLO COMPRAS Y VENTAS)
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("FASE 1A: FILTRADO DE DATOS POR PERÍODO")
    print("=" * 80)
    
    # Convertir fechas a datetime si no lo son
    compras_df['Fecha'] = pd.to_datetime(compras_df['Fecha'], errors='coerce')
    ventas_df['Fecha'] = pd.to_datetime(ventas_df['Fecha'], errors='coerce')
    
    # Filtrar compras por período
    filas_antes_compras = len(compras_df)
    compras_df = compras_df[
        (compras_df['Fecha'] >= FECHA_INICIO) & 
        (compras_df['Fecha'] <= FECHA_FIN)
    ].copy()
    filas_despues_compras = len(compras_df)
    print(f"COMPRAS filtradas por período: {filas_antes_compras} → {filas_despues_compras} registros")
    print(f"   Período: {FECHA_INICIO.strftime('%d/%m/%Y')} - {FECHA_FIN.strftime('%d/%m/%Y')}")
    
    # Filtrar ventas por período
    filas_antes_ventas = len(ventas_df)
    ventas_df = ventas_df[
        (ventas_df['Fecha'] >= FECHA_INICIO) & 
        (ventas_df['Fecha'] <= FECHA_FIN)
    ].copy()
    filas_despues_ventas = len(ventas_df)
    print(f"VENTAS filtradas por período: {filas_antes_ventas} → {filas_despues_ventas} registros")
    print(f"   Período: {FECHA_INICIO.strftime('%d/%m/%Y')} - {FECHA_FIN.strftime('%d/%m/%Y')}")
    
    if len(compras_df) == 0:
        print("ADVERTENCIA: No hay datos de compras en el período especificado.")
    if len(ventas_df) == 0:
        print("ADVERTENCIA: No hay datos de ventas en el período especificado.")
    
    # Filtrar filas con Artículo vacío en Compras
    filas_antes = len(compras_df)
    compras_df = compras_df[compras_df['Artículo'].notna() & (compras_df['Artículo'] != '')]
    filas_eliminadas = filas_antes - len(compras_df)
    if filas_eliminadas > 0:
        print(f"Eliminadas {filas_eliminadas} filas con artículo vacío en Compras")
    
    # Rellenar celdas vacías en STOCK
    filas_vacias_stock = stock_df['Artículo'].isna().sum()
    if filas_vacias_stock > 0:
        stock_df['Artículo'] = stock_df['Artículo'].ffill()
        stock_df['Nombre artículo'] = stock_df['Nombre artículo'].ffill()
        print(f"STOCK: {len(stock_df)} registros ({filas_vacias_stock} celdas vacías preenchidas)")
    else:
        print(f"STOCK: {len(stock_df)} registros")
    
    # =========================================================================
    # PROCESAR DATOS DE VENTAS - Calcular Coste y Beneficio
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("FASE 1B: CÁLCULO DE COSTE Y BENEFICIO EN VENTAS")
    print("=" * 80)
    
    # Filtrar solo filas de tipo 'Detalle'
    filas_ventas_total = len(ventas_df)
    ventas_df = ventas_df[ventas_df['Tipo registro'] == 'Detalle'].copy()
    print(f"VENTAS: {filas_ventas_total} filas totales → {len(ventas_df)} filas de Detalle")
    
    # Normalizar claves de unión en Coste
    # Ordenar por fecha de última compra (buscando columna automáticamente)
    columna_fecha = encontrar_columna(list(coste_df.columns), 'ultima compra')
    if columna_fecha:
        coste_df_sorted = coste_df.sort_values(columna_fecha, ascending=False)
    else:
        print("ADVERTENCIA: No se encontró columna de fecha de última compra, usando orden original")
        coste_df_sorted = coste_df.copy()
    coste_df_latest = coste_df_sorted.drop_duplicates(subset=['Artículo', 'Talla', 'Color'], keep='first').copy()
    
    def normalize_keys(df):
        df = df.copy()
        df['Artículo'] = df['Artículo'].astype(str).str.replace(r'\.0$', '', regex=True)
        df['Talla'] = df['Talla'].fillna('').astype(str).str.strip()
        df['Color'] = df['Color'].fillna('').astype(str).str.strip()
        return df
    
    def normalize_keys_coste(df):
        """Normalizar claves para el archivo SPA_Coste.xlsx (ya tiene columna 'Artículo')"""
        df = df.copy()
        df['Artículo'] = df['Artículo'].astype(str).str.replace(r'\.0$', '', regex=True)
        df['Talla'] = df['Talla'].fillna('').astype(str).str.strip()
        df['Color'] = df['Color'].fillna('').astype(str).str.strip()
        return df
    
    ventas_normalized = normalize_keys(ventas_df)
    coste_normalized = normalize_keys_coste(coste_df_latest)
    
    # Seleccionar solo las columnas necesarias de coste (ya renombrado a Artículo)
    coste_for_merge = coste_normalized[['Artículo', 'Talla', 'Color', 'Coste']].copy()
    
    # Merge de ventas con costes
    ventas_with_costs = pd.merge(
        ventas_normalized,
        coste_for_merge,
        on=['Artículo', 'Talla', 'Color'],
        how='left'
    )
    
    # Calcular Coste total
    def calculate_cost(row):
        try:
            unidades = row['Unidades'] if pd.notna(row['Unidades']) else 1
            importe = row['Importe'] if pd.notna(row['Importe']) else 0
            coste_unitario = row['Coste'] if pd.notna(row['Coste']) else 0
            
            if coste_unitario > 0:
                return unidades * coste_unitario
            
            if unidades > 0 and importe > 0:
                pvp = importe / unidades
                iva = obtener_iva_articulo(row['Artículo'])
                
                if iva == 10:
                    coste_calculado = (pvp / 1.10) / 2.3
                else:
                    coste_calculado = (pvp / 1.21) / 2
                
                return unidades * coste_calculado
            
            return 0
        except:
            return 0
    
    ventas_with_costs['Coste'] = ventas_with_costs.apply(calculate_cost, axis=1)
    
    # Calcular Beneficio
    def calculate_beneficio(row):
        try:
            importe = row['Importe'] if pd.notna(row['Importe']) else 0
            coste = row['Coste'] if pd.notna(row['Coste']) else 0
            iva = obtener_iva_articulo(row['Artículo'])
            beneficio = (importe / (1 + iva / 100)) - coste
            return beneficio
        except:
            return 0
    
    ventas_with_costs['Beneficio'] = ventas_with_costs.apply(calculate_beneficio, axis=1)
    
    # Seleccionar solo las columnas necesarias
    columnas_ventas = ['Vendedor', 'Serie', 'Documento', 'Fecha', 'Factura', 
                       'Artículo', 'Nombre artículo', 'Talla', 'Color', 
                       'Unidades', 'Precio', 'Importe', 'Comisión', 'Tipo registro',
                       'Coste', 'Beneficio']
    
    ventas_df = ventas_with_costs[columnas_ventas].copy()
    
    # Convertir columnas a tipos numéricos correctos
    ventas_df['Unidades'] = pd.to_numeric(ventas_df['Unidades'], errors='coerce').fillna(0)
    ventas_df['Importe'] = pd.to_numeric(ventas_df['Importe'], errors='coerce').fillna(0)
    ventas_df['Coste'] = pd.to_numeric(ventas_df['Coste'], errors='coerce').fillna(0)
    ventas_df['Beneficio'] = pd.to_numeric(ventas_df['Beneficio'], errors='coerce').fillna(0)
    
    # Resumen del procesamiento
    ventas_sin_coste = (ventas_with_costs['Coste'] == 0).sum()
    print(f"VENTAS procesadas: {len(ventas_df)} registros")
    print(f"  - Con coste encontrado: {len(ventas_df) - ventas_sin_coste}")
    print(f"  - Sin coste (asignado 0): {ventas_sin_coste}")
    print(f"\nTotal importe ventas: {ventas_df['Importe'].sum():.2f} €")
    print(f"Total coste ventas: {ventas_df['Coste'].sum():.2f} €")
    print(f"Total beneficio: {ventas_df['Beneficio'].sum():.2f} €")
    
    # Las fechas ya fueron convertidas y filtradas en FASE 1A
    # No es necesario convertirlas nuevamente
    
    # =========================================================================
    # NORMALIZACIÓN DE DATOS
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("FASE 2: NORMALIZACIÓN DE DATOS")
    print("=" * 80)
    
    def normalizar_articulo(df):
        df = df.copy()
        
        def convertir_articulo(valor):
            if pd.isna(valor):
                return ''
            valor_str = str(valor)
            if valor_str.endswith('.0'):
                valor_str = valor_str[:-2]
            return valor_str
        
        df['codigo_str'] = df['Artículo'].apply(convertir_articulo)
        df['nombre_str'] = df['Nombre artículo'].astype(str).str.rstrip()
        df['talla_str'] = df['Talla'].fillna('').astype(str).str.strip()
        df['color_str'] = df['Color'].fillna('').astype(str).str.strip()
        return df
    
    ventas_df = normalizar_articulo(ventas_df)
    compras_df = normalizar_articulo(compras_df)
    stock_df = normalizar_articulo(stock_df)
    
    print("Columnas normalizadas creadas para comparación")
    
    # =========================================================================
    # FILTRAR ARTÍCULOS CON MENOS DE 10 DÍGITOS (REGLA PRIORITARIA)
    # =========================================================================
    
    # Esta regla tiene prioridad sobre todas las demás
    # Los artículos con códigos menores a 10 dígitos no se procesarán
    
    def codigo_valido(codigo):
        """Verifica que el código tenga al menos 10 dígitos"""
        if not codigo or codigo == 'nan':
            return False
        return len(codigo) >= 10
    
    compras_filas_antes = len(compras_df)
    ventas_filas_antes = len(ventas_df)
    stock_filas_antes = len(stock_df)
    
    # Filtrar artículos con códigos menores a 10 dígitos
    compras_df = compras_df[compras_df['codigo_str'].apply(codigo_valido)].copy()
    ventas_df = ventas_df[ventas_df['codigo_str'].apply(codigo_valido)].copy()
    stock_df = stock_df[stock_df['codigo_str'].apply(codigo_valido)].copy()
    
    print(f"\nFiltrados {compras_filas_antes - len(compras_df)} artículos con menos de 10 dígitos en COMPRAS")
    print(f"Filtrados {ventas_filas_antes - len(ventas_df)} artículos con menos de 10 dígitos en VENTAS")
    print(f"Filtrados {stock_filas_antes - len(stock_df)} artículos con menos de 10 dígitos en STOCK")
    
    # =========================================================================
    # FILTRAR FILAS CON UNIDADES = 0
    # =========================================================================
    
    compras_filas_antes = len(compras_df)
    ventas_filas_antes = len(ventas_df)
    stock_filas_antes = len(stock_df)
    
    compras_df = compras_df[compras_df['Unidades'].notna() & (compras_df['Unidades'] > 0)].copy()
    ventas_df = ventas_df[ventas_df['Unidades'].notna() & (ventas_df['Unidades'] > 0)].copy()
    stock_df = stock_df[stock_df['Unidades'].notna() & (stock_df['Unidades'] > 0)].copy()
    
    print(f"\nFiltradas {compras_filas_antes - len(compras_df)} filas con 0 unidades en COMPRAS")
    print(f"Filtradas {ventas_filas_antes - len(ventas_df)} filas con 0 unidades en VENTAS")
    print(f"Filtradas {stock_filas_antes - len(stock_df)} filas con 0 unidades en STOCK")
    
    # =========================================================================
    # PROCESAR SECCIONES
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("FASE 3: PROCESAMIENTO DE SECCIONES")
    print("=" * 80)
    
    # Determinar qué secciones procesar
    if seccion_especifica:
        secciones_a_procesar = [(seccion_especifica, SECCIONES[seccion_especifica])]
    else:
        secciones_a_procesar = list(SECCIONES.items())
    
    # Procesar cada sección
    estadisticas = []
    secciones_procesadas = []
    secciones_sin_datos = []
    
    for nombre_seccion, seccion_info in secciones_a_procesar:
        resultado = procesar_seccion(
            compras_df, ventas_df, stock_df, coste_df,
            nombre_seccion, seccion_info
        )
        
        if resultado:
            estadisticas.append(resultado)
            secciones_procesadas.append(nombre_seccion)
            print(f"\n✓ Sección '{nombre_seccion}' completada: {resultado['archivo']}")
        else:
            secciones_sin_datos.append(nombre_seccion)
    
    # =========================================================================
    # RESUMEN FINAL
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("RESUMEN DEL PROCESAMIENTO")
    print("=" * 80)
    
    # Mostrar información del período y año
    print(f"\nPeríodo procesado: {periodo_seleccionado}")
    print(f"Año de los datos: {año_datos}")
    print(f"Fechas de análisis: {FECHA_INICIO.strftime('%d/%m/%Y')} - {FECHA_FIN.strftime('%d/%m/%Y')} ({DIAS_PERIODO} días)")
    
    print(f"\nSecciones procesadas: {len(secciones_procesadas)}")
    if secciones_procesadas:
        print("  - " + "\n  - ".join(sorted(secciones_procesadas)))
    
    if secciones_sin_datos:
        print(f"\nSecciones sin datos (saltadas): {len(secciones_sin_datos)}")
        print("  - " + "\n  - ".join(sorted(secciones_sin_datos)))
    
    if estadisticas:
        print(f"\nArchivos generados:")
        total_articulos = 0
        for stat in estadisticas:
            print(f"  - {stat['archivo']}: {stat['total_articulos']} artículos "
                  f"(A:{stat['categoria_a']}, B:{stat['categoria_b']}, "
                  f"C:{stat['categoria_c']}, D:{stat['categoria_d']})")
            total_articulos += stat['total_articulos']
        
        print(f"\nTotal artículos en todos los archivos: {total_articulos}")
    
    print("\n" + "=" * 80)
    print("PROCESO COMPLETADO CORRECTAMENTE")
    print("=" * 80)

# ============================================================================
# PUNTO DE ENTRADA
# ============================================================================

if __name__ == "__main__":
    main()
