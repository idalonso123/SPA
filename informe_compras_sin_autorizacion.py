"""
Script para generar informe semanal de compras sin pedido

Este script identifica artículos que han sido comprados pero no estaban
en el pedido semanal de la semana anterior.

NUEVA LÓGICA (Febrero 2026):
1. Cargar el pedido de la semana anterior (artículos autorizados para compra)
2. Cargar el stock de la semana anterior (artículos que existían antes)
3. Cargar el stock actual

Comparación:
- Si el artículo ESTÁ en el pedido_semana_anterior → NO incluido (autorizado)
- Si el artículo NO está en pedido Y SÍ está en stock_semana_anterior → 
  NO incluido (ya existía antes, no es compra reciente)
- Si el artículo NO está en pedido Y NO está en stock_semana_anterior → 
  SÍ incluido (comprado sin autorización en esta semana)

INTEGRACIÓN DE ALERTAS: Este script está integrado con el sistema de alertas.
Los errores y advertencias se enviarán por email automáticamente.

Estructura de datos de entrada:
- SPA_stock_actual.xlsx (stock actual)
- SPA_stock_semana_{fecha}.xlsx (stock semanal guardado, en stocks_semanales/)
- Pedido_Semana_{semana}_{sección}.xlsx (pedidos semanales)

Estructura de datos de salida:
- Excel con 11 hojas (una por sección)
- Archivo de stock semanal guardado para uso futuro

Autor: Sistema de Pedidos VIVEVERDE
Fecha: 2026-02-28
"""

import pandas as pd
import json
import os
import re
from datetime import datetime
from pathlib import Path
from src.paths import INPUT_DIR, OUTPUT_DIR, HISTORICO_COMPRAS_SIN_PEDIDO, COMPRAS_SIN_AUTORIZACION_DIR, PEDIDOS_SEMANALES_DIR
import glob
import warnings
import smtplib
import ssl
import logging
import traceback
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import formatdate

# Ignorar warnings de openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================================
# INTEGRACIÓN DE ALERTAS - IMPORTS Y INICIALIZACIÓN
# ============================================================================

# Importar el módulo de integración de alertas
try:
    from src.integracion_alertas import crear_integrador
    from src.alert_service import crear_alert_service, iniciar_sistema_alertas
    from src.config_loader import cargar_configuracion
    
    # Variable global para el integrador de alertas
    INTEGRADOR_ALERTAS = None
    ALERTAS_DISPONIBLES = True
    logger.info("Módulo de alertas importado correctamente")
except ImportError as e:
    ALERTAS_DISPONIBLES = False
    logger.warning(f"No se pudo importar módulo de alertas: {e}. Continuando sin alertas.")

# Configuración
BASE_PATH = Path(__file__).parent
DATA_INPUT_PATH = INPUT_DIR
# Directorio de salida para compras sin autorización
DATA_OUTPUT_PATH = COMPRAS_SIN_AUTORIZACION_DIR
# Directorio donde están los pedidos semanales
PEDIDOS_DIR = PEDIDOS_SEMANALES_DIR
HISTORY_FILE = HISTORICO_COMPRAS_SIN_PEDIDO

# Secciones del sistema
SECCIONES = [
    "maf",
    "interior", 
    "deco_exterior",
    "deco_interior",
    "fitos",
    "mascotas_manufacturado",
    "mascotas_vivo",
    "semillas",
    "utiles_jardin",
    "vivero",
    "tierras_aridos"
]

# Períodos de stock histórico
PERIODOS = ["P1", "P2", "P3", "P4"]

# ============================================================================
# CONFIGURACIÓN DE EMAIL
# ============================================================================

# Destinatarios: Ivan y Sandra
DESTINATARIOS = [
    {'nombre': 'Ivan', 'email': 'ivan.delgado@viveverde.es'},
    {'nombre': 'Sandra', 'email': 'ivan.delgado@viveverde.es'}
]

# Configuración del servidor SMTP
SMTP_CONFIG = {
    'servidor': 'smtp.serviciodecorreo.es',
    'puerto': 465,
    'remitente_email': 'ivan.delgado@viveverde.es',
    'remitente_nombre': 'Sistema de Pedidos Viveverde'
}


# ============================================================================
# FUNCIONES AUXILIARES PARA EMAIL
# ============================================================================

def obtener_semana_desde_archivos() -> str:
    """
    Extrae el número de semana del archivo de pedido más reciente.
    
    Returns:
        str: Número de semana (ej: '06') o 'desconocida' si no se encuentra
    """
    try:
        # Buscar archivos de pedidos en el directorio de pedidos semanales
        todos_archivos = list(PEDIDOS_SEMANALES_DIR.glob('Pedido_Semana_*'))
        
        if not todos_archivos:
            return 'desconocida'
        
        # Ordenar por fecha y tomar el más reciente
        todos_archivos.sort(key=lambda x: x.name, reverse=True)
        archivo_mas_reciente = todos_archivos[0]
        
        # Extraer la semana del nombre del archivo (formato: Pedido_Semana_XX_...)
        nombre = archivo_mas_reciente.name
        match = re.search(r'Pedido_Semana_(\d+)_', nombre)
        if match:
            return match.group(1)
        
        return 'desconocida'
    except Exception as e:
        print(f"  AVISO: No se pudo extraer la semana: {e}")
        return 'desconocida'


def enviar_email_informe(archivo_informe: str, semana: str = 'desconocida') -> bool:
    """
    Envía un email a Ivan y Sandra con el informe de compras sin autorización adjunto.
    
    Args:
        archivo_informe: Ruta del archivo Excel generado
        semana: Número de semana (ej: '06')
    
    Returns:
        bool: True si el email fue enviado exitosamente, False en caso contrario
    """
    if not archivo_informe:
        print("  AVISO: No hay informe para enviar. No se enviará email.")
        return False
    
    # Verificar que el archivo existe
    if not Path(archivo_informe).exists():
        print(f"  AVISO: El archivo '{archivo_informe}' no existe. No se enviará email.")
        return False
    
    # Verificar contraseña en variable de entorno
    password = os.environ.get('EMAIL_PASSWORD')
    if not password:
        print(f"  AVISO: Variable de entorno 'EMAIL_PASSWORD' no configurada. No se enviará email.")
        return False
    
    # Enviar email a cada destinatario
    emails_enviados = 0
    
    for destinatario in DESTINATARIOS:
        nombre_destinatario = destinatario['nombre']
        email_destinatario = destinatario['email']
        
        try:
            # Crear mensaje MIME
            msg = MIMEMultipart()
            msg['From'] = f"{SMTP_CONFIG['remitente_nombre']} <{SMTP_CONFIG['remitente_email']}>"
            msg['To'] = email_destinatario
            msg['Subject'] = f"Viveverde: Informe Compras Sin Autorización - Semana {semana}"
            msg['Date'] = formatdate(localtime=True)
            
            # Cuerpo del email
            cuerpo = f"""Buenos días {nombre_destinatario},

Te adjunto en este correo el informe de compras sin autorización de la semana {semana}.

Este informe muestra los artículos que han sido comprados pero no estaban en el pedido semanal.

Este informe te permitirá identificar posibles problemas en el proceso de compra.

Atentamente,

Sistema de Pedidos Viveverde."""
            
            msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
            
            # Adjuntar archivo Excel
            filename = Path(archivo_informe).name
            with open(archivo_informe, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename= "{filename}"')
            part.add_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            msg.attach(part)
            
            # Enviar email mediante SSL
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(SMTP_CONFIG['servidor'], SMTP_CONFIG['puerto'], context=context) as server:
                server.login(SMTP_CONFIG['remitente_email'], password)
                server.sendmail(SMTP_CONFIG['remitente_email'], email_destinatario, msg.as_string())
            
            print(f"  Email enviado a {nombre_destinatario} ({email_destinatario})")
            emails_enviados += 1
            
        except smtplib.SMTPException as e:
            print(f"  ERROR SMTP al enviar email a {nombre_destinatario}: {e}")
        except Exception as e:
            print(f"  ERROR al enviar email a {nombre_destinatario}: {e}")
    
    return emails_enviados > 0

# Mapeo de prefijos de códigos de artículo por sección
# Basado en el análisis de los archivos de pedido
SECCION_PREFIX = {
    "maf": ["7"],           # 72, 74, 75, 78, 79
    "interior": ["1"],      # 11, 12, 14, 15, 16
    "deco_exterior": ["9"], # 92, 93, 94, 95, 97
    "deco_interior": ["6"], # 61, 62, 63, 64, 65
    "fitos": ["3"],         # 33, 34, 35, 37, 39
    "mascotas_manufacturado": ["2"],  # 21, 22, 23, 24, 27
    "mascotas_vivo": ["2"],  # Compartido con mascotas_manufacturado
    "semillas": ["5"],      # 53, 54
    "utiles_jardin": ["4"], # 41, 42, 44, 46, 48
    "vivero": ["8"],        # 81, 82, 83, 85, 86
    "tierras_aridos": ["3"] # Compartido con fitos
}


def obtener_prefijo_articulo(articulo):
    """
    Obtiene el prefijo (primer dígito) del código de artículo.
    """
    try:
        articulo_str = str(articulo).replace('.0', '').strip()
        if articulo_str and articulo_str != 'nan':
            return articulo_str[0]
    except:
        pass
    return ''


def es_articulo_de_seccion(articulo, seccion):
    """
    Verifica si un artículo pertenece a una sección basándose en su código.
    """
    prefijo = obtener_prefijo_articulo(articulo)
    return prefijo in SECCION_PREFIX.get(seccion, [])


def cargar_stock_historico():
    """
    Carga todos los archivos de stock histórico (P1-P4) y los combina.
    Retorna un DataFrame con todos los artículos históricos.
    """
    dfs = []
    
    for periodo in PERIODOS:
        archivo = DATA_INPUT_PATH / f"SPA_stock_{periodo}.xlsx"
        if archivo.exists():
            df = pd.read_excel(archivo)
            # Rellenar celdas en blanco hacia abajo para Artículo y Nombre
            df = fill_forward_blank_cells(df, ['Artículo', 'Nombre artículo'])
            df['Periodo'] = periodo
            dfs.append(df)
            print(f"  - Cargado: {archivo.name}")
    
    if dfs:
        return pd.concat(dfs, ignore_index=True)
    return pd.DataFrame()


def cargar_stock_actual():
    """
    Carga el archivo de stock actual.
    """
    archivo = DATA_INPUT_PATH / "SPA_stock_actual.xlsx"
    if archivo.exists():
        df = pd.read_excel(archivo)
        # Rellenar celdas en blanco hacia abajo para Artículo y Nombre
        df = fill_forward_blank_cells(df, ['Artículo', 'Nombre artículo'])
        print(f"  - Cargado: {archivo.name}")
        return df
    return pd.DataFrame()


def cargar_pedido_semana(seccion, semana=None):
    """
    Carga el archivo de pedido semanal para una sección específica.
    Si no se especifica semana, busca el archivo más reciente.
    
    Nota: Los archivos de pedido tienen la primera fila como encabezados.
    """
    if semana is None:
        # Buscar el archivo de pedido más reciente para la sección
        patron = PEDIDOS_DIR / f"Pedido_Semana_*_{seccion}.xlsx"
        archivos = list(patron.glob("*"))
        if not archivos:
            return pd.DataFrame()
        # Ordenar por fecha y tomar el más reciente
        archivos.sort(key=lambda x: x.name, reverse=True)
        archivo = archivos[0]
    else:
        archivo = PEDIDOS_DIR / f"Pedido_Semana_{semana}_{seccion}.xlsx"
    
    if archivo.exists():
        # Los archivos de pedido tienen la primera fila como encabezados
        df = pd.read_excel(archivo, header=1)
        # Rellenar celdas en blanco hacia abajo para Código artículo y Nombre
        df = fill_forward_blank_cells(df, ['Código artículo', 'Nombre Artículo', 'Nombre artículo'])
        print(f"  - Cargado: {archivo.name}")
        return df
    return pd.DataFrame()


def obtener_semana_anterior():
    """
    Obtiene el número de la semana anterior basándose en los archivos de pedido existentes.
    
    Returns:
        str: Número de semana anterior (ej: '06') o None si no se encuentra
    """
    try:
        # Buscar todos los archivos de pedido
        todos_archivos = list(PEDIDOS_SEMANALES_DIR.glob('Pedido_Semana_*'))
        
        if not todos_archivos:
            print("  AVISO: No se encontraron archivos de pedido para determinar la semana anterior")
            return None
        
        # Extraer números de semana de los nombres de archivo
        semanas = set()
        for archivo in todos_archivos:
            match = re.search(r'Pedido_Semana_(\d+)_', archivo.name)
            if match:
                semanas.add(int(match.group(1)))
        
        if not semanas:
            return None
        
        # Obtener la semana máxima (la más reciente)
        semana_mas_reciente = max(semanas)
        
        # La semana anterior es la semana actual - 1
        # Si la semana más reciente es 7, entonces la semana anterior es 6
        semana_anterior = semana_mas_reciente - 1
        
        # Formatear con cero a la izquierda
        return str(semana_anterior).zfill(2)
    
    except Exception as e:
        print(f"  AVISO: Error al obtener la semana anterior: {e}")
        return None


def cargar_pedido_semana_anterior(seccion):
    """
    Carga el archivo de pedido de la semana anterior para una sección específica.
    
    Args:
        seccion: Nombre de la sección (ej: 'maf', 'interior')
    
    Returns:
        DataFrame con los artículos del pedido de la semana anterior
    """
    semana_anterior = obtener_semana_anterior()
    
    if semana_anterior is None:
        print(f"  AVISO: No se pudo determinar la semana anterior para {seccion}")
        return pd.DataFrame()
    
    # Buscar archivo de pedido de la semana anterior
    # El formato del nombre es: Pedido_Semana_{semana}_{fecha}_{seccion}.xlsx
    # Ejemplo: Pedido_Semana_07_26022026_maf.xlsx
    # Buscamos: Pedido_Semana_07_*_maf.xlsx (con _ antes de la sección)
    archivo = PEDIDOS_DIR / f"Pedido_Semana_{semana_anterior}_*_{seccion}.xlsx"
    archivos_encontrados = list(archivo.glob("*"))
    
    if not archivos_encontrados:
        print(f"  AVISO: No se encontró pedido de semana {semana_anterior} para {seccion}")
        return pd.DataFrame()
    
    # Tomar el primero (debería haber solo uno)
    archivo_pedido = archivos_encontrados[0]
    
    try:
        df = pd.read_excel(archivo_pedido, header=1)
        df = fill_forward_blank_cells(df, ['Código artículo', 'Nombre Artículo', 'Nombre artículo'])
        print(f"  - Cargado pedido semana {semana_anterior}: {archivo_pedido.name}")
        return df
    except Exception as e:
        print(f"  ERROR al cargar pedido de semana {semana_anterior}: {e}")
        return pd.DataFrame()


def cargar_stock_semana_anterior():
    """
    Carga el archivo de stock de la semana anterior.
    
    Busca el archivo SPA_stock_semana_anterior.xlsx en el directorio data/input.
    Si no existe, usa el archivo P más antiguo disponible como referencia.
    
    Returns:
        DataFrame con los artículos del stock de la semana anterior
    """
    # Buscar el archivo固定 SPA_stock_semana_anterior.xlsx en data/input
    archivo_stock_semana_anterior = DATA_INPUT_PATH / "SPA_stock_semana_anterior.xlsx"
    
    if archivo_stock_semana_anterior.exists():
        try:
            df = pd.read_excel(archivo_stock_semana_anterior)
            df = fill_forward_blank_cells(df, ['Artículo', 'Nombre artículo'])
            print(f"  - Cargado stock semana anterior: {archivo_stock_semana_anterior.name}")
            return df
        except Exception as e:
            print(f"  AVISO: Error al cargar stock semana anterior: {e}")
    
    # Si no hay stock de semana anterior, usar el archivo P más antiguo disponible como referencia
    # (esto es un fallback - idealmente debería haber un stock de semana anterior guardado)
    print("  AVISO: No se encontró SPA_stock_semana_anterior.xlsx. Usando archivo P más antiguo como referencia.")
    
    for periodo in ["P1", "P2", "P3", "P4"]:
        archivo = DATA_INPUT_PATH / f"SPA_stock_{periodo}.xlsx"
        if archivo.exists():
            try:
                df = pd.read_excel(archivo)
                df = fill_forward_blank_cells(df, ['Artículo', 'Nombre artículo'])
                df['Periodo'] = periodo
                print(f"  - Cargado stock histórico ({periodo}): {archivo.name}")
                return df
            except Exception as e:
                print(f"  AVISO: Error al cargar {archivo.name}: {e}")
                continue
    
    print("  ERROR: No se encontró ningún archivo de stock histórico")
    return pd.DataFrame()


def guardar_stock_semana_actual():
    """
    Guarda una copia del stock actual como SPA_stock_semana_anterior.xlsx en data/input.
    Esto permite tener un registro del stock al final de cada semana para
    compararlo en la próxima ejecución. Se sobrescribe en cada ejecución.
    
    Returns:
        str: Ruta del archivo guardado o None si falló
    """
    try:
        # Cargar stock actual
        stock_actual = cargar_stock_actual()
        
        if stock_actual.empty:
            print("  AVISO: No hay stock actual para guardar")
            return None
        
        # Nombre固定 para el archivo de stock de semana anterior (en data/input)
        nombre_archivo = "SPA_stock_semana_anterior.xlsx"
        archivo_destino = DATA_INPUT_PATH / nombre_archivo
        
        # Guardar archivo (se sobrescribe en cada ejecución)
        stock_actual.to_excel(archivo_destino, index=False)
        print(f"  - Stock semana anterior guardado: {nombre_archivo}")
        
        return str(archivo_destino)
    
    except Exception as e:
        print(f"  ERROR al guardar stock semana anterior: {e}")
        return None


def normalizar_codigo_articulo(codigo):
    """
    Normaliza el código de artículo para comparación.
    Elimina decimales y limpia espacios.
    """
    if pd.isna(codigo):
        return ''
    codigo_str = str(codigo).strip()
    # Eliminar decimales como .0
    if codigo_str.endswith('.0'):
        codigo_str = codigo_str[:-2]
    return codigo_str


def fill_forward_blank_cells(df, columns):
    """
    Rellena celdas en blanco hacia abajo.
    Cuando una celda está vacía, usa el valor de la celda anterior.
    
    Args:
        df: DataFrame
        columns: Lista de columnas a procesar
    
    Returns:
        DataFrame con las celdas rellenadas
    """
    df = df.copy()
    for col in columns:
        if col in df.columns:
            # Convertir a string para manejar NaN correctamente
            df[col] = df[col].astype(str)
            # Reemplazar 'nan' y '' con el valor anterior (forward fill)
            df[col] = df[col].replace(['nan', 'None', ''], pd.NA)
            df[col] = df[col].ffill()
            # Convertir de nuevo a string original
            df[col] = df[col].astype(str).replace('nan', '')
    return df


def normalizar_columnas(df, tipo='stock'):
    """
    Normaliza los nombres de columnas según el tipo de archivo.
    """
    if df.empty:
        return df
    
    if tipo == 'stock':
        # Columnas esperada: Artículo, Nombre artículo, Talla, Color, Unidades
        columnas_normales = {
            'Artículo': 'Artículo',
            'Nombre artículo': 'Nombre Artículo', 
            'Talla': 'Talla',
            'Color': 'Color',
            'Unidades': 'Stock'
        }
    elif tipo == 'pedido':
        # Columnas esperada: Código artículo, Nombre Artículo, Talla, Color
        # Los archivos de pedido pueden tener diferentes nombres de columnas
        columnas_normales = {
            'Código artículo': 'Artículo',
            'Código': 'Artículo', 
            'Nombre Artículo': 'Nombre Artículo',
            'Nombre artículo': 'Nombre Artículo',
            'Talla': 'Talla',
            'Color': 'Color'
        }
    
    # Renombrar columnas
    df = df.rename(columns=columnas_normales)
    
    # Normalizar códigos de artículo
    if 'Artículo' in df.columns:
        df['Artículo'] = df['Artículo'].apply(normalizar_codigo_articulo)
        # Filtrar valores vacíos
        df = df[df['Artículo'] != '']
    
    return df


def crear_clave_articulo(articulo, talla, color):
    """
    Crea una clave única para un artículo combinando código, talla y color.
    """
    articulo_norm = normalizar_codigo_articulo(articulo)
    return f"{articulo_norm}_{talla}_{color}"


def cargar_historico_json():
    """
    Carga el archivo JSON con el histórico de compras sin pedido.
    Estructura: {seccion: {articulo_clave: {stock: float, fecha_actualizacion: str}}}
    
    Nota: Si existe un archivo con estructura antigua (solo fechas), lo migra automáticamente.
    """
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Detectar si es la estructura antigua (sin stock)
        # Estructura antigua: {seccion: {articulo: {fecha_deteccion: date}}}
        # Estructura nueva: {seccion: {articulo: {stock: float, fecha_actualizacion: date}}}
        first_key = list(data.keys())[0] if data else None
        if first_key and first_key in SECCIONES:
            first_item = list(data[first_key].values())[0] if data[first_key] else {}
            if isinstance(first_item, dict) and 'stock' not in first_item:
                print("  - Migrando estructura antigua del JSON a nueva estructura...")
                # Migrar estructura antigua a nueva
                nueva_data = {}
                for seccion, articulos in data.items():
                    nueva_data[seccion] = {}
                    if isinstance(articulos, dict):
                        for articulo, info in articulos.items():
                            if isinstance(info, dict):
                                nueva_data[seccion][articulo] = {
                                    'stock': 0,
                                    'fecha_actualizacion': info.get('fecha_deteccion', datetime.now().strftime('%Y-%m-%d'))
                                }
                            else:
                                # Si el valor no es un dict, inicializar con stock 0
                                nueva_data[seccion][articulo] = {
                                    'stock': 0,
                                    'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d')
                                }
                return nueva_data
        
        return data
    return {}


def guardar_historico_json(datos):
    """
    Guarda el histórico en el archivo JSON.
    """
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(datos, f, indent=2, ensure_ascii=False)
    print(f"  - Historico actualizado: {HISTORY_FILE.name}")


def identificar_compras_sin_pedido(seccion):
    """
    Identifica los artículos comprados sin estar en el pedido para una sección.
    
    NUEVA LÓGICA (según especificación del usuario):
    1. Cargar el pedido de la semana anterior (artículos autorizados para compra)
    2. Cargar el stock de la semana anterior (artículos que existían antes)
    3. Cargar el stock actual
    
    Comparación:
    - Si el artículo ESTÁ en el pedido_semana_anterior → NO incluido (autorizado)
    - Si el artículo NO está en pedido_semana_anterior Y SÍ está en stock_semana_anterior → 
      NO incluido (ya existía antes, no es compra reciente)
    - Si el artículo NO está en pedido_semana_anterior Y NO está en stock_semana_anterior → 
      SÍ incluido (comprado sin autorización en esta semana)
    """
    print(f"\nProcesando sección: {seccion}")
    
    # Cargar datos
    print("  Cargando archivos...")
    stock_actual = cargar_stock_actual()
    pedido_semana_anterior = cargar_pedido_semana_anterior(seccion)
    stock_semana_anterior = cargar_stock_semana_anterior()
    
    if stock_actual.empty:
        print(f"  - No hay stock actual para {seccion}")
        return pd.DataFrame()
    
    # Normalizar columnas
    stock_actual = normalizar_columnas(stock_actual, tipo='stock')
    pedido_semana_anterior = normalizar_columnas(pedido_semana_anterior, tipo='pedido')
    stock_semana_anterior = normalizar_columnas(stock_semana_anterior, tipo='stock')
    
    # ============================================================
    # Paso 1: Obtener artículos autorizados (pedido semana anterior)
    # ============================================================
    if not pedido_semana_anterior.empty and 'Artículo' in pedido_semana_anterior.columns:
        # Crear clave única para pedido
        pedido_semana_anterior['clave'] = pedido_semana_anterior.apply(
            lambda x: crear_clave_articulo(
                str(x.get('Artículo', '')), 
                str(x.get('Talla', '')), 
                str(x.get('Color', ''))
            ), axis=1
        )
        articulos_autorizados = set(pedido_semana_anterior['clave'].dropna())
        print(f"  - Artículos autorizados (pedido semana anterior): {len(articulos_autorizados)}")
    else:
        articulos_autorizados = set()
        print(f"  - No se encontró pedido de semana anterior (0 artículos autorizados)")
    
    # ============================================================
    # Paso 2: Obtener artículos que existían la semana pasada
    # ============================================================
    if not stock_semana_anterior.empty and 'Artículo' in stock_semana_anterior.columns:
        # Crear clave única para stock semana anterior
        stock_semana_anterior['clave'] = stock_semana_anterior.apply(
            lambda x: crear_clave_articulo(
                str(x.get('Artículo', '')), 
                str(x.get('Talla', '')), 
                str(x.get('Color', ''))
            ), axis=1
        )
        articulos_stock_semana_anterior = set(stock_semana_anterior['clave'].dropna())
        print(f"  - Artículos en stock semana anterior: {len(articulos_stock_semana_anterior)}")
    else:
        articulos_stock_semana_anterior = set()
        print(f"  - No se encontró stock de semana anterior (0 artículos)")
    
    # ============================================================
    # Paso 3: Procesar stock actual y aplicar lógica de comparación
    # ============================================================
    if 'Artículo' in stock_actual.columns:
        # Crear clave única para stock actual
        stock_actual['clave'] = stock_actual.apply(
            lambda x: crear_clave_articulo(
                str(x.get('Artículo', '')), 
                str(x.get('Talla', '')), 
                str(x.get('Color', ''))
            ), axis=1
        )
        
        # Filtrar: solo artículos con stock > 0
        stock_con_stock = stock_actual[stock_actual['Stock'].fillna(0) > 0].copy()
        
        # Aplicar la lógica de comparación según el flujo especificado:
        # 1. Si está en pedido_semana_anterior → NO incluir (autorizado)
        # 2. Si NO está en pedido Y SÍ está en stock_semana_anterior → NO incluir (ya existía)
        # 3. Si NO está en pedido Y NO está en stock_semana_anterior → SÍ incluir (compra sin autorización)
        
        resultados = []
        
        for idx, row in stock_con_stock.iterrows():
            clave = row['clave']
            stock_actual_val = float(row.get('Stock', 0))
            
            # Opción 1: ¿El artículo está en el pedido de la semana anterior?
            if clave in articulos_autorizados:
                # Artículo autorizado - NO incluir en el informe
                continue
            
            # Opción 2: ¿El artículo NO está en pedido pero SÍ existía la semana pasada?
            if clave in articulos_stock_semana_anterior:
                # El artículo ya existía antes - NO incluir en el informe
                continue
            
            # Opción 3: El artículo NO está en pedido Y NO estaba en stock semana anterior
            # → ES UNA COMPRA SIN AUTORIZACIÓN - SÍ incluir
            resultados.append({
                'Artículo': row.get('Artículo', ''),
                'Nombre Artículo': row.get('Nombre Artículo', ''),
                'Talla': row.get('Talla', ''),
                'Color': row.get('Color', ''),
                'Stock': stock_actual_val
            })
        
        # Filtrar también por prefijo de sección para asegurar que
        # solojamos artículos que pertenecen a esta sección
        if resultados:
            df_resultados = pd.DataFrame(resultados)
            df_resultados = df_resultados[
                df_resultados['Artículo'].apply(lambda x: es_articulo_de_seccion(x, seccion))
            ]
            
            if not df_resultados.empty:
                print(f"  - Encontrados {len(df_resultados)} artículos comprados sin autorización")
                return df_resultados
            else:
                print(f"  - No hay compras sin autorización para {seccion}")
                return pd.DataFrame()
        else:
            print(f"  - No hay compras sin autorización para {seccion}")
            return pd.DataFrame()
    
    return pd.DataFrame()


def aplicar_estilo_excel(worksheet):
    """
    Aplica formato visual a la hoja de cálculo para que coincida con los pedidos semanales.
    """
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    
    # Fuentes y estilos (color verde oscuro igual que los pedidos semanales)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar formato a encabezados
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Aplicar formato a datos
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column in [1, 2]:  # Artículo y Nombre
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # Talla, Color, Stock
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 15,  # Artículo
        'B': 45,  # Nombre Artículo
        'C': 12,  # Talla
        'D': 12,  # Color
        'E': 10   # Stock
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width


def generar_informe_excel(resultados_por_seccion, nombre_archivo=None):
    """
    Genera el archivo Excel con los resultados de todas las secciones.
    """
    if nombre_archivo is None:
        fecha = datetime.now().strftime('%d%m%Y')
        nombre_archivo = f"Compras_sin_autorizacion_{fecha}.xlsx"
    
    output_path = DATA_OUTPUT_PATH / nombre_archivo
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for seccion in SECCIONES:
            if seccion in resultados_por_seccion and not resultados_por_seccion[seccion].empty:
                df = resultados_por_seccion[seccion]
                df.to_excel(writer, sheet_name=seccion.capitalize(), index=False)
                print(f"  - Hoja '{seccion.capitalize()}' creada con {len(df)} registros")
            else:
                # Crear hoja vacía con encabezados
                df_vacio = pd.DataFrame(columns=['Artículo', 'Nombre Artículo', 'Talla', 'Color', 'Stock'])
                df_vacio.to_excel(writer, sheet_name=seccion.capitalize(), index=False)
    
    # Aplicar formato después de crear el archivo
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        aplicar_estilo_excel(ws)
    
    wb.save(output_path)
    print(f"\nInforme generado: {output_path}")
    return output_path


def main():
    """
    Función principal que orquesta todo el proceso.
    """
    print("=" * 60)
    print("INFORME DE COMPRAS SIN PEDIDO")
    print("=" * 60)
    print(f"Fecha de ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Verificar que existen los archivos necesarios
    print("\nVerificando archivos de entrada...")
    
    stock_actual = DATA_INPUT_PATH / "SPA_stock_actual.xlsx"
    if not stock_actual.exists():
        print(f"ERROR: No se encuentra el archivo: {stock_actual}")
        return
    
    print(f"  - {stock_actual.name} ✓")
    
    # Procesar cada sección
    resultados_por_seccion = {}
    
    for seccion in SECCIONES:
        df_resultado = identificar_compras_sin_pedido(seccion)
        resultados_por_seccion[seccion] = df_resultado
    
    # Generar informe Excel
    print("\n" + "=" * 60)
    print("GENERANDO INFORME EXCEL")
    print("=" * 60)
    
    output_file = generar_informe_excel(resultados_por_seccion)
    
    # Extraer la semana del nombre del archivo de pedido más reciente
    semana = obtener_semana_desde_archivos()
    
    # Enviar email con el informe adjunto
    print("\n" + "=" * 60)
    print("ENVIANDO EMAIL")
    print("=" * 60)
    
    email_enviado = enviar_email_informe(output_file, semana)
    
    # Guardar stock actual para uso en próximas ejecuciones
    print("\n" + "=" * 60)
    print("GUARDANDO STOCK SEMANAL")
    print("=" * 60)
    guardar_stock_semana_actual()
    
    print("\n" + "=" * 60)
    print("PROCESO COMPLETADO")
    print("=" * 60)
    print(f"Archivo de salida: {output_file}")
    print(f"Histórico guardado en: {HISTORY_FILE}")
    if email_enviado:
        print(f"Email enviado a los destinatarios: Ivan y Sandra")


if __name__ == "__main__":
    # ============================================================================
    # INTEGRACIÓN DE ALERTAS - INICIALIZACIÓN Y EJECUCIÓN
    # ============================================================================
    
    alert_service = None
    
    if ALERTAS_DISPONIBLES:
        try:
            # Inicializar el sistema de alertas
            alert_service = crear_integrador("informe_compras_sin_autorizacion")
            if alert_service:
                logger.info("Sistema de alertas inicializado correctamente")
            else:
                logger.warning("No se pudo crear el integrador de alertas")
        except Exception as e:
            logger.error(f"Error al inicializar alertas: {e}")
            alert_service = None
    
    try:
        # Ejecutar el proceso principal
        main()
        logger.info("Proceso de informe de compras sin autorización completado exitosamente.")
    except Exception as e:
        logger.critical(f"Error crítico en el script informe_compras_sin_autorizacion: {e}", exc_info=True)
        if alert_service:
            alert_service.reportar_error("ERROR_EJECUCION", {
                "script": "informe_compras_sin_autorizacion",
                "error": str(e),
                "traceback": traceback.format_exc()
            })
        sys.exit(1)
    finally:
        # Enviar resumen de alertas si el servicio está disponible
        if alert_service:
            try:
                alert_service.enviar_resumen_alertas("informe_compras_sin_autorizacion")
            except Exception as e:
                logger.error(f"Error al enviar resumen de alertas: {e}")
