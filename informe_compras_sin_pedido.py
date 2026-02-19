"""
Script para generar informe semanal de compras sin pedido

Este script identifica artículos que han sido comprados pero no estaban
en el pedido semanal. Compara el stock histórico con el actual y los pedidos.

Estructura de datos de entrada:
- SPA_stock_P1.xlsx, SPA_stock_P2.xlsx, SPA_stock_P3.xlsx, SPA_stock_P4.xlsx
- SPA_stock_actual.xlsx
- Pedido_Semana_{semana}_{sección}.xlsx

Estructura de datos de salida:
- Excel con 11 hojas (una por sección)
- JSON histórico para seguimiento dinámico de stock
"""

import pandas as pd
import json
import os
from datetime import datetime
from pathlib import Path
import glob
import warnings

# Ignorar warnings de openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Configuración
BASE_PATH = Path(__file__).parent
DATA_INPUT_PATH = BASE_PATH / "data" / "input"
DATA_OUTPUT_PATH = BASE_PATH / "data" / "output"
HISTORY_FILE = BASE_PATH / "data" / "compras_sin_pedido_historico.json"

# Secciones del sistema
SECCIONES = [
    "maf",
    "interior", 
    "deco_exterior",
    "deco_interior",
    "fitos",
    "mascotas_manufacturado",
    "mascotas_vivo",
    "semillas",
    "utiles_jardin",
    "vivero",
    "tierras_aridos"
]

# Períodos de stock histórico
PERIODOS = ["P1", "P2", "P3", "P4"]

# Mapeo de prefijos de códigos de artículo por sección
# Basado en el análisis de los archivos de pedido
SECCION_PREFIX = {
    "maf": ["7"],           # 72, 74, 75, 78, 79
    "interior": ["1"],      # 11, 12, 14, 15, 16
    "deco_exterior": ["9"], # 92, 93, 94, 95, 97
    "deco_interior": ["6"], # 61, 62, 63, 64, 65
    "fitos": ["3"],         # 33, 34, 35, 37, 39
    "mascotas_manufacturado": ["2"],  # 21, 22, 23, 24, 27
    "mascotas_vivo": ["2"],  # Compartido con mascotas_manufacturado
    "semillas": ["5"],      # 53, 54
    "utiles_jardin": ["4"], # 41, 42, 44, 46, 48
    "vivero": ["8"],        # 81, 82, 83, 85, 86
    "tierras_aridos": ["3"] # Compartido con fitos
}


def obtener_prefijo_articulo(articulo):
    """
    Obtiene el prefijo (primer dígito) del código de artículo.
    """
    try:
        articulo_str = str(articulo).replace('.0', '').strip()
        if articulo_str and articulo_str != 'nan':
            return articulo_str[0]
    except:
        pass
    return ''


def es_articulo_de_seccion(articulo, seccion):
    """
    Verifica si un artículo pertenece a una sección basándose en su código.
    """
    prefijo = obtener_prefijo_articulo(articulo)
    return prefijo in SECCION_PREFIX.get(seccion, [])


def cargar_stock_historico():
    """
    Carga todos los archivos de stock histórico (P1-P4) y los combina.
    Retorna un DataFrame con todos los artículos históricos.
    """
    dfs = []
    
    for periodo in PERIODOS:
        archivo = DATA_INPUT_PATH / f"SPA_stock_{periodo}.xlsx"
        if archivo.exists():
            df = pd.read_excel(archivo)
            # Rellenar celdas en blanco hacia abajo para Artículo y Nombre
            df = fill_forward_blank_cells(df, ['Artículo', 'Nombre artículo'])
            df['Periodo'] = periodo
            dfs.append(df)
            print(f"  - Cargado: {archivo.name}")
    
    if dfs:
        return pd.concat(dfs, ignore_index=True)
    return pd.DataFrame()


def cargar_stock_actual():
    """
    Carga el archivo de stock actual.
    """
    archivo = DATA_INPUT_PATH / "SPA_stock_actual.xlsx"
    if archivo.exists():
        df = pd.read_excel(archivo)
        # Rellenar celdas en blanco hacia abajo para Artículo y Nombre
        df = fill_forward_blank_cells(df, ['Artículo', 'Nombre artículo'])
        print(f"  - Cargado: {archivo.name}")
        return df
    return pd.DataFrame()


def cargar_pedido_semana(seccion, semana=None):
    """
    Carga el archivo de pedido semanal para una sección específica.
    Si no se especifica semana, busca el archivo más reciente.
    
    Nota: Los archivos de pedido tienen la primera fila como encabezados.
    """
    if semana is None:
        # Buscar el archivo de pedido más reciente para la sección
        patron = DATA_OUTPUT_PATH / f"Pedido_Semana_*_{seccion}.xlsx"
        archivos = list(patron.glob("*"))
        if not archivos:
            return pd.DataFrame()
        # Ordenar por fecha y tomar el más reciente
        archivos.sort(key=lambda x: x.name, reverse=True)
        archivo = archivos[0]
    else:
        archivo = DATA_OUTPUT_PATH / f"Pedido_Semana_{semana}_{seccion}.xlsx"
    
    if archivo.exists():
        # Los archivos de pedido tienen la primera fila como encabezados
        df = pd.read_excel(archivo, header=1)
        # Rellenar celdas en blanco hacia abajo para Código artículo y Nombre
        df = fill_forward_blank_cells(df, ['Código artículo', 'Nombre Artículo', 'Nombre artículo'])
        print(f"  - Cargado: {archivo.name}")
        return df
    return pd.DataFrame()


def normalizar_codigo_articulo(codigo):
    """
    Normaliza el código de artículo para comparación.
    Elimina decimales y limpia espacios.
    """
    if pd.isna(codigo):
        return ''
    codigo_str = str(codigo).strip()
    # Eliminar decimales como .0
    if codigo_str.endswith('.0'):
        codigo_str = codigo_str[:-2]
    return codigo_str


def fill_forward_blank_cells(df, columns):
    """
    Rellena celdas en blanco hacia abajo.
    Cuando una celda está vacía, usa el valor de la celda anterior.
    
    Args:
        df: DataFrame
        columns: Lista de columnas a procesar
    
    Returns:
        DataFrame con las celdas rellenadas
    """
    df = df.copy()
    for col in columns:
        if col in df.columns:
            # Convertir a string para manejar NaN correctamente
            df[col] = df[col].astype(str)
            # Reemplazar 'nan' y '' con el valor anterior (forward fill)
            df[col] = df[col].replace(['nan', 'None', ''], pd.NA)
            df[col] = df[col].ffill()
            # Convertir de nuevo a string original
            df[col] = df[col].astype(str).replace('nan', '')
    return df


def normalizar_columnas(df, tipo='stock'):
    """
    Normaliza los nombres de columnas según el tipo de archivo.
    """
    if df.empty:
        return df
    
    if tipo == 'stock':
        # Columnas esperada: Artículo, Nombre artículo, Talla, Color, Unidades
        columnas_normales = {
            'Artículo': 'Artículo',
            'Nombre artículo': 'Nombre Artículo', 
            'Talla': 'Talla',
            'Color': 'Color',
            'Unidades': 'Stock'
        }
    elif tipo == 'pedido':
        # Columnas esperada: Código artículo, Nombre Artículo, Talla, Color
        # Los archivos de pedido pueden tener diferentes nombres de columnas
        columnas_normales = {
            'Código artículo': 'Artículo',
            'Código': 'Artículo', 
            'Nombre Artículo': 'Nombre Artículo',
            'Nombre artículo': 'Nombre Artículo',
            'Talla': 'Talla',
            'Color': 'Color'
        }
    
    # Renombrar columnas
    df = df.rename(columns=columnas_normales)
    
    # Normalizar códigos de artículo
    if 'Artículo' in df.columns:
        df['Artículo'] = df['Artículo'].apply(normalizar_codigo_articulo)
        # Filtrar valores vacíos
        df = df[df['Artículo'] != '']
    
    return df


def crear_clave_articulo(articulo, talla, color):
    """
    Crea una clave única para un artículo combinando código, talla y color.
    """
    articulo_norm = normalizar_codigo_articulo(articulo)
    return f"{articulo_norm}_{talla}_{color}"


def cargar_historico_json():
    """
    Carga el archivo JSON con el histórico de compras sin pedido.
    Estructura: {seccion: {articulo_clave: {stock: float, fecha_actualizacion: str}}}
    
    Nota: Si existe un archivo con estructura antigua (solo fechas), lo migra automáticamente.
    """
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Detectar si es la estructura antigua (sin stock)
        # Estructura antigua: {seccion: {articulo: {fecha_deteccion: date}}}
        # Estructura nueva: {seccion: {articulo: {stock: float, fecha_actualizacion: date}}}
        first_key = list(data.keys())[0] if data else None
        if first_key and first_key in SECCIONES:
            first_item = list(data[first_key].values())[0] if data[first_key] else {}
            if isinstance(first_item, dict) and 'stock' not in first_item:
                print("  - Migrando estructura antigua del JSON a nueva estructura...")
                # Migrar estructura antigua a nueva
                nueva_data = {}
                for seccion, articulos in data.items():
                    nueva_data[seccion] = {}
                    if isinstance(articulos, dict):
                        for articulo, info in articulos.items():
                            if isinstance(info, dict):
                                nueva_data[seccion][articulo] = {
                                    'stock': 0,
                                    'fecha_actualizacion': info.get('fecha_deteccion', datetime.now().strftime('%Y-%m-%d'))
                                }
                            else:
                                # Si el valor no es un dict, inicializar con stock 0
                                nueva_data[seccion][articulo] = {
                                    'stock': 0,
                                    'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d')
                                }
                return nueva_data
        
        return data
    return {}


def guardar_historico_json(datos):
    """
    Guarda el histórico en el archivo JSON.
    """
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(datos, f, indent=2, ensure_ascii=False)
    print(f"  - Historico actualizado: {HISTORY_FILE.name}")


def identificar_compras_sin_pedido(seccion):
    """
    Identifica los artículos comprados sin estar en el pedido para una sección.
    
    Lógica:
    1. El artículo debe estar en SPA_stock_actual con stock > 0
    2. El artículo NO debe estar en ningún período histórico (P1-P4)
    3. El artículo NO debe estar en el pedido semanal
    4. Si ya estaba en el histórico y el stock no ha cambiado, no aparece
    5. Si el stock aumenta respecto al histórico, sí aparece (nueva compra)
    """
    print(f"\nProcesando sección: {seccion}")
    
    # Cargar datos
    print("  Cargando archivos...")
    stock_historico = cargar_stock_historico()
    stock_actual = cargar_stock_actual()
    pedido = cargar_pedido_semana(seccion)
    
    if stock_actual.empty:
        print(f"  - No hay stock actual para {seccion}")
        return pd.DataFrame()
    
    # Normalizar columnas
    stock_historico = normalizar_columnas(stock_historico, tipo='stock')
    stock_actual = normalizar_columnas(stock_actual, tipo='stock')
    pedido = normalizar_columnas(pedido, tipo='pedido')
    
    # Obtener artículos del pedido (para exclusión)
    if not pedido.empty and 'Artículo' in pedido.columns:
        articulos_pedido = set(pedido['Artículo'].dropna().astype(str).str.strip())
    else:
        articulos_pedido = set()
    
    # Obtener artículos históricos
    if not stock_historico.empty and 'Artículo' in stock_historico.columns:
        # Crear clave única para histórico
        stock_historico['clave'] = stock_historico.apply(
            lambda x: crear_clave_articulo(
                str(x.get('Artículo', '')), 
                str(x.get('Talla', '')), 
                str(x.get('Color', ''))
            ), axis=1
        )
        articulos_historicos = set(stock_historico['clave'].dropna())
    else:
        articulos_historicos = set()
    
    # Obtener artículos en stock actual
    if 'Artículo' in stock_actual.columns:
        stock_actual['clave'] = stock_actual.apply(
            lambda x: crear_clave_articulo(
                str(x.get('Artículo', '')), 
                str(x.get('Talla', '')), 
                str(x.get('Color', ''))
            ), axis=1
        )
        
        # Filtrar: stock > 0 y no está en histórico
        stock_actual_filtrado = stock_actual[
            (stock_actual['Stock'].fillna(0) > 0) &
            (~stock_actual['clave'].isin(articulos_historicos))
        ].copy()
        
        # Filtrar: no está en el pedido
        stock_actual_filtrado = stock_actual_filtrado[
            ~stock_actual_filtrado['Artículo'].isin(articulos_pedido)
        ]
        
        # NUEVO: Filtrar también por prefijo de sección para asegurar que
        # solojamos artículos que pertenecen a esta sección
        stock_actual_filtrado = stock_actual_filtrado[
            stock_actual_filtrado['Artículo'].apply(lambda x: es_articulo_de_seccion(x, seccion))
        ]
        
        # Cargar histórico JSON (estructura por sección)
        historico_completo = cargar_historico_json()
        historico_seccion = historico_completo.get(seccion, {})
        
        # Filtrar según el histórico (dinámico)
        resultados = []
        for idx, row in stock_actual_filtrado.iterrows():
            clave = row['clave']
            stock_actual_val = float(row.get('Stock', 0))
            
            # Buscar en histórico de la sección
            if clave in historico_seccion:
                stock_anterior = float(historico_seccion[clave].get('stock', 0))
                
                # Solo incluir si el stock ha cambiado (aumentado o disminuido)
                if stock_actual_val != stock_anterior:
                    resultados.append({
                        'Artículo': row.get('Artículo', ''),
                        'Nombre Artículo': row.get('Nombre Artículo', ''),
                        'Talla': row.get('Talla', ''),
                        'Color': row.get('Color', ''),
                        'Stock': stock_actual_val
                    })
                    
                    # Actualizar histórico de la sección
                    historico_seccion[clave] = {
                        'stock': stock_actual_val,
                        'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
            else:
                # Es nuevo, incluirlo
                resultados.append({
                    'Artículo': row.get('Artículo', ''),
                    'Nombre Artículo': row.get('Nombre Artículo', ''),
                    'Talla': row.get('Talla', ''),
                    'Color': row.get('Color', ''),
                    'Stock': stock_actual_val
                })
                
                # Agregar al histórico de la sección
                historico_seccion[clave] = {
                    'stock': stock_actual_val,
                    'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
        
        # Actualizar el histórico completo
        historico_completo[seccion] = historico_seccion
        
        # Guardar histórico actualizado
        guardar_historico_json(historico_completo)
        
        if resultados:
            print(f"  - Encontrados {len(resultados)} artículos comprados sin pedido")
            return pd.DataFrame(resultados)
        else:
            print(f"  - No hay compras sin pedido para {seccion}")
            return pd.DataFrame()
    
    return pd.DataFrame()


def aplicar_estilo_excel(worksheet):
    """
    Aplica formato visual a la hoja de cálculo para que coincida con los pedidos semanales.
    """
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    
    # Fuentes y estilos
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar formato a encabezados
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Aplicar formato a datos
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column in [1, 2]:  # Artículo y Nombre
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # Talla, Color, Stock
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 15,  # Artículo
        'B': 45,  # Nombre Artículo
        'C': 12,  # Talla
        'D': 12,  # Color
        'E': 10   # Stock
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width


def generar_informe_excel(resultados_por_seccion, nombre_archivo=None):
    """
    Genera el archivo Excel con los resultados de todas las secciones.
    """
    if nombre_archivo is None:
        fecha = datetime.now().strftime('%d%m%Y')
        nombre_archivo = f"Compras_Sin_Pedido_{fecha}.xlsx"
    
    output_path = DATA_OUTPUT_PATH / nombre_archivo
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for seccion in SECCIONES:
            if seccion in resultados_por_seccion and not resultados_por_seccion[seccion].empty:
                df = resultados_por_seccion[seccion]
                df.to_excel(writer, sheet_name=seccion.capitalize(), index=False)
                print(f"  - Hoja '{seccion.capitalize()}' creada con {len(df)} registros")
            else:
                # Crear hoja vacía con encabezados
                df_vacio = pd.DataFrame(columns=['Artículo', 'Nombre Artículo', 'Talla', 'Color', 'Stock'])
                df_vacio.to_excel(writer, sheet_name=seccion.capitalize(), index=False)
    
    # Aplicar formato después de crear el archivo
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        aplicar_estilo_excel(ws)
    
    wb.save(output_path)
    print(f"\nInforme generado: {output_path}")
    return output_path


def main():
    """
    Función principal que orquesta todo el proceso.
    """
    print("=" * 60)
    print("INFORME DE COMPRAS SIN PEDIDO")
    print("=" * 60)
    print(f"Fecha de ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Verificar que existen los archivos necesarios
    print("\nVerificando archivos de entrada...")
    
    stock_actual = DATA_INPUT_PATH / "SPA_stock_actual.xlsx"
    if not stock_actual.exists():
        print(f"ERROR: No se encuentra el archivo: {stock_actual}")
        return
    
    print(f"  - {stock_actual.name} ✓")
    
    # Procesar cada sección
    resultados_por_seccion = {}
    
    for seccion in SECCIONES:
        df_resultado = identificar_compras_sin_pedido(seccion)
        resultados_por_seccion[seccion] = df_resultado
    
    # Generar informe Excel
    print("\n" + "=" * 60)
    print("GENERANDO INFORME EXCEL")
    print("=" * 60)
    
    output_file = generar_informe_excel(resultados_por_seccion)
    
    print("\n" + "=" * 60)
    print("PROCESO COMPLETADO")
    print("=" * 60)
    print(f"Archivo de salida: {output_file}")
    print(f"Histórico guardado en: {HISTORY_FILE}")


if __name__ == "__main__":
    main()
