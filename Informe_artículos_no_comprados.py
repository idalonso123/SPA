"""
Script para generar informe semanal de artículos NO comprados

Este script identifica artículos que según el pedido semanal deberían haberse comprado
pero NO se encontraron ni en ventas ni en stock.

INTEGRACIÓN DE ALERTAS: Este script está integrado con el sistema de alertas.
Los errores y advertencias se enviarán por email automáticamente.

Estructura de datos de entrada:
- Pedido_Semana_{semana}_{sección}.xlsx
- SPA_ventas_semana.xlsx
- SPA_stock_actual.xlsx

Estructura de datos de salida:
- Excel con 11 hojas (una por sección)

Autor: Sistema de Pedidos VIVEVERDE
Fecha: 2026-02-25
"""

import pandas as pd
import json
import re
from datetime import datetime
from pathlib import Path
from src.paths import INPUT_DIR, OUTPUT_DIR, ARTICULOS_NO_COMPRADOS_DIR, PEDIDOS_SEMANALES_DIR
import glob
import warnings
import smtplib
import ssl
import os
import logging
import traceback
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import formatdate

# Ignorar warnings de openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================================
# INTEGRACIÓN DE ALERTAS - IMPORTS Y INICIALIZACIÓN
# ============================================================================

# Importar el módulo de integración de alertas
try:
    from src.integracion_alertas import crear_integrador
    from src.alert_service import crear_alert_service, iniciar_sistema_alertas
    from src.config_loader import cargar_configuracion
    
    # Variable global para el integrador de alertas
    INTEGRADOR_ALERTAS = None
    ALERTAS_DISPONIBLES = True
    logger.info("Módulo de alertas importado correctamente")
except ImportError as e:
    ALERTAS_DISPONIBLES = False
    logger.warning(f"No se pudo importar módulo de alertas: {e}. Continuando sin alertas.")

# Configuración
BASE_PATH = Path(__file__).parent
DATA_INPUT_PATH = INPUT_DIR
# Usar directorio centralizado para artículos no comprados
DATA_OUTPUT_PATH = ARTICULOS_NO_COMPRADOS_DIR

# Secciones del sistema
SECCIONES = [
    "maf",
    "interior", 
    "deco_exterior",
    "deco_interior",
    "fitos",
    "mascotas_manufacturado",
    "mascotas_vivo",
    "semillas",
    "utiles_jardin",
    "vivero",
    "tierras_aridos"
]

# ============================================================================
# CONFIGURACIÓN DE EMAIL
# ============================================================================

# Destinatarios: Ivan y Sandra
DESTINATARIOS = [
    {'nombre': 'Ivan', 'email': 'ivan.delgado@viveverde.es'},
    {'nombre': 'Sandra', 'email': 'ivan.delgado@viveverde.es'}
]


# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def obtener_semana_desde_archivos() -> str:
    """
    Extrae el número de semana del archivo de pedido más reciente.
    
    Returns:
        str: Número de semana (ej: '06') o 'desconocida' si no se encuentra
    """
    try:
        # Buscar archivos de pedidos en el directorio de pedidos semanales
        todos_archivos = list(PEDIDOS_SEMANALES_DIR.glob('Pedido_Semana_*'))
        
        if not todos_archivos:
            return 'desconocida'
        
        # Ordenar por fecha y tomar el más reciente
        todos_archivos.sort(key=lambda x: x.name, reverse=True)
        archivo_mas_reciente = todos_archivos[0]
        
        # Extraer la semana del nombre del archivo (formato: Pedido_Semana_XX_...)
        nombre = archivo_mas_reciente.name
        # Buscar patrón: Pedido_Semana_XX_ donde XX es el número de semana
        import re
        match = re.search(r'Pedido_Semana_(\d+)_', nombre)
        if match:
            return match.group(1)
        
        return 'desconocida'
    except Exception as e:
        print(f"  AVISO: No se pudo extraer la semana: {e}")
        return 'desconocida'

# Configuración del servidor SMTP
SMTP_CONFIG = {
    'servidor': 'smtp.serviciodecorreo.es',
    'puerto': 465,
    'remitente_email': 'ivan.delgado@viveverde.es',
    'remitente_nombre': 'Sistema de Pedidos Viveverde'
}


def normalizar_codigo_articulo(codigo):
    """
    Normaliza el código de artículo para comparación y visualización.
    Elimina decimales y limpia espacios.
    """
    if pd.isna(codigo):
        return ''
    codigo_str = str(codigo).strip()
    # Eliminar decimales como .0
    if codigo_str.endswith('.0'):
        codigo_str = codigo_str[:-2]
    return codigo_str


def fill_forward_blank_cells(df, columns):
    """
    Rellena celdas en blanco hacia abajo.
    Cuando una celda está vacía, usa el valor de la celda anterior.
    """
    df = df.copy()
    for col in columns:
        if col in df.columns:
            df[col] = df[col].astype(str)
            df[col] = df[col].replace(['nan', 'None', ''], pd.NA)
            df[col] = df[col].ffill()
            df[col] = df[col].astype(str).replace('nan', '')
    return df


def crear_clave_articulo(articulo, talla, color):
    """
    Crea una clave única para un artículo combinando código, talla y color.
    """
    articulo_norm = normalizar_codigo_articulo(articulo)
    return f"{articulo_norm}_{talla}_{color}"


def cargar_pedido_semana(seccion, semana=None):
    """
    Carga el archivo de pedido semanal para una sección específica.
    """
    # Buscar todos los archivos de pedido
    # Buscar archivos de pedidos en el directorio de pedidos semanales
    todos_archivos = list(PEDIDOS_SEMANALES_DIR.glob('Pedido_Semana_*'))
    
    # Filtrar por sección
    archivos = [f for f in todos_archivos if f.name.endswith(f'_{seccion}.xlsx')]
    
    if not archivos:
        return pd.DataFrame()
    
    # Ordenar por fecha y tomar el más reciente
    archivos.sort(key=lambda x: x.name, reverse=True)
    archivo = archivos[0]
    
    if archivo.exists():
        df = pd.read_excel(archivo, header=1)
        # Rellenar celdas en blanco hacia abajo
        df = fill_forward_blank_cells(df, ['Código artículo', 'Nombre Artículo', 'Nombre artículo', 'Talla', 'Color'])
        
        # Eliminar filas de "Métrica de resúmenes" y otras filas de resumen
        if 'Código artículo' in df.columns:
            df = df[df['Código artículo'].notna()]
            df = df[df['Código artículo'] != '']
            # Filtrar filas que contengan texto de métricas/resumen en cualquier columna
            # Incluimos: Métrica, Resumen, Total, Subtotal, Articulos_A, Articulos_B, Articulos_C, Stock_Minimo, Objetivo_Semana, Factor_Crecimiento, Factor_Festivo
            df = df[~df.apply(lambda row: row.astype(str).str.contains(
                'Métrica|Resumen|Total|Subtotal|Articulos_A:|Articulos_B:|Articulos_C:|Stock_Minimo|Objetivo_Semana:|Factor_Crecimiento:|Factor_Festivo:',
                case=False, na=False
            ).any(), axis=1)]
        
        print(f"  - Cargado: {archivo.name}")
        return df
    return pd.DataFrame()


def cargar_ventas_semana():
    """
    Carga el archivo de ventas de la semana.
    """
    archivo = DATA_INPUT_PATH / "SPA_ventas_semana.xlsx"
    if archivo.exists():
        df = pd.read_excel(archivo)
        # Rellenar celdas en blanco
        df = fill_forward_blank_cells(df, ['Artículo', 'Nombre artículo'])
        print(f"  - Cargado: {archivo.name}")
        return df
    return pd.DataFrame()


def cargar_stock_actual():
    """
    Carga el archivo de stock actual.
    """
    archivo = DATA_INPUT_PATH / "SPA_stock_actual.xlsx"
    if archivo.exists():
        df = pd.read_excel(archivo)
        # Rellenar celdas en blanco
        df = fill_forward_blank_cells(df, ['Artículo', 'Nombre artículo'])
        print(f"  - Cargado: {archivo.name}")
        return df
    return pd.DataFrame()


def identificar_articulos_no_comprados(seccion):
    """
    Identifica los artículos que según el pedido deberían haberse comprado
    pero no se encontraron ni en ventas ni en stock.
    
    Lógica:
    1. Leer todos los artículos del pedido semanal
    2. Crear clave única (artículo + talla + color)
    3. Cargar ventas de la semana y stock actual
    4. Verificar si cada artículo del pedido aparece en ventas o stock
    5. Incluir solo los que NO aparecen en ninguno (Opción C)
    """
    print(f"\nProcesando sección: {seccion}")
    
    # Cargar datos
    print("  Cargando archivos...")
    pedido = cargar_pedido_semana(seccion)
    ventas = cargar_ventas_semana()
    stock = cargar_stock_actual()
    
    if pedido.empty:
        print(f"  - No hay pedido para {seccion}")
        return pd.DataFrame()
    
    # Normalizar columna de código de artículo en el pedido
    if 'Código artículo' in pedido.columns:
        pedido['Artículo'] = pedido['Código artículo']
    
    # Normalizar códigos
    if 'Artículo' in pedido.columns:
        pedido['Artículo_norm'] = pedido['Artículo'].apply(normalizar_codigo_articulo)
    else:
        pedido['Artículo_norm'] = ''
    
    # Crear clave para el pedido
    pedido['clave'] = pedido.apply(
        lambda x: crear_clave_articulo(
            x.get('Artículo_norm', ''),
            str(x.get('Talla', '')),
            str(x.get('Color', ''))
        ), axis=1
    )
    
    # Obtener las unidades de compra del pedido (columna "Pedido Final" o similar)
    # Buscar la columna de unidades
    unidades_col = None
    for col in ['Pedido Final', 'Unidades Calculadas', 'Unidades']:
        if col in pedido.columns:
            unidades_col = col
            break
    
    # Obtener conjunto de claves de ventas
    claves_ventas = set()
    if not ventas.empty and 'Artículo' in ventas.columns:
        ventas['clave'] = ventas.apply(
            lambda x: crear_clave_articulo(
                normalizar_codigo_articulo(x.get('Artículo', '')),
                str(x.get('Talla', '')),
                str(x.get('Color', ''))
            ), axis=1
        )
        claves_ventas = set(ventas['clave'].dropna())
    
    # Obtener conjunto de claves de stock
    claves_stock = set()
    if not stock.empty and 'Artículo' in stock.columns:
        stock['clave'] = stock.apply(
            lambda x: crear_clave_articulo(
                normalizar_codigo_articulo(x.get('Artículo', '')),
                str(x.get('Talla', '')),
                str(x.get('Color', ''))
            ), axis=1
        )
        claves_stock = set(stock['clave'].dropna())
    
    # Filtrar: solo artículos que NO están en ventas NI en stock
    resultados = []
    for idx, row in pedido.iterrows():
        clave = row.get('clave', '')
        
        # Solo procesar si la clave no está vacía
        if not clave or clave == '__':
            continue
            
        # Verificar si aparece en ventas o stock
        if clave not in claves_ventas and clave not in claves_stock:
            # Obtener unidades de compra
            unidades = ''
            if unidades_col and unidades_col in row.index:
                unidades = row[unidades_col]
            
            resultados.append({
                'Artículo': normalizar_codigo_articulo(row.get('Artículo', '')),
                'Nombre artículo': row.get('Nombre Artículo', row.get('Nombre artículo', '')),
                'Talla': row.get('Talla', ''),
                'Color': row.get('Color', ''),
                'Unidades compra': unidades
            })
    
    if resultados:
        print(f"  - Encontrados {len(resultados)} artículos NO comprados")
        return pd.DataFrame(resultados)
    else:
        print(f"  - No hay artículos NO comprados para {seccion}")
        return pd.DataFrame()


def aplicar_estilo_excel(worksheet):
    """
    Aplica formato visual a la hoja de cálculo.
    """
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    
    # Fuentes y estilos (color verde RGB 0,128,0 = #008000)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar formato a encabezados
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Aplicar formato a datos
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column in [1, 2]:  # Artículo y Nombre
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # Talla, Color, Stock
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 15,  # Artículo
        'B': 45,  # Nombre Artículo
        'C': 12,  # Talla
        'D': 12,  # Color
        'E': 15   # Unidades compra
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width


def generar_informe_excel(resultados_por_seccion, nombre_archivo=None):
    """
    Genera el archivo Excel con los resultados de todas las secciones.
    """
    if nombre_archivo is None:
        fecha = datetime.now().strftime('%d%m%Y')
        nombre_archivo = f"Articulos_no_comprados_{fecha}.xlsx"
    
    output_path = DATA_OUTPUT_PATH / nombre_archivo
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for seccion in SECCIONES:
            if seccion in resultados_por_seccion and not resultados_por_seccion[seccion].empty:
                df = resultados_por_seccion[seccion]
                df.to_excel(writer, sheet_name=seccion.capitalize(), index=False)
                print(f"  - Hoja '{seccion.capitalize()}' creada con {len(df)} registros")
            else:
                # Crear hoja vacía con encabezados
                df_vacio = pd.DataFrame(columns=['Artículo', 'Nombre artículo', 'Talla', 'Color', 'Unidades compra'])
                df_vacio.to_excel(writer, sheet_name=seccion.capitalize(), index=False)
    
    # Aplicar formato después de crear el archivo
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        aplicar_estilo_excel(ws)
    
    wb.save(output_path)
    print(f"\nInforme generado: {output_path}")
    return output_path


def main():
    """
    Función principal que orquesta todo el proceso.
    """
    print("=" * 60)
    print("INFORME DE ARTÍCULOS NO COMPRADOS")
    print("=" * 60)
    print(f"Fecha de ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Verificar que existen los archivos necesarios
    print("\nVerificando archivos de entrada...")
    
    ventas_file = DATA_INPUT_PATH / "SPA_ventas_semana.xlsx"
    stock_file = DATA_INPUT_PATH / "SPA_stock_actual.xlsx"
    
    if not ventas_file.exists():
        print(f"ERROR: No se encuentra el archivo: {ventas_file}")
        return
    
    if not stock_file.exists():
        print(f"ERROR: No se encuentra el archivo: {stock_file}")
        return
    
    print(f"  - {ventas_file.name} ✓")
    print(f"  - {stock_file.name} ✓")
    
    # Procesar cada sección
    resultados_por_seccion = {}
    
    for seccion in SECCIONES:
        df_resultado = identificar_articulos_no_comprados(seccion)
        resultados_por_seccion[seccion] = df_resultado
    
    # Extraer la semana del nombre del archivo de pedido más reciente
    semana = obtener_semana_desde_archivos()
    
    # Generar informe Excel
    print("\n" + "=" * 60)
    print("GENERANDO INFORME EXCEL")
    print("=" * 60)
    
    output_file = generar_informe_excel(resultados_por_seccion)
    
    # Enviar email con el informe adjunto
    print("\n" + "=" * 60)
    print("ENVIANDO EMAIL")
    print("=" * 60)
    
    email_enviado = enviar_email_informe(output_file, semana)
    
    print("\n" + "=" * 60)
    print("PROCESO COMPLETADO")
    print("=" * 60)
    print(f"Archivo de salida: {output_file}")
    if email_enviado:
        print(f"Email enviado a los destinatarios: Ivan y Sandra")


# ============================================================================
# FUNCIÓN PARA ENVIAR EMAIL CON INFORME ADJUNTO
# ============================================================================

def enviar_email_informe(archivo_informe: str, semana: str = 'desconocida') -> bool:
    """
    Envía un email a Ivan y Sandra con el informe de artículos no comprados adjunto.
    
    Args:
        archivo_informe: Ruta del archivo Excel generado
        semana: Número de semana (ej: '06')
    
    Returns:
        bool: True si el email fue enviado exitosamente, False en caso contrario
    """
    if not archivo_informe:
        print("  AVISO: No hay informe para enviar. No se enviará email.")
        return False
    
    # Verificar que el archivo existe
    if not Path(archivo_informe).exists():
        print(f"  AVISO: El archivo '{archivo_informe}' no existe. No se enviará email.")
        return False
    
    # Verificar contraseña en variable de entorno
    password = os.environ.get('EMAIL_PASSWORD')
    if not password:
        print(f"  AVISO: Variable de entorno 'EMAIL_PASSWORD' no configurada. No se enviará email.")
        return False
    
    # Enviar email a cada destinatario
    emails_enviados = 0
    
    for destinatario in DESTINATARIOS:
        nombre_destinatario = destinatario['nombre']
        email_destinatario = destinatario['email']
        
        try:
            # Crear mensaje MIME
            msg = MIMEMultipart()
            msg['From'] = f"{SMTP_CONFIG['remitente_nombre']} <{SMTP_CONFIG['remitente_email']}>"
            msg['To'] = email_destinatario
            msg['Subject'] = f"Viveverde: Informe Artículos No Comprados - Semana {semana}"
            msg['Date'] = formatdate(localtime=True)
            
            # Cuerpo del email
            cuerpo = f"""Buenos días {nombre_destinatario},

Te adjunto en este correo el informe de artículos no comprados de la semana {semana}. 

Este informe muestra los artículos que aparecen en los pedidos semanales pero que no se han comprado (no aparecen en ventas ni en stock).

Este informe te permitirá identificar posibles problemas en el proceso de compra.

Atentamente,

Sistema de Pedidos Viveverde."""
            
            msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
            
            # Adjuntar archivo Excel
            filename = Path(archivo_informe).name
            with open(archivo_informe, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename= "{filename}"')
            part.add_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            msg.attach(part)
            
            # Enviar email mediante SSL
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(SMTP_CONFIG['servidor'], SMTP_CONFIG['puerto'], context=context) as server:
                server.login(SMTP_CONFIG['remitente_email'], password)
                server.sendmail(SMTP_CONFIG['remitente_email'], email_destinatario, msg.as_string())
            
            print(f"  Email enviado a {nombre_destinatario} ({email_destinatario})")
            emails_enviados += 1
            
        except smtplib.SMTPException as e:
            print(f"  ERROR SMTP al enviar email a {nombre_destinatario}: {e}")
        except Exception as e:
            print(f"  ERROR al enviar email a {nombre_destinatario}: {e}")
    
    return emails_enviados > 0


if __name__ == "__main__":
    # ============================================================================
    # INTEGRACIÓN DE ALERTAS - INICIALIZACIÓN Y EJECUCIÓN
    # ============================================================================
    
    alert_service = None
    
    if ALERTAS_DISPONIBLES:
        try:
            # Inicializar el sistema de alertas
            alert_service = crear_integrador("Informe_articulos_no_comprados")
            if alert_service:
                logger.info("Sistema de alertas inicializado correctamente")
            else:
                logger.warning("No se pudo crear el integrador de alertas")
        except Exception as e:
            logger.error(f"Error al inicializar alertas: {e}")
            alert_service = None
    
    try:
        # Ejecutar el proceso principal
        main()
        logger.info("Proceso de informe de artículos no comprados completado exitosamente.")
    except Exception as e:
        logger.critical(f"Error crítico en el script Informe_artículos_no_comprados: {e}", exc_info=True)
        if alert_service:
            alert_service.reportar_error("ERROR_EJECUCION", {
                "script": "Informe_artículos_no_comprados",
                "error": str(e),
                "traceback": traceback.format_exc()
            })
        sys.exit(1)
    finally:
        # Enviar resumen de alertas si el servicio está disponible
        if alert_service:
            try:
                alert_service.enviar_resumen_alertas("Informe_artículos_no_comprados")
            except Exception as e:
                logger.error(f"Error al enviar resumen de alertas: {e}")
